"""
Octofleet Inventory Backend
FastAPI server for receiving and storing inventory data from Windows Agents
"""
from contextlib import asynccontextmanager
from fastapi import FastAPI, Depends, HTTPException, Header, status, Request, BackgroundTasks, Body, WebSocket, WebSocketDisconnect
from fastapi.middleware.cors import CORSMiddleware
import asyncpg
from typing import Optional, Any, Dict, List
from pydantic import BaseModel, Field
import os
import json
import time
from uuid import UUID
import uuid
import re

def _is_uuid(val: str) -> bool:
    try:
        UUID(val)
        return True
    except (ValueError, AttributeError):
        return False
import secrets
from datetime import datetime, timedelta, timezone

# E7: Alerting imports
from alerting import get_alert_manager, update_node_health, check_node_health

# E19: Provisioning imports
from routers.provisioning import router as provisioning_router, pxe_router
from routers.provisioning_vm import vm_router as provisioning_vm_router
from routers.provisioning_iso import iso_router as provisioning_iso_router
from routers.software_metering import router as software_metering_router
from routers.query_engine import router as query_router
from routers.nodes import router as nodes_router, pending_router
from routers.inventory import router as inventory_router
from routers.jobs import router as jobs_router, agent_router as jobs_agent_router
from routers.mssql import router as mssql_router
from routers.metrics import router as metrics_router
from routers.deployments import router as deployments_router
from routers.alerting import router as alerting_router
from routers.security import router as security_router
from routers.auth import router as auth_router, users_router
from routers.dashboard import router as dashboard_router
from routers.groups import router as groups_router
from routers.content_lifecycle import router as content_lifecycle_router
from app.core.rules import evaluate_dynamic_rule
from app.db.nodes import update_dynamic_device_groupships, auto_onboard_node, upsert_node
from auth import (
    UserCreate, UserUpdate, UserResponse, LoginRequest, TokenResponse,
    RoleCreate, RoleResponse, APIKeyCreate, APIKeyResponse,
    hash_password, verify_password, hash_api_key,
    create_access_token, create_refresh_token, decode_token,
    get_current_user, require_auth, require_permission, CurrentUser,
    get_permissions_for_roles
)
from routers.provisioning_domain import generate_autounattend, generate_djoin_script
from routers.provisioning_postinstall import execute_post_install, handle_provisioning_complete
from routers.remediation import router as remediation_router
from routers.patches import router as patches_router
from routers.baselines import router as baselines_router
from routers.services_mgmt import router as services_mgmt_router
from routers.reports import router as reports_router
from routers.monitoring_mgmt import router as monitoring_mgmt_router
from routers.terminal import router as terminal_router
from routers.packages_mgmt import router as packages_mgmt_router
from routers.vulnerabilities import router as vulnerabilities_router
from routers.hardware import router as hardware_router

# E19: PDF Report imports
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.barcharts import VerticalBarChart
import matplotlib
matplotlib.use('Agg')  # Non-interactive backend
import matplotlib.pyplot as plt

# Centralized dependencies - single source of truth for auth and config
from dependencies import (
    not_found, bad_request, conflict, internal_error,
    API_KEY, DATABASE_URL, GATEWAY_URL, GATEWAY_TOKEN, INVENTORY_API_URL,
    verify_api_key, verify_api_key_or_query, require_scope,
    sanitize_for_postgres, parse_datetime, get_db, set_db_pool,
    db_pool as _deps_db_pool
)

# Local db_pool reference (set during lifespan)
db_pool: Optional[asyncpg.Pool] = None


@asynccontextmanager
async def lifespan(app: FastAPI):
    """Startup and shutdown events"""
    global db_pool
    # Startup
    db_pool = await asyncpg.create_pool(DATABASE_URL, min_size=2, max_size=10)
    set_db_pool(db_pool)  # Set in dependencies for shared access
    
    # Auto-provision default admin user if table is empty
    try:
        async with db_pool.acquire() as conn:
            user_count = await conn.fetchval("SELECT COUNT(*) FROM users")
            if user_count == 0:
                print("⚠️ No users found in database. Provisioning default admin user...")
                # Hash for 'admin' with cost 12
                admin_hash = "$2b$12$rYyY2JO.Uh/NBVp7kXBlY.p9Iv.S0GeIZbgttoYjjyS.hQQQJkkJK"
                await conn.execute("""
                    INSERT INTO users (username, password_hash, display_name, is_superuser, is_active)
                    VALUES ('admin', $1, 'System Administrator', true, true)
                """, admin_hash)
                print("✅ Default admin user created (admin / admin)")
    except Exception as e:
        print(f"❌ Failed to check/provision admin user: {e}")
        
    print(f"✅ Database pool created")
    yield
    # Shutdown
    if db_pool:
        await db_pool.close()
        print("Database pool closed")


# Create app
app = FastAPI(
    title="Octofleet API",
    description="Receives and stores inventory data from Windows Agents",
    version="1.0.0",
    lifespan=lifespan
)

# CORS - configurable via CORS_ORIGINS env var (comma-separated), default: allow all
_cors_env = os.environ.get("CORS_ORIGINS", "*")
CORS_ORIGINS = ["*"] if _cors_env.strip() == "*" else [o.strip() for o in _cors_env.split(",") if o.strip()]

# When allow_origins=["*"], credentials must be False per CORS spec.
# Browsers silently block responses when both are set.
_allow_creds = CORS_ORIGINS != ["*"]

app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS,
    allow_credentials=_allow_creds,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Agent activity tracking middleware
import re as _re
_AGENT_PATH_PATTERN = _re.compile(
    r"/api/v1/(?:jobs/pending|terminal/pending|shell/pending|screen/pending|remediation/jobs/pending|live-data|agents)/(?:([^/]+))"
)
_AGENT_POST_PATTERNS = [
    (_re.compile(r"/api/v1/live-data"), "live-data"),
    (_re.compile(r"/api/v1/agents/([^/]+)/health"), "health-report"),
    (_re.compile(r"/api/v1/remediation/jobs/(\d+)/result"), "remediation-result"),
    (_re.compile(r"/api/v1/jobs/instances/([^/]+)/result"), "job-result"),
]

from collections import deque
import time as _time

_agent_activity_buffer = deque(maxlen=500)

def log_agent_activity(hostname: str, action: str, detail: str = "", ip: str = "", status_code: int = 200):
    _agent_activity_buffer.append({
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "hostname": hostname,
        "action": action,
        "detail": detail,
        "ip": ip,
        "status_code": status_code,
        "ts": _time.time()
    })

@app.middleware("http")
async def agent_activity_middleware(request: Request, call_next):
    response = await call_next(request)
    path = request.url.path
    ip = request.client.host if request.client else ""

    # Track polling endpoints
    if "/pending/" in path:
        m = _re.search(r"/pending/([^/]+)", path)
        if m:
            hostname = m.group(1)
            action = "poll"
            if "terminal" in path: action = "terminal-poll"
            elif "shell" in path: action = "shell-poll"
            elif "screen" in path: action = "screen-poll"
            elif "remediation" in path: action = "remediation-poll"
            elif "jobs" in path: action = "job-poll"
            log_agent_activity(hostname, action, path, ip, response.status_code)

    # Track POST endpoints (health, live-data, results)
    elif request.method == "POST":
        for pattern, action in _AGENT_POST_PATTERNS:
            m = pattern.search(path)
            if m:
                hostname = m.group(1) if m.lastindex else "unknown"
                log_agent_activity(hostname, action, path, ip, response.status_code)
                break

    return response

# Global exception handler to ensure CORS headers on 500 errors
from starlette.responses import JSONResponse
@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    import traceback
    traceback.print_exc()
    return JSONResponse(
        status_code=500,
        content={"detail": "Internal Server Error"},
    )

# Include routers
app.include_router(nodes_router)
app.include_router(pending_router)
app.include_router(inventory_router)
app.include_router(dashboard_router)
app.include_router(groups_router)
app.include_router(content_lifecycle_router, tags=["Content Lifecycle"])
app.include_router(jobs_router)
app.include_router(jobs_agent_router)
app.include_router(mssql_router)
app.include_router(metrics_router)
app.include_router(deployments_router)
app.include_router(alerting_router)
app.include_router(security_router)
app.include_router(auth_router)
app.include_router(users_router)
app.include_router(provisioning_router)
app.include_router(query_router, tags=["Query Engine"])
app.include_router(provisioning_vm_router)
app.include_router(provisioning_iso_router)
app.include_router(pxe_router)
app.include_router(software_metering_router, tags=["Software Metering"])
app.include_router(remediation_router)
app.include_router(patches_router)
app.include_router(baselines_router)
app.include_router(services_mgmt_router)
app.include_router(reports_router)
app.include_router(monitoring_mgmt_router)
app.include_router(terminal_router)
app.include_router(packages_mgmt_router)
app.include_router(vulnerabilities_router)
app.include_router(hardware_router)


# === API Endpoints ===

@app.get("/api/v1/test-sse")
async def test_sse():
    async def generate():
        for i in range(5):
            yield f"event: tick\ndata: {i}\n\n"
            await asyncio.sleep(1)
        yield f"event: done\ndata: finished\n\n"
    return StreamingResponse(generate(), media_type="text/event-stream")


@app.get("/health")
async def health_check():
    """Health check endpoint"""
    try:
        async with db_pool.acquire() as conn:
            await conn.fetchval("SELECT 1")
        return {"status": "ok", "service": "octofleet", "database": "connected"}
    except Exception as e:
        return {"status": "degraded", "service": "octofleet", "database": str(e)}


@app.get("/api/v1/test-sse")
async def test_sse():
    async def generate():
        for i in range(5):
            yield f"event: tick\ndata: {i}\n\n"
            await asyncio.sleep(1)
        yield f"event: done\ndata: finished\n\n"
    return StreamingResponse(generate(), media_type="text/event-stream")


@app.get("/api/v1/health")
async def api_health_check():
    """Health check endpoint (API versioned path)"""
    return await health_check()


@app.get("/api/v1/nodes/{node_id}", dependencies=[Depends(verify_api_key)])
async def get_node_detail(node_id: str, db: asyncpg.Pool = Depends(get_db)):
    """Get detailed info for a single node"""
    async with db.acquire() as conn:
        # Get node basic info
        node = await conn.fetchrow("""
            SELECT id, node_id, hostname, os_name, os_version, os_build, 
                   first_seen, last_seen, is_online, agent_version, created_at, updated_at
            FROM nodes WHERE node_id = $1 OR id::text = $1
        """, node_id)
        
        if not node:
            raise not_found("Node", node_id)
        
        node_uuid = node['id']
        result = dict(node)
        
        # Get hardware summary
        hw = await conn.fetchrow("""
            SELECT cpu->>'name' as cpu_name,
                   (ram->>'totalGb')::numeric as total_memory_gb,
                   updated_at as hardware_updated_at
            FROM hardware_current WHERE node_id = $1
        """, node_uuid)
        if hw:
            result['cpuName'] = hw['cpu_name']
            result['totalMemoryGb'] = float(hw['total_memory_gb']) if hw['total_memory_gb'] else None
            result['hardwareUpdatedAt'] = hw['hardware_updated_at'].isoformat() if hw['hardware_updated_at'] else None
        
        # Get software count
        sw_count = await conn.fetchval(
            "SELECT COUNT(*) FROM software_current WHERE node_id = $1", node_uuid)
        result['softwareCount'] = sw_count
        
        # Get groups
        groups = await conn.fetch("""
            SELECT g.id, g.name, g.color, g.icon
            FROM device_groups dg
            JOIN groups g ON dg.group_id = g.id
            WHERE dg.node_id = $1
        """, node_uuid)
        result['groups'] = [dict(g) for g in groups]
        
        # Get tags
        tags = await conn.fetch("""
            SELECT t.id, t.name, t.color
            FROM device_tags dt
            JOIN tags t ON dt.tag_id = t.id
            WHERE dt.node_id = $1
        """, node_uuid)
        result['tags'] = [dict(t) for t in tags]
        
        return result


@app.get("/api/v1/nodes/{node_id}/history", dependencies=[Depends(verify_api_key)])
async def get_node_history(node_id: str, limit: int = 50, db: asyncpg.Pool = Depends(get_db)):
    """Get change history for a node (hardware snapshots)"""
    async with db.acquire() as conn:
        # Get node UUID
        node = await conn.fetchrow("SELECT id FROM nodes WHERE node_id = $1 OR id::text = $1", node_id)
        if not node:
            raise not_found("Node", node_id)
        
        node_uuid = node['id']
        
        # Get hardware change history
        changes = await conn.fetch("""
            SELECT time as detected_at, change_type, component as category, 
                   old_value, new_value
            FROM hardware_changes 
            WHERE node_id = $1
            ORDER BY time DESC
            LIMIT $2
        """, node_uuid, limit)
        
        result = []
        for i, change in enumerate(changes):
            result.append({
                "id": i + 1,
                "category": change['category'] or "hardware",
                "changeType": change['change_type'] or "snapshot",
                "fieldName": None,
                "oldValue": json.dumps(change['old_value'])[:100] if change['old_value'] else None,
                "newValue": json.dumps(change['new_value'])[:100] if change['new_value'] else None,
                "detectedAt": change['detected_at'].isoformat() if change['detected_at'] else None
            })
        
        return {"changes": result}


@app.post("/api/v1/nodes/register")
async def register_pending_node(request: Request):
    """
    Register a new node as pending approval.
    No authentication required - agents call this on first startup.
    """
    pool = await get_db()
    data = await request.json()
    
    hostname = data.get("hostname", "Unknown")
    os_name = data.get("osName")
    os_version = data.get("osVersion")
    agent_version = data.get("agentVersion")
    machine_id = data.get("machineId")  # Unique hardware identifier
    
    # Get client IP
    ip_address = request.client.host if request.client else None
    
    # Check if already registered (by machine_id or hostname+ip)
    existing = None
    if machine_id:
        existing = await pool.fetchrow(
            "SELECT id, status FROM pending_nodes WHERE machine_id = $1",
            machine_id
        )
    if not existing:
        existing = await pool.fetchrow(
            "SELECT id, status FROM pending_nodes WHERE hostname = $1 AND ip_address = $2",
            hostname, ip_address
        )
    
    if existing:
        # Return existing pending ID so agent can poll
        return {
            "status": existing['status'],
            "pendingId": str(existing['id']),
            "message": f"Node already registered with status: {existing['status']}"
        }
    
    # Create new pending node
    pending_id = await pool.fetchval("""
        INSERT INTO pending_nodes (hostname, os_name, os_version, ip_address, agent_version, machine_id)
        VALUES ($1, $2, $3, $4, $5, $6)
        RETURNING id
    """, hostname, os_name, os_version, ip_address, agent_version, machine_id)
    
    return {
        "status": "pending",
        "pendingId": str(pending_id),
        "message": f"Node {hostname} registered. Awaiting admin approval."
    }

# ============================================

# ============================================
# E3: Job System API
# ============================================

@app.post("/api/v1/jobs", dependencies=[Depends(verify_api_key)])
async def create_job(data: Dict[str, Any], db: asyncpg.Pool = Depends(get_db)):
    """Create a new job targeting devices, groups, or tags"""
    async with db.acquire() as conn:
        job_uuid = uuid.uuid4()
        job_id = str(job_uuid)
        
        # Required fields
        target_type = data.get("targetType", "device")  # device, group, tag, all
        command_type = data.get("commandType", "run")   # run, script, inventory
        command_data = data.get("commandData", {})
        
        # Optional fields
        name = data.get("name", f"Job {job_id[:8]}")
        description = data.get("description", "")
        # Support both targetId and targetDeviceId/targetGroupId
        target_id_input = data.get("targetId") or data.get("targetDeviceId") or data.get("targetGroupId")
        target_tag = data.get("targetTag")
        priority = data.get("priority", 5)
        scheduled_at = data.get("scheduledAt")
        expires_at = data.get("expiresAt")
        created_by = data.get("createdBy", "api")
        timeout_seconds = data.get("timeoutSeconds", 300)
        
        # For device target type, target_id is a text node_id - look up UUID
        target_id = None
        target_node_id = None  # Text node_id for instance creation
        if target_type == "device" and target_id_input:
            # Try to find node by text node_id first
            node = await conn.fetchrow(
                "SELECT id, node_id FROM nodes WHERE node_id = $1 OR id::text = $1", 
                target_id_input
            )
            if node:
                target_id = str(node["id"])
                target_node_id = node["node_id"]
            else:
                # Maybe it's already a UUID?
                try:
                    uuid.UUID(target_id_input)
                    target_id = target_id_input
                    # Look up text node_id
                    node = await conn.fetchrow(
                        "SELECT node_id FROM nodes WHERE id = $1::uuid",
                        target_id_input
                    )
                    if node:
                        target_node_id = node["node_id"]
                except ValueError:
                    raise HTTPException(status_code=404, detail=f"Node not found: {target_id_input}")
        elif target_id_input:
            # For group/tag target types, expect UUID
            target_id = target_id_input
        
        # Insert job
        row = await conn.fetchrow("""
            INSERT INTO jobs (id, name, description, target_type, target_id, target_tag, 
                             command_type, command_data, priority, scheduled_at, expires_at, 
                             created_by, timeout_seconds)
            VALUES ($1::uuid, $2, $3, $4, $5::uuid, $6, $7, $8::jsonb, $9, $10, $11, $12, $13)
            RETURNING id, created_at
        """, str(job_uuid), name, description, target_type, target_id, target_tag,
             command_type, json.dumps(command_data), priority, scheduled_at, expires_at, 
             created_by, timeout_seconds)
        
        # Expand job to instances based on target
        instances_created = 0
        
        if target_type == "device" and target_node_id:
            # target_node_id is already resolved text node_id
            await conn.execute("""
                INSERT INTO job_instances (job_id, node_id, status)
                VALUES ($1::uuid, $2, 'pending')
            """, str(job_uuid), target_node_id)
            instances_created = 1
        
        elif target_type == "group" and target_id:
            # device_groups.node_id is UUID, need to join with nodes table
            nodes = await conn.fetch("""
                SELECT n.node_id FROM device_groups dg
                JOIN nodes n ON n.id = dg.node_id
                WHERE dg.group_id = $1::uuid
            """, target_id)
            for node in nodes:
                await conn.execute("""
                    INSERT INTO job_instances (job_id, node_id, status)
                    VALUES ($1::uuid, $2, 'pending')
                """, str(job_uuid), node["node_id"])
                instances_created += 1
        
        elif target_type == "tag" and target_tag:
            # device_tags.node_id is also UUID
            nodes = await conn.fetch("""
                SELECT n.node_id FROM device_tags dt
                JOIN nodes n ON n.id = dt.node_id
                JOIN tags t ON t.id = dt.tag_id
                WHERE t.name = $1
            """, target_tag)
            for node in nodes:
                await conn.execute("""
                    INSERT INTO job_instances (job_id, node_id, status)
                    VALUES ($1::uuid, $2, 'pending')
                """, str(job_uuid), node["node_id"])
                instances_created += 1
        
        elif target_type == "all":
            # Get node_id (text) from nodes table via system_current
            nodes = await conn.fetch("""
                SELECT n.node_id 
                FROM nodes n 
                INNER JOIN system_current sc ON sc.node_id = n.id
            """)
            for node in nodes:
                await conn.execute("""
                    INSERT INTO job_instances (job_id, node_id, status)
                    VALUES ($1::uuid, $2, 'pending')
                """, str(job_uuid), node["node_id"])
                instances_created += 1
        
        return {
            "id": job_id,
            "name": name,
            "targetType": target_type,
            "commandType": command_type,
            "instancesCreated": instances_created,
            "createdAt": row["created_at"].isoformat() if row["created_at"] else None
        }


# Popular Chocolatey packages for quick reference
CHOCO_PACKAGES = {
    "notepad++": "notepadplusplus",
    "notepadplusplus": "notepadplusplus",
    "7zip": "7zip",
    "7-zip": "7zip",
    "vlc": "vlc",
    "git": "git",
    "vscode": "vscode",
    "vs-code": "vscode",
    "python": "python",
    "nodejs": "nodejs-lts",
    "node": "nodejs-lts",
    "chrome": "googlechrome",
    "firefox": "firefox",
    "putty": "putty",
    "winscp": "winscp",
    "filezilla": "filezilla",
    "sysinternals": "sysinternals",
    "powershell": "powershell-core",
    "pwsh": "powershell-core",
}


class SmartInstallRequest(BaseModel):
    """Request body for smart-install endpoint"""
    package: str = Field(..., description="Package name (e.g., 'notepad++', '7zip', 'git')")
    targets: List[str] = Field(..., description="List of node IDs or hostnames")
    timeout_seconds: int = Field(600, description="Timeout for installation")


@app.post("/api/v1/smart-install")
async def smart_install(
    request: SmartInstallRequest,
    db: asyncpg.Pool = Depends(get_db)
):
    """
    Smart software installation with automatic Chocolatey bootstrap.
    
    Installs software on Windows nodes using Chocolatey. If Chocolatey is not
    installed on the target node, it will be automatically installed first.
    
    Example:
    ```json
    {
        "package": "notepad++",
        "targets": ["CONTROLLER", "DESKTOP-PC1"]
    }
    ```
    """
    # Normalize package name
    package_name = request.package.lower().strip()
    choco_package = CHOCO_PACKAGES.get(package_name, package_name)
    
    # Build smart-install command (installs choco if missing, then package)
    smart_cmd = (
        f'$c="C:\\ProgramData\\chocolatey\\bin\\choco.exe";'
        f'if(!(Test-Path $c))'
        f'{{Set-ExecutionPolicy Bypass -Scope Process -Force;'
        f'[Net.ServicePointManager]::SecurityProtocol='
        f'[Net.ServicePointManager]::SecurityProtocol -bor 3072;'
        f'iex((New-Object Net.WebClient).DownloadString('
        f"'https://community.chocolatey.org/install.ps1'))}}"
        f';& $c install {choco_package} -y --no-progress'
    )
    
    results = []
    async with db.acquire() as conn:
        for target in request.targets:
            # Find node
            node = await conn.fetchrow(
                "SELECT id, node_id, hostname FROM nodes WHERE node_id = $1 OR hostname = $1",
                target
            )
            if not node:
                results.append({
                    "target": target,
                    "status": "error",
                    "error": f"Node not found: {target}"
                })
                continue
            
            # Create job
            job_uuid = uuid.uuid4()
            job_name = f"Smart-Install {request.package}"
            
            await conn.execute("""
                INSERT INTO jobs (id, name, description, target_type, target_id, 
                                 command_type, command_data, timeout_seconds, created_by)
                VALUES ($1::uuid, $2, $3, 'device', $4::uuid, 'run', $5::jsonb, $6, $7)
            """, str(job_uuid), job_name, 
                f"Auto-install {choco_package} via Chocolatey (with bootstrap)",
                str(node["id"]), json.dumps({"command": smart_cmd}),
                request.timeout_seconds, "api")
            
            # Create instance
            await conn.execute("""
                INSERT INTO job_instances (job_id, node_id, status)
                VALUES ($1::uuid, $2, 'pending')
            """, str(job_uuid), node["node_id"])
            
            results.append({
                "target": target,
                "hostname": node["hostname"],
                "status": "created",
                "jobId": str(job_uuid),
                "package": choco_package
            })
    
    return {
        "package": request.package,
        "chocoPackage": choco_package,
        "results": results,
        "totalTargets": len(request.targets),
        "jobsCreated": sum(1 for r in results if r["status"] == "created")
    }


@app.get("/api/v1/smart-install/packages")
async def list_smart_install_packages():
    """List available package aliases for smart-install"""
    return {
        "packages": [
            {"alias": k, "chocoName": v, "description": f"Installs {v} via Chocolatey"}
            for k, v in sorted(CHOCO_PACKAGES.items())
        ]
    }


# ============================================
# E19: MSSQL Deployment Module
# ============================================

from mssql_module import (
    generate_disk_prep_script,
    generate_install_script,
    MssqlInstallRequest,
    DiskConfig,
    DiskConfigSection,
    SqlPaths,
    MSSQL_DOWNLOADS,
    EDITIONS as MSSQL_EDITIONS
)


@app.get("/api/v1/onboarding/config")
async def get_onboarding_config(db: asyncpg.Pool = Depends(get_db)):
    """Get onboarding configuration"""
    async with db.acquire() as conn:
        row = await conn.fetchrow("""
            SELECT oc.*, g.name as group_name
            FROM onboarding_config oc
            LEFT JOIN groups g ON g.id = oc.onboarding_group_id
            LIMIT 1
        """)
        
        if not row:
            return {"configured": False}
        
        return {
            "configured": True,
            "onboardingGroupId": str(row["onboarding_group_id"]) if row["onboarding_group_id"] else None,
            "onboardingGroupName": row["group_name"],
            "autoAssignNewNodes": row["auto_assign_new_nodes"]
        }


@app.put("/api/v1/onboarding/config")
async def update_onboarding_config(data: Dict[str, Any], db: asyncpg.Pool = Depends(get_db)):
    """Update onboarding configuration"""
    async with db.acquire() as conn:
        await conn.execute("""
            INSERT INTO onboarding_config (id, onboarding_group_id, auto_assign_new_nodes, updated_at)
            VALUES (gen_random_uuid(), $1::uuid, $2, now())
            ON CONFLICT (id) DO UPDATE SET
                onboarding_group_id = $1::uuid,
                auto_assign_new_nodes = $2,
                updated_at = now()
        """, data.get("onboardingGroupId"), data.get("autoAssignNewNodes", True))
        
        # Actually update the existing row
        await conn.execute("""
            UPDATE onboarding_config SET
                onboarding_group_id = $1::uuid,
                auto_assign_new_nodes = $2,
                updated_at = now()
        """, data.get("onboardingGroupId"), data.get("autoAssignNewNodes", True))
        
        return {"status": "updated"}


@app.get("/api/v1/jobs", dependencies=[Depends(verify_api_key)])
async def list_jobs(limit: int = 50, offset: int = 0, db: asyncpg.Pool = Depends(get_db)):
    """List all jobs with summary"""
    async with db.acquire() as conn:
        rows = await conn.fetch("""
            SELECT js.job_id, js.name, js.command_type, js.target_type, js.target_id, js.created_at,
                   js.total_instances, js.pending, js.queued, js.running, js.success, js.failed, js.cancelled,
                   n.hostname as target_name
            FROM job_summary js
            LEFT JOIN nodes n ON n.id::text = js.target_id::text
            ORDER BY js.created_at DESC
            LIMIT $1 OFFSET $2
        """, limit, offset)
        
        jobs = [{
            "id": str(row["job_id"]),
            "name": row["name"],
            "commandType": row["command_type"],
            "targetType": row["target_type"],
            "targetName": row["target_name"] or (str(row["target_id"])[:8] + "…" if row["target_id"] else "—"),
            "createdAt": row["created_at"].isoformat() if row["created_at"] else None,
            "summary": {
                "total": row["total_instances"],
                "pending": row["pending"],
                "queued": row["queued"],
                "running": row["running"],
                "success": row["success"],
                "failed": row["failed"],
                "cancelled": row["cancelled"]
            }
        } for row in rows]
        
        return {"jobs": jobs}


@app.get("/api/v1/jobs/{job_id}", dependencies=[Depends(verify_api_key)])
async def get_job(job_id: str, db: asyncpg.Pool = Depends(get_db)):
    """Get job details with all instances"""
    async with db.acquire() as conn:
        job = await conn.fetchrow("""
            SELECT id, name, description, target_type, target_id, target_tag,
                   command_type, command_data, priority, scheduled_at, expires_at,
                   created_by, created_at, timeout_seconds
            FROM jobs WHERE id = $1
        """, job_id)
        
        if not job:
            raise not_found("Job", job_id)
        
        instances = await conn.fetch("""
            SELECT id, node_id, status, queued_at, started_at, completed_at,
                   exit_code, stdout, stderr, error_message, duration_ms, attempt
            FROM job_instances
            WHERE job_id = $1
            ORDER BY queued_at
        """, job_id)
        
        return {
            "id": str(job["id"]),
            "name": job["name"],
            "description": job["description"],
            "targetType": job["target_type"],
            "targetId": str(job["target_id"]) if job["target_id"] else None,
            "targetTag": job["target_tag"],
            "commandType": job["command_type"],
            "commandData": json.loads(job["command_data"]) if job["command_data"] else {},
            "priority": job["priority"],
            "scheduledAt": job["scheduled_at"].isoformat() if job["scheduled_at"] else None,
            "expiresAt": job["expires_at"].isoformat() if job["expires_at"] else None,
            "createdBy": job["created_by"],
            "createdAt": job["created_at"].isoformat() if job["created_at"] else None,
            "timeoutSeconds": job["timeout_seconds"],
            "instances": [{
                "id": str(i["id"]),
                "nodeId": i["node_id"],
                "status": i["status"],
                "queuedAt": i["queued_at"].isoformat() if i["queued_at"] else None,
                "startedAt": i["started_at"].isoformat() if i["started_at"] else None,
                "completedAt": i["completed_at"].isoformat() if i["completed_at"] else None,
                "exitCode": i["exit_code"],
                "stdout": i["stdout"],
                "stderr": i["stderr"],
                "errorMessage": i["error_message"],
                "durationMs": i["duration_ms"],
                "attempt": i["attempt"]
            } for i in instances]
        }


@app.get("/api/v1/jobs/pending/{node_id}")
async def get_pending_jobs(node_id: str, db: asyncpg.Pool = Depends(get_db)):
    """Agent endpoint: Get pending jobs for a specific node"""
    # Support both formats: "win-baltasa" and "BALTASA"
    # Agent uses win-{hostname.lower()}, DB stores HOSTNAME
    lookup_id = node_id
    if node_id.startswith("win-"):
        lookup_id = node_id[4:].upper()  # win-baltasa -> BALTASA
    
    async with db.acquire() as conn:
        # Resolve hostname to UUID if needed
        resolved_id = lookup_id
        node_row = await conn.fetchrow(
            "SELECT id FROM nodes WHERE UPPER(hostname) = UPPER($1) OR UPPER(node_id) = UPPER($1) OR id::text = $1",
            lookup_id
        )
        if node_row:
            resolved_id = str(node_row["id"])
        # Reset stuck 'queued' jobs back to 'pending' (jobs that were picked up but never completed)
        # Uses job timeout or 5 minutes default
        await conn.execute("""
            UPDATE job_instances ji
            SET status = 'pending', updated_at = NOW()
            FROM jobs j
            WHERE ji.job_id = j.id
              AND ji.status = 'queued'
              AND ji.updated_at < NOW() - INTERVAL '1 second' * COALESCE(j.timeout_seconds, 300)
        """)
        
        rows = await conn.fetch("""
            SELECT ji.id, ji.job_id, j.name, j.command_type, j.command_data, j.priority,
                   ji.attempt, ji.max_attempts, j.timeout_seconds
            FROM job_instances ji
            JOIN jobs j ON j.id = ji.job_id
            WHERE (ji.node_id::text = $1 OR UPPER(ji.node_id::text) = UPPER($1))
              AND ji.status = 'pending'
              AND (j.scheduled_at IS NULL OR j.scheduled_at <= NOW())
              AND (j.expires_at IS NULL OR j.expires_at > NOW())
            ORDER BY j.priority ASC, ji.queued_at ASC
            LIMIT 10
        """, resolved_id)
        
        jobs = []
        for row in rows:
            # Mark as queued
            await conn.execute("""
                UPDATE job_instances SET status = 'queued', updated_at = NOW()
                WHERE id = $1
            """, row["id"])
            
            command_data = row["command_data"]
            if isinstance(command_data, str):
                try:
                    command_data = json.loads(command_data)
                except:
                    command_data = {}
            
            command_type = row["command_type"] or "run"
            command_payload = command_data
            
            # Convert install_package to a run command with PowerShell script
            if command_type == "install_package":
                package_id = command_data.get("packageId")
                version_id = command_data.get("versionId")
                
                # Look up version details to get download URL and install command
                if package_id and version_id:
                    version_row = await conn.fetchrow("""
                        SELECT pv.filename, pv.download_url, pv.install_command, pv.sha256_hash,
                               p.name as package_name, p.display_name
                        FROM package_versions pv
                        JOIN packages p ON p.id = pv.package_id
                        WHERE pv.id = $1 AND p.id = $2
                    """, version_id, package_id)
                    
                    if version_row and version_row["download_url"]:
                        download_url = version_row["download_url"]
                        filename = version_row["filename"] or "installer.exe"
                        package_name = version_row["display_name"] or version_row["package_name"]
                        
                        ps_script = f'''
$ErrorActionPreference = "Stop"
$ProgressPreference = "SilentlyContinue"
$installerPath = "$env:TEMP\\{filename}"

Write-Host "=== Installing {package_name} ==="
Write-Host "Downloading from: {download_url}"

try {{
    Invoke-WebRequest -Uri "{download_url}" -OutFile $installerPath -UseBasicParsing
    Write-Host "Download complete: $installerPath"
    
    Write-Host "Running installation..."
    if ($installerPath -like "*.msi") {{
        $msiArgs = @("/i", $installerPath, "/qn", "/norestart")
        Write-Host "msiexec /i $installerPath /qn /norestart"
        $process = Start-Process -FilePath "msiexec.exe" -ArgumentList $msiArgs -Wait -PassThru
    }} else {{
        Write-Host "$installerPath /S"
        $process = Start-Process -FilePath $installerPath -ArgumentList "/S" -Wait -PassThru
    }}
    $exitCode = $process.ExitCode
    
    if ($exitCode -eq 0 -or $exitCode -eq 3010) {{
        Write-Host "Installation successful (exit code: $exitCode)"
    }} else {{
        Write-Host "Installation failed with exit code: $exitCode"
        exit $exitCode
    }}
    
    Remove-Item $installerPath -Force -ErrorAction SilentlyContinue
    Write-Host "=== Done ==="
}} catch {{
    Write-Host "ERROR: $_"
    exit 1
}}
'''
                        # Convert to run command
                        command_type = "run"
                        command_payload = {
                            "command": ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", ps_script.strip()],
                            "timeout": row["timeout_seconds"] or 600
                        }
                    else:
                        # No download URL found
                        command_payload = {
                            "command": ["echo", "ERROR: Package version not found or missing download URL"],
                            "timeout": 30
                        }
                        command_type = "run"

            # Resolve patch catalog IDs to KB IDs for patch_install jobs
            if command_type == "patch_install":
                patches = command_data.get("patches", []) if isinstance(command_data, dict) else []
                if patches:
                    kb_rows = await conn.fetch("""
                        SELECT id::text, kb_id FROM patch_catalog WHERE id::text = ANY($1::text[])
                    """, patches)
                    kb_ids = [r["kb_id"] for r in kb_rows if r["kb_id"]]
                    if isinstance(command_data, dict):
                        command_payload = {
                            **command_data,
                            "kb_ids": kb_ids,
                        }
                    else:
                        command_payload = {"kb_ids": kb_ids, "reboot_policy": "no_reboot"}

            jobs.append({
                # camelCase (new agents)
                "instanceId": str(row["id"]),
                "jobId": str(row["job_id"]),
                "jobName": row["name"] or "Unnamed Job",
                "commandType": command_type,
                "commandPayload": json.dumps(command_payload) if isinstance(command_payload, dict) else str(command_payload),
                "priority": row["priority"],
                "attempt": row["attempt"],
                "maxAttempts": row["max_attempts"],
                "timeoutSeconds": row["timeout_seconds"] or 300,
                # snake_case (legacy Linux agent compatibility)
                "instance_id": str(row["id"]),
                "job_id": str(row["job_id"]),
                "job_name": row["name"] or "Unnamed Job",
                "command_type": command_type,
                "command_payload": json.dumps(command_payload) if isinstance(command_payload, dict) else str(command_payload),
            })
        
        return {"jobs": jobs, "count": len(jobs)}


@app.post("/api/v1/jobs/instances/{instance_id}/start", dependencies=[Depends(verify_api_key)])
async def start_job_instance(instance_id: str, db: asyncpg.Pool = Depends(get_db)):
    """Agent endpoint: Mark job as started"""
    async with db.acquire() as conn:
        row = await conn.fetchrow("""
            UPDATE job_instances 
            SET status = 'running', started_at = NOW(), updated_at = NOW()
            WHERE id = $1
            RETURNING id, job_id, node_id
        """, instance_id)
        
        if not row:
            raise not_found("Job instance", instance_id)
        
        await conn.execute("""
            INSERT INTO job_logs (instance_id, level, message)
            VALUES ($1, 'info', 'Job execution started')
        """, instance_id)
        
        return {"status": "running", "instanceId": instance_id}


@app.post("/api/v1/jobs/instances/{instance_id}/result", dependencies=[Depends(verify_api_key)])
async def submit_job_result(instance_id: str, data: Dict[str, Any], db: asyncpg.Pool = Depends(get_db)):
    """Agent endpoint: Submit job execution result"""
    success = data.get("success", False)
    exit_code = data.get("exitCode", -1)
    stdout = data.get("stdout", "")
    stderr = data.get("stderr", "")
    error_message = data.get("errorMessage", "")
    duration_ms = data.get("durationMs", 0)
    
    status = "success" if success else "failed"
    
    async with db.acquire() as conn:
        row = await conn.fetchrow("""
            UPDATE job_instances 
            SET status = $1, completed_at = NOW(), updated_at = NOW(),
                exit_code = $2, stdout = $3, stderr = $4, error_message = $5, 
                duration_ms = $6
            WHERE id = $7
            RETURNING id, job_id, node_id, attempt, max_attempts
        """, status, exit_code, stdout[:50000] if stdout else None, 
             stderr[:50000] if stderr else None, error_message, duration_ms, instance_id)
        
        if not row:
            raise not_found("Job instance", instance_id)
        
        # Log completion
        await conn.execute("""
            INSERT INTO job_logs (instance_id, level, message)
            VALUES ($1, $2, $3)
        """, instance_id, "info" if success else "error", 
             f"Job completed: exit_code={exit_code}")
        
        # Trigger alert on failure
        if not success:
            # Get job and node info for alert
            job_info = await conn.fetchrow("""
                SELECT j.name as job_name, n.hostname 
                FROM job_instances ji
                JOIN jobs j ON ji.job_id = j.id
                JOIN nodes n ON ji.node_id = n.node_id
                WHERE ji.id = $1
            """, instance_id)
            if job_info:
                try:
                    await trigger_alert('job_failed', {
                        'message': f"Job '{job_info['job_name']}' failed on {job_info['hostname']}",
                        'job_name': job_info['job_name'],
                        'hostname': job_info['hostname'],
                        'exit_code': exit_code,
                        'error': error_message or stderr[:500] if stderr else 'Unknown error'
                    })
                except Exception as e:
                    print(f"Alert trigger error: {e}")
        
        # Check if should retry
        should_retry = False
        if not success and row["attempt"] < row["max_attempts"]:
            should_retry = True
            await conn.execute("""
                UPDATE job_instances 
                SET status = 'pending', attempt = attempt + 1, updated_at = NOW()
                WHERE id = $1
            """, instance_id)
        
        # Handle VM creation job completion - extract MAC and update provisioning task
        if success:
            job_info = await conn.fetchrow("""
                SELECT name FROM jobs WHERE id = $1
            """, row["job_id"])
            
            if job_info and job_info['name'] and job_info['name'].startswith('Create VM:'):
                # This is a VM creation job - parse MAC from output
                try:
                    import re
                    mac_match = re.search(r'"macAddress"\s*:\s*"([0-9A-Fa-f:]{17})"', stdout or '')
                    if mac_match:
                        mac_address = mac_match.group(1).upper()
                        
                        # Find provisioning task linked to this job
                        task = await conn.fetchrow("""
                            SELECT id, hostname FROM provisioning_tasks 
                            WHERE network_config::jsonb->>'vm_job_id' = $1
                        """, row["job_id"])
                        
                        if task:
                            # Update task with MAC address and activate it
                            await conn.execute("""
                                UPDATE provisioning_tasks 
                                SET mac_address = $1, 
                                    status = 'pending',
                                    status_message = 'VM created, waiting for PXE boot',
                                    updated_at = NOW()
                                WHERE id = $2
                            """, mac_address, task['id'])
                            print(f"✅ VM created: {task['hostname']} - MAC: {mac_address}")
                except Exception as e:
                    print(f"Error processing VM creation result: {e}")

        # Handle patch_install job results
        job_detail = await conn.fetchrow(
            "SELECT command_type, command_data FROM jobs WHERE id = $1", row["job_id"])
        if job_detail and job_detail["command_type"] == "patch_install":
            try:
                cmd_data = job_detail["command_data"]
                if isinstance(cmd_data, str):
                    cmd_data = json.loads(cmd_data)
                deployment_id = cmd_data.get("deployment_id")
                patch_ids = cmd_data.get("patches", [])
                node_id = str(row["node_id"])

                # Parse installed/failed KBs from stdout JSON
                installed_kbs = []
                failed_kbs = []
                reboot_required = False
                try:
                    patch_output = json.loads(stdout) if stdout else {}
                    installed_kbs = patch_output.get("installed", [])
                    failed_kbs = patch_output.get("failed", [])
                    reboot_required = patch_output.get("reboot_required", False)
                except:
                    pass

                # Map installed KB IDs back to patch catalog IDs
                if installed_kbs:
                    installed_rows = await conn.fetch(
                        "SELECT id::text, kb_id FROM patch_catalog WHERE kb_id = ANY($1::text[])",
                        installed_kbs)
                    installed_patch_ids = {r["id"] for r in installed_rows}
                else:
                    installed_patch_ids = set()

                for patch_id in patch_ids:
                    p_status = "installed" if patch_id in installed_patch_ids else (
                        "failed" if not success else "installed")

                    # Update deployment results
                    if deployment_id:
                        await conn.execute("""
                            UPDATE patch_deployment_results
                            SET status = $1, installed_at = CASE WHEN $1='installed' THEN NOW() ELSE NULL END
                            WHERE deployment_id = $2 AND node_id = $3 AND patch_id = $4
                        """, p_status, deployment_id, node_id, patch_id)

                    # Update patch_catalog_nodes
                    await conn.execute("""
                        UPDATE patch_catalog_nodes SET status = $1,
                            installed_at = CASE WHEN $1='installed' THEN NOW() ELSE NULL END
                        WHERE patch_id = $2 AND node_id = $3::uuid
                    """, p_status, patch_id, node_id)

                # Check if all results for deployment are done
                if deployment_id:
                    remaining = await conn.fetchval("""
                        SELECT COUNT(*) FROM patch_deployment_results
                        WHERE deployment_id = $1 AND status = 'pending'
                    """, deployment_id)
                    if remaining == 0:
                        await conn.execute(
                            "UPDATE patch_deployments SET status='completed', completed_at=NOW() WHERE id=$1",
                            deployment_id)

                print(f"✅ patch_install result: installed={installed_kbs}, failed={failed_kbs}, reboot={reboot_required}")
            except Exception as e:
                print(f"Error processing patch_install result: {e}")
        
        return {
            "status": status,
            "instanceId": instance_id,
            "willRetry": should_retry
        }


# Legacy endpoint for old Linux agent (snake_case fields, different path)
@app.post("/api/v1/jobs/result", dependencies=[Depends(verify_api_key)])
async def submit_job_result_legacy(data: Dict[str, Any], db: asyncpg.Pool = Depends(get_db)):
    """Legacy agent endpoint: Submit job result (old format with instance_id in body)"""
    instance_id = data.get("instance_id")
    if not instance_id:
        raise HTTPException(status_code=400, detail="instance_id required")
    
    # Support both old (snake_case) and new (camelCase) field names
    exit_code = data.get("exit_code", data.get("exitCode", -1))
    status = data.get("status", "failed")
    stdout = data.get("output", data.get("stdout", ""))
    stderr = data.get("stderr", "")
    
    success = status == "success" or exit_code == 0
    
    async with db.acquire() as conn:
        row = await conn.fetchrow("""
            UPDATE job_instances 
            SET status = $1, completed_at = NOW(), updated_at = NOW(),
                exit_code = $2, stdout = $3, stderr = $4
            WHERE id = $5
            RETURNING id
        """, "success" if success else "failed", exit_code, 
             stdout[:50000] if stdout else None, stderr[:50000] if stderr else None, 
             instance_id)
        
        if not row:
            raise not_found("Job instance", instance_id)
        
        return {"status": "success" if success else "failed", "instanceId": instance_id}


@app.post("/api/v1/jobs/instances/{instance_id}/retry", dependencies=[Depends(verify_api_key)])
async def retry_job_instance(instance_id: str, db: asyncpg.Pool = Depends(get_db)):
    """Manually retry a failed or cancelled job instance"""
    async with db.acquire() as conn:
        row = await conn.fetchrow("""
            SELECT id, status, attempt, max_attempts FROM job_instances WHERE id = $1
        """, instance_id)
        
        if not row:
            raise not_found("Job instance", instance_id)
        
        if row["status"] not in ("failed", "cancelled"):
            raise HTTPException(status_code=400, detail=f"Cannot retry instance with status: {row['status']}")
        
        # Reset to pending with incremented attempt (or reset if at max)
        new_attempt = 1  # Reset attempts for manual retry
        await conn.execute("""
            UPDATE job_instances 
            SET status = 'pending', 
                attempt = $2,
                started_at = NULL,
                completed_at = NULL,
                exit_code = NULL,
                stdout = NULL,
                stderr = NULL,
                error_message = NULL,
                updated_at = NOW()
            WHERE id = $1
        """, instance_id, new_attempt)
        
        await conn.execute("""
            INSERT INTO job_logs (instance_id, level, message)
            VALUES ($1, 'info', 'Job manually retried')
        """, instance_id)
        
        return {"status": "pending", "instanceId": instance_id, "attempt": new_attempt}


@app.delete("/api/v1/jobs/{job_id}", dependencies=[Depends(verify_api_key)])
async def cancel_job(job_id: str, db: asyncpg.Pool = Depends(get_db)):
    """Cancel a job and all pending instances"""
    async with db.acquire() as conn:
        rows = await conn.fetch("""
            UPDATE job_instances 
            SET status = 'cancelled', updated_at = NOW()
            WHERE job_id = $1 AND status IN ('pending', 'queued')
            RETURNING id
        """, job_id)
        
        return {"status": "cancelled", "instancesCancelled": len(rows)}

# PACKAGE MANAGEMENT API (E4)
# ============================================


# ============================================
# PACKAGE VERSIONS API
# ============================================


# ============================================
# DETECTION RULES API
# ============================================


# ============================================
# PACKAGE SOURCES API
# ============================================


# ============================================
# AGENT: PACKAGE DETECTION/DOWNLOAD ENDPOINTS
# ============================================


# ============================================
# EVENTLOG COLLECTION ENDPOINTS
# ============================================

@app.post("/api/v1/nodes/{node_id}/eventlog", dependencies=[Depends(verify_api_key)])
async def push_eventlog(node_id: str, request: Request, db: asyncpg.Pool = Depends(get_db)):
    """Receive eventlog entries from agent"""
    data = await request.json()
    events = data.get("events", [])
    
    if not events:
        return {"status": "ok", "inserted": 0}
    
    async with db.acquire() as conn:
        # Verify node exists
        node = await conn.fetchrow("SELECT node_id FROM nodes WHERE node_id = $1 OR id::text = $1", node_id)
        if not node:
            raise not_found("Node", node_id)
        
        inserted = 0
        for event in events:
            try:
                await conn.execute("""
                    INSERT INTO eventlog_entries 
                    (node_id, log_name, event_id, level, level_name, source, message, event_time, raw_data)
                    VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9)
                """,
                    node_id,
                    sanitize_for_postgres(event.get("logName", "Unknown")),
                    event.get("eventId", 0),
                    event.get("level", 4),
                    sanitize_for_postgres(event.get("levelName")),
                    sanitize_for_postgres(event.get("source")),
                    sanitize_for_postgres(event.get("message", ""))[:4000],  # Limit message size
                    parse_datetime(event.get("timestamp") or event.get("eventTime")) or datetime.utcnow(),
                    json.dumps(sanitize_for_postgres(event.get("rawData"))) if event.get("rawData") else None
                )
                inserted += 1
            except Exception as e:
                print(f"Error inserting event: {e}")
                continue
        
        return {"status": "ok", "inserted": inserted, "total": len(events)}


@app.get("/api/v1/nodes/{node_id}/eventlog", dependencies=[Depends(verify_api_key)])
async def get_node_eventlog(
    node_id: str,
    log_name: Optional[str] = None,
    level: Optional[int] = None,  # Max level (1=Critical only, 2=+Error, etc.)
    event_id: Optional[int] = None,
    hours: int = 24,
    limit: int = 100,
    offset: int = 0,
    db: asyncpg.Pool = Depends(get_db)
):
    """Get eventlog entries for a node with filtering"""
    async with db.acquire() as conn:
        # Build query with filters
        conditions = ["node_id = $1", "collected_at > NOW() - $2 * INTERVAL '1 hour'"]
        params = [node_id, hours]
        param_idx = 3
        
        if log_name:
            conditions.append(f"log_name = ${param_idx}")
            params.append(log_name)
            param_idx += 1
        
        if level:
            conditions.append(f"level <= ${param_idx}")
            params.append(level)
            param_idx += 1
        
        if event_id:
            conditions.append(f"event_id = ${param_idx}")
            params.append(event_id)
            param_idx += 1
        
        where_clause = " AND ".join(conditions)
        
        # Get total count
        count = await conn.fetchval(f"""
            SELECT COUNT(*) FROM eventlog_entries WHERE {where_clause}
        """, *params)
        
        # Get events
        params.extend([limit, offset])
        rows = await conn.fetch(f"""
            SELECT id, log_name, event_id, level, level_name, source, 
                   LEFT(message, 500) as message, event_time, collected_at
            FROM eventlog_entries 
            WHERE {where_clause}
            ORDER BY event_time DESC
            LIMIT ${param_idx} OFFSET ${param_idx + 1}
        """, *params)
        
        events = [dict(r) for r in rows]
        
        # Convert datetimes to ISO strings
        for e in events:
            if e.get("event_time"):
                e["eventTime"] = e.pop("event_time").isoformat()
            if e.get("collected_at"):
                e["collectedAt"] = e.pop("collected_at").isoformat()
            e["eventId"] = e.pop("event_id")
            e["logName"] = e.pop("log_name")
            e["levelName"] = e.pop("level_name")
        
        return {
            "nodeId": node_id,
            "events": events,
            "total": count,
            "limit": limit,
            "offset": offset
        }


@app.get("/api/v1/eventlog/summary")
async def get_eventlog_summary(hours: int = 24, db: asyncpg.Pool = Depends(get_db)):
    """Get eventlog summary across all nodes for dashboard"""
    async with db.acquire() as conn:
        # Summary per node
        rows = await conn.fetch("""
            SELECT 
                e.node_id,
                n.hostname,
                e.log_name,
                COUNT(*) FILTER (WHERE e.level = 1) as critical_count,
                COUNT(*) FILTER (WHERE e.level = 2) as error_count,
                COUNT(*) FILTER (WHERE e.level = 3) as warning_count,
                COUNT(*) as total_count,
                MAX(e.collected_at) as last_collected
            FROM eventlog_entries e
            JOIN nodes n ON n.node_id = e.node_id
            WHERE e.collected_at > NOW() - $1 * INTERVAL '1 hour'
            GROUP BY e.node_id, n.hostname, e.log_name
            ORDER BY critical_count DESC, error_count DESC
        """, hours)
        
        summary = [dict(r) for r in rows]
        for s in summary:
            if s.get("last_collected"):
                s["lastCollected"] = s.pop("last_collected").isoformat()
            s["criticalCount"] = s.pop("critical_count")
            s["errorCount"] = s.pop("error_count")
            s["warningCount"] = s.pop("warning_count")
            s["totalCount"] = s.pop("total_count")
            s["nodeId"] = s.pop("node_id")
            s["logName"] = s.pop("log_name")
        
        # Recent critical/error events
        critical_events = await conn.fetch("""
            SELECT e.id, e.node_id, n.hostname, e.log_name, e.event_id, 
                   e.level, e.level_name, e.source, LEFT(e.message, 200) as message, e.event_time
            FROM eventlog_entries e
            JOIN nodes n ON n.node_id = e.node_id
            WHERE e.level <= 2 AND e.collected_at > NOW() - $1 * INTERVAL '1 hour'
            ORDER BY e.event_time DESC
            LIMIT 20
        """, hours)
        
        recent = []
        for r in critical_events:
            event = dict(r)
            if event.get("event_time"):
                event["eventTime"] = event.pop("event_time").isoformat()
            event["eventId"] = event.pop("event_id")
            event["nodeId"] = event.pop("node_id")
            event["logName"] = event.pop("log_name")
            event["levelName"] = event.pop("level_name")
            recent.append(event)
        
        return {
            "hours": hours,
            "summaryByNode": summary,
            "recentCritical": recent
        }


@app.get("/api/v1/eventlog/important-events")
async def get_important_events(hours: int = 24, limit: int = 50, db: asyncpg.Pool = Depends(get_db)):
    """Get security-relevant events (login failures, user changes, etc.)"""
    # Important Security Event IDs
    important_event_ids = [
        4624,  # Successful login
        4625,  # Failed login
        4634,  # Logoff
        4648,  # Explicit credential logon
        4720,  # User created
        4722,  # User enabled
        4725,  # User disabled
        4726,  # User deleted
        4732,  # Member added to security group
        4733,  # Member removed from security group
        4672,  # Special privileges assigned
        4688,  # Process created
        4697,  # Service installed
        7045,  # Service installed (System log)
        1102,  # Audit log cleared
    ]
    
    async with db.acquire() as conn:
        rows = await conn.fetch("""
            SELECT e.id, e.node_id, n.hostname, e.log_name, e.event_id, 
                   e.level, e.level_name, e.source, LEFT(e.message, 500) as message, e.event_time
            FROM eventlog_entries e
            JOIN nodes n ON n.node_id = e.node_id
            WHERE e.event_id = ANY($1) 
              AND e.collected_at > NOW() - $2 * INTERVAL '1 hour'
            ORDER BY e.event_time DESC
            LIMIT $3
        """, important_event_ids, hours, limit)
        
        events = []
        for r in rows:
            event = dict(r)
            if event.get("event_time"):
                event["eventTime"] = event.pop("event_time").isoformat()
            event["eventId"] = event.pop("event_id")
            event["nodeId"] = event.pop("node_id")
            event["logName"] = event.pop("log_name")
            event["levelName"] = event.pop("level_name")
            events.append(event)
        
        return {
            "hours": hours,
            "events": events,
            "monitoredEventIds": important_event_ids
        }


@app.post("/api/v1/jobs/{job_id}/parse-eventlog", dependencies=[Depends(verify_api_key)])
async def parse_eventlog_from_job(job_id: str, db: asyncpg.Pool = Depends(get_db)):
    """Parse eventlog data from a completed job and store in eventlog_entries"""
    async with db.acquire() as conn:
        # Get job instances with stdout
        instances = await conn.fetch("""
            SELECT id, node_id, stdout, status
            FROM job_instances 
            WHERE job_id = $1 AND status = 'success' AND stdout IS NOT NULL
        """, job_id)
        
        if not instances:
            return {"status": "no_data", "message": "No successful instances with output found"}
        
        total_inserted = 0
        results = []
        
        for inst in instances:
            node_id = inst["node_id"]
            stdout = inst["stdout"]
            
            # Extract JSON from stdout (skip "=== MAIN COMMAND ===" prefix)
            json_start = stdout.find('[')
            if json_start == -1:
                results.append({"nodeId": node_id, "status": "no_json", "inserted": 0})
                continue
            
            json_data = stdout[json_start:]
            
            try:
                events = json.loads(json_data)
                if not isinstance(events, list):
                    events = [events]
                
                inserted = 0
                for event in events:
                    try:
                        await conn.execute("""
                            INSERT INTO eventlog_entries 
                            (node_id, log_name, event_id, level, level_name, source, message, event_time)
                            VALUES ($1, $2, $3, $4, $5, $6, $7, $8)
                            ON CONFLICT DO NOTHING
                        """,
                            node_id,
                            sanitize_for_postgres(event.get("logName", "Unknown")),
                            event.get("eventId", 0),
                            event.get("level", 4),
                            sanitize_for_postgres(event.get("levelName")),
                            sanitize_for_postgres(event.get("source")),
                            sanitize_for_postgres(event.get("message", ""))[:4000],
                            parse_datetime(event.get("eventTime")) or datetime.utcnow()
                        )
                        inserted += 1
                    except Exception as e:
                        print(f"Error inserting event for {node_id}: {e}")
                        continue
                
                total_inserted += inserted
                results.append({"nodeId": node_id, "status": "ok", "inserted": inserted, "total": len(events)})
                
            except json.JSONDecodeError as e:
                results.append({"nodeId": node_id, "status": "json_error", "error": str(e)})
        
        return {
            "jobId": job_id,
            "totalInserted": total_inserted,
            "results": results
        }


# ============== METRICS API ==============

@app.post("/api/v1/nodes/{node_id}/metrics", dependencies=[Depends(verify_api_key)])
async def push_metrics(node_id: str, data: Dict[str, Any], db: asyncpg.Pool = Depends(get_db)):
    """Receive metrics from a node"""
    async with db.acquire() as conn:
        # Get node UUID (case-insensitive)
        node = await conn.fetchrow("SELECT id FROM nodes WHERE UPPER(node_id) = UPPER($1)", node_id)
        if not node:
            raise not_found("Node", node_id)
        
        node_uuid = node["id"]
        
        # Insert metrics
        await conn.execute("""
            INSERT INTO node_metrics (time, node_id, cpu_percent, ram_percent, disk_percent, network_in_mb, network_out_mb)
            VALUES (NOW(), $1, $2, $3, $4, $5, $6)
        """, node_uuid, 
            data.get("cpuPercent"),
            data.get("ramPercent"),
            data.get("diskPercent"),
            data.get("networkInMb"),
            data.get("networkOutMb")
        )
        
        # Update last_seen timestamp - metrics prove the node is alive
        await conn.execute("UPDATE nodes SET last_seen = NOW() WHERE id = $1", node_uuid)
        
        return {"status": "ok"}


@app.get("/api/v1/nodes/{node_id}/metrics", dependencies=[Depends(verify_api_key)])
async def get_node_metrics(node_id: str, hours: int = 24, db: asyncpg.Pool = Depends(get_db)):
    """Get metrics for a node with time series data"""
    async with db.acquire() as conn:
        node = await conn.fetchrow("SELECT id FROM nodes WHERE node_id = $1 OR id::text = $1", node_id)
        if not node:
            raise not_found("Node", node_id)
        
        node_uuid = node["id"]
        
        # Get raw metrics
        rows = await conn.fetch("""
            SELECT time, cpu_percent, ram_percent, disk_percent, network_in_mb, network_out_mb
            FROM node_metrics
            WHERE node_id = $1 AND time > NOW() - INTERVAL '%s hours'
            ORDER BY time ASC
        """ % hours, node_uuid)
        
        metrics = [{
            "time": row["time"].isoformat(),
            "cpuPercent": row["cpu_percent"],
            "ramPercent": row["ram_percent"],
            "diskPercent": row["disk_percent"],
            "networkInMb": row["network_in_mb"],
            "networkOutMb": row["network_out_mb"]
        } for row in rows]
        
        # Calculate averages
        if metrics:
            avg_cpu = sum(m["cpuPercent"] or 0 for m in metrics) / len(metrics)
            avg_ram = sum(m["ramPercent"] or 0 for m in metrics) / len(metrics)
            avg_disk = sum(m["diskPercent"] or 0 for m in metrics) / len(metrics)
        else:
            avg_cpu = avg_ram = avg_disk = None
        
        return {
            "nodeId": node_id,
            "hours": hours,
            "dataPoints": len(metrics),
            "averages": {
                "cpuPercent": round(avg_cpu, 1) if avg_cpu else None,
                "ramPercent": round(avg_ram, 1) if avg_ram else None,
                "diskPercent": round(avg_disk, 1) if avg_disk else None
            },
            "metrics": metrics
        }


@app.get("/api/v1/metrics/summary", dependencies=[Depends(verify_api_key)])
async def get_metrics_summary(db: asyncpg.Pool = Depends(get_db)):
    """Get metrics summary for all nodes (for dashboard)"""
    async with db.acquire() as conn:
        # Get latest metrics per node
        rows = await conn.fetch("""
            WITH latest AS (
                SELECT DISTINCT ON (node_id) 
                    node_id, time, cpu_percent, ram_percent, disk_percent
                FROM node_metrics
                WHERE time > NOW() - INTERVAL '1 hour'
                ORDER BY node_id, time DESC
            )
            SELECT n.node_id as text_node_id, n.hostname, 
                   l.time, l.cpu_percent, l.ram_percent, l.disk_percent
            FROM nodes n
            LEFT JOIN latest l ON l.node_id = n.id
            ORDER BY n.hostname
        """)
        
        nodes = [{
            "nodeId": row["text_node_id"],
            "hostname": row["hostname"],
            "lastMetricTime": row["time"].isoformat() if row["time"] else None,
            "cpuPercent": row["cpu_percent"],
            "ramPercent": row["ram_percent"],
            "diskPercent": row["disk_percent"]
        } for row in rows]
        
        # Calculate fleet averages
        active_nodes = [n for n in nodes if n["cpuPercent"] is not None]
        if active_nodes:
            fleet_avg = {
                "cpuPercent": round(sum(n["cpuPercent"] for n in active_nodes) / len(active_nodes), 1),
                "ramPercent": round(sum(n["ramPercent"] for n in active_nodes) / len(active_nodes), 1),
                "diskPercent": round(sum(n["diskPercent"] for n in active_nodes) / len(active_nodes), 1)
            }
        else:
            fleet_avg = {"cpuPercent": None, "ramPercent": None, "diskPercent": None}
        
        return {
            "nodesWithMetrics": len(active_nodes),
            "totalNodes": len(nodes),
            "fleetAverages": fleet_avg,
            "nodes": nodes
        }


@app.get("/api/v1/metrics/fleet", dependencies=[Depends(verify_api_key)])
async def get_fleet_performance(hours: int = 1, db: asyncpg.Pool = Depends(get_db)):
    """
    Get detailed fleet performance data for the performance overview table.
    Returns avg/max/min for CPU, RAM, Disk, Network for each node.
    """
    async with db.acquire() as conn:
        # Get aggregated metrics per node for the last N hours
        rows = await conn.fetch("""
            SELECT 
                n.id,
                n.node_id as text_node_id, 
                n.hostname,
                n.os_name,
                n.last_seen,
                n.is_online,
                -- CPU stats
                ROUND(AVG(m.cpu_percent)::numeric, 1) as avg_cpu,
                ROUND(MAX(m.cpu_percent)::numeric, 1) as max_cpu,
                ROUND(MIN(m.cpu_percent)::numeric, 1) as min_cpu,
                -- RAM stats
                ROUND(AVG(m.ram_percent)::numeric, 1) as avg_ram,
                ROUND(MAX(m.ram_percent)::numeric, 1) as max_ram,
                -- Disk stats
                ROUND(AVG(m.disk_percent)::numeric, 1) as avg_disk,
                ROUND(MAX(m.disk_percent)::numeric, 1) as max_disk,
                -- Network stats
                ROUND(AVG(m.network_in_mb)::numeric, 2) as avg_net_in,
                ROUND(AVG(m.network_out_mb)::numeric, 2) as avg_net_out,
                ROUND(MAX(m.network_in_mb)::numeric, 2) as max_net_in,
                ROUND(MAX(m.network_out_mb)::numeric, 2) as max_net_out,
                -- Data point count
                COUNT(m.*)::int as data_points
            FROM nodes n
            LEFT JOIN node_metrics m ON m.node_id = n.id 
                AND m.time > NOW() - INTERVAL '1 hour' * $1
            GROUP BY n.id, n.node_id, n.hostname, n.os_name, n.last_seen, n.is_online
            ORDER BY n.hostname
        """, hours)
        
        nodes = []
        for row in rows:
            nodes.append({
                "id": str(row["id"]),
                "nodeId": row["text_node_id"],
                "hostname": row["hostname"],
                "osName": row["os_name"],
                "lastSeen": row["last_seen"].isoformat() if row["last_seen"] else None,
                "isOnline": row["is_online"],
                "dataPoints": row["data_points"],
                "cpu": {
                    "avg": float(row["avg_cpu"]) if row["avg_cpu"] else None,
                    "max": float(row["max_cpu"]) if row["max_cpu"] else None,
                    "min": float(row["min_cpu"]) if row["min_cpu"] else None,
                },
                "ram": {
                    "avg": float(row["avg_ram"]) if row["avg_ram"] else None,
                    "max": float(row["max_ram"]) if row["max_ram"] else None,
                },
                "disk": {
                    "avg": float(row["avg_disk"]) if row["avg_disk"] else None,
                    "max": float(row["max_disk"]) if row["max_disk"] else None,
                },
                "network": {
                    "avgIn": float(row["avg_net_in"]) if row["avg_net_in"] else None,
                    "avgOut": float(row["avg_net_out"]) if row["avg_net_out"] else None,
                    "maxIn": float(row["max_net_in"]) if row["max_net_in"] else None,
                    "maxOut": float(row["max_net_out"]) if row["max_net_out"] else None,
                },
            })
        
        # Calculate fleet totals
        active = [n for n in nodes if n["cpu"]["avg"] is not None]
        fleet = {
            "avgCpu": round(sum(n["cpu"]["avg"] for n in active) / len(active), 1) if active else None,
            "avgRam": round(sum(n["ram"]["avg"] for n in active) / len(active), 1) if active else None,
            "avgDisk": round(sum(n["disk"]["avg"] for n in active) / len(active), 1) if active else None,
        }
        
        return {
            "timestamp": datetime.utcnow().isoformat(),
            "hoursAggregated": hours,
            "totalNodes": len(nodes),
            "nodesWithMetrics": len(active),
            "fleet": fleet,
            "nodes": nodes
        }


@app.get("/api/v1/metrics/timeseries", dependencies=[Depends(verify_api_key)])
async def get_fleet_timeseries(
    hours: int = 1, 
    bucket_minutes: int = 5,
    group_id: Optional[str] = None,
    db: asyncpg.Pool = Depends(get_db)
):
    """
    Get fleet-wide time series data for sparkline charts.
    Aggregates all nodes (or filtered by group) into time buckets.
    """
    from datetime import timedelta
    
    async with db.acquire() as conn:
        # Calculate start time as datetime
        start_time = datetime.utcnow() - timedelta(hours=hours)
        bucket_delta = timedelta(minutes=bucket_minutes)
        
        # Build group filter
        group_filter = ""
        params = [start_time, bucket_delta]
        if group_id:
            group_filter = "AND n.id IN (SELECT node_id FROM group_members WHERE group_id = $3::uuid)"
            params.append(group_id)
        
        # Use TimescaleDB time_bucket for efficient aggregation
        query = f"""
            SELECT 
                time_bucket($2, m.time) as bucket,
                ROUND(AVG(m.cpu_percent)::numeric, 1) as avg_cpu,
                ROUND(AVG(m.ram_percent)::numeric, 1) as avg_ram,
                ROUND(AVG(m.disk_percent)::numeric, 1) as avg_disk,
                COUNT(DISTINCT m.node_id) as node_count
            FROM node_metrics m
            JOIN nodes n ON n.id = m.node_id
            WHERE m.time > $1
            {group_filter}
            GROUP BY bucket
            ORDER BY bucket ASC
        """
        
        rows = await conn.fetch(query, *params)
        
        timeseries = [{
            "time": row["bucket"].isoformat(),
            "cpu": float(row["avg_cpu"]) if row["avg_cpu"] else None,
            "ram": float(row["avg_ram"]) if row["avg_ram"] else None,
            "disk": float(row["avg_disk"]) if row["avg_disk"] else None,
            "nodes": row["node_count"]
        } for row in rows]
        
        # Get current averages
        if timeseries:
            latest = timeseries[-1]
            current = {"cpu": latest["cpu"], "ram": latest["ram"], "disk": latest["disk"]}
        else:
            current = {"cpu": None, "ram": None, "disk": None}
        
        return {
            "hours": hours,
            "bucketMinutes": bucket_minutes,
            "groupId": group_id,
            "dataPoints": len(timeseries),
            "current": current,
            "timeseries": timeseries
        }


# NOTE: metrics/history endpoint moved to Line ~6990 with more parameters


# ========== E5: Deployment Engine ==========

@app.post("/api/v1/deployments", dependencies=[Depends(verify_api_key)])
async def create_deployment(data: dict):
    """Create a new deployment (package version -> target)"""
    required = ["name", "packageVersionId", "targetType"]
    for f in required:
        if f not in data:
            raise HTTPException(400, f"Missing required field: {f}")
    
    package_version_id = data["packageVersionId"]
    target_type = data["targetType"]
    target_id = data.get("targetId")
    mode = data.get("mode", "required")
    
    if target_type not in ("node", "group", "all"):
        raise HTTPException(400, "targetType must be node, group, or all")
    if mode not in ("required", "available", "uninstall"):
        raise HTTPException(400, "mode must be required, available, or uninstall")
    if target_type in ("node", "group") and not target_id:
        raise HTTPException(400, f"targetId required for targetType={target_type}")
    
    async with db_pool.acquire() as conn:
        # Verify package version exists
        pv = await conn.fetchrow("SELECT id FROM package_versions WHERE id = $1", package_version_id)
        if not pv:
            raise HTTPException(404, "Package version not found")
        
        # Create deployment
        dep_id = await conn.fetchval("""
            INSERT INTO deployments (name, description, package_version_id, target_type, target_id, mode, 
                                      status, scheduled_start, scheduled_end, maintenance_window_only, created_by)
            VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11)
            RETURNING id
        """, data["name"], data.get("description"), package_version_id, target_type,
            target_id, mode, data.get("status", "active"),
            data.get("scheduledStart"), data.get("scheduledEnd"),
            data.get("maintenanceWindowOnly", False), data.get("createdBy"))
        
        # Create deployment_status entries for targeted nodes
        if target_type == "all":
            await conn.execute("""
                INSERT INTO deployment_status (deployment_id, node_id, status)
                SELECT $1, id, 'pending' FROM nodes
            """, dep_id)
        elif target_type == "node":
            await conn.execute("""
                INSERT INTO deployment_status (deployment_id, node_id, status)
                VALUES ($1, $2, 'pending')
            """, dep_id, target_id)
        elif target_type == "group":
            await conn.execute("""
                INSERT INTO deployment_status (deployment_id, node_id, status)
                SELECT $1, node_id, 'pending' FROM device_groups WHERE group_id = $2
            """, dep_id, target_id)
        
        return {"id": str(dep_id), "status": "created"}


@app.get("/api/v1/deployments", dependencies=[Depends(verify_api_key)])
async def list_deployments(status: str = None, limit: int = 50):
    """List all deployments with aggregated status"""
    async with db_pool.acquire() as conn:
        query = """
            SELECT d.*, p.name as package_name, pv.version as package_version,
                   COUNT(ds.id) as total_nodes,
                   COUNT(ds.id) FILTER (WHERE ds.status = 'success') as success_count,
                   COUNT(ds.id) FILTER (WHERE ds.status = 'failed') as failed_count,
                   COUNT(ds.id) FILTER (WHERE ds.status = 'pending') as pending_count,
                   COUNT(ds.id) FILTER (WHERE ds.status IN ('downloading', 'installing')) as in_progress_count
            FROM deployments d
            JOIN package_versions pv ON d.package_version_id = pv.id
            JOIN packages p ON pv.package_id = p.id
            LEFT JOIN deployment_status ds ON d.id = ds.deployment_id
        """
        params = []
        if status:
            query += " WHERE d.status = $1"
            params.append(status)
        query += " GROUP BY d.id, p.name, pv.version ORDER BY d.created_at DESC LIMIT $" + str(len(params) + 1)
        params.append(limit)
        
        rows = await conn.fetch(query, *params)
        return [dict(r) for r in rows]


@app.get("/api/v1/deployments/{deployment_id}", dependencies=[Depends(verify_api_key)])
async def get_deployment(deployment_id: str):
    """Get deployment details with per-node status"""
    async with db_pool.acquire() as conn:
        dep = await conn.fetchrow("""
            SELECT d.*, p.name as package_name, pv.version as package_version
            FROM deployments d
            JOIN package_versions pv ON d.package_version_id = pv.id
            JOIN packages p ON pv.package_id = p.id
            WHERE d.id = $1
        """, deployment_id)
        if not dep:
            raise HTTPException(404, "Deployment not found")
        
        statuses = await conn.fetch("""
            SELECT ds.*, n.node_id as node_name, n.hostname
            FROM deployment_status ds
            JOIN nodes n ON ds.node_id = n.id
            WHERE ds.deployment_id = $1
            ORDER BY ds.status, n.hostname
        """, deployment_id)
        
        result = dict(dep)
        result["nodes"] = [dict(s) for s in statuses]
        return result


@app.patch("/api/v1/deployments/{deployment_id}", dependencies=[Depends(verify_api_key)])
async def update_deployment(deployment_id: str, data: dict):
    """Update deployment (pause, resume, cancel)"""
    allowed_fields = {"status", "scheduled_start", "scheduled_end", "maintenance_window_only"}
    updates = {k: v for k, v in data.items() if k in allowed_fields or k.replace("_", "") in ["scheduledStart", "scheduledEnd", "maintenanceWindowOnly"]}
    
    if not updates:
        raise HTTPException(400, "No valid fields to update")
    
    async with db_pool.acquire() as conn:
        field_map = {"scheduledStart": "scheduled_start", "scheduledEnd": "scheduled_end", "maintenanceWindowOnly": "maintenance_window_only"}
        set_clauses = []
        params = [deployment_id]
        i = 2
        for k, v in updates.items():
            col = field_map.get(k, k)
            set_clauses.append(f"{col} = ${i}")
            params.append(v)
            i += 1
        
        set_clauses.append("updated_at = NOW()")
        await conn.execute(f"UPDATE deployments SET {', '.join(set_clauses)} WHERE id = $1", *params)
        return {"status": "updated"}


@app.delete("/api/v1/deployments/{deployment_id}", dependencies=[Depends(verify_api_key)])
async def delete_deployment(deployment_id: str):
    """Delete a deployment"""
    async with db_pool.acquire() as conn:
        result = await conn.execute("DELETE FROM deployments WHERE id = $1", deployment_id)
        if result == "DELETE 0":
            raise HTTPException(404, "Deployment not found")
        return {"status": "deleted"}


@app.get("/api/v1/nodes/{node_id}/deployments", dependencies=[Depends(verify_api_key)])
async def get_node_deployments(node_id: str):
    """Get pending deployments for a specific node (agent polling endpoint)"""
    async with db_pool.acquire() as conn:
        node = await conn.fetchrow("SELECT id FROM nodes WHERE node_id = $1 OR id::text = $1 OR id::text = $1", node_id)
        if not node:
            raise HTTPException(404, "Node not found")
        
        deployments = await conn.fetch("""
            SELECT d.id as deployment_id, d.name, d.mode, d.maintenance_window_only,
                   p.name as package_name, pv.version as package_version,
                   pv.install_command as installer_type, 
                   pv.download_url as installer_url,
                   pv.install_args,
                   pv.uninstall_args, 
                   pv.sha256_hash as expected_hash,
                   ds.status as node_status, ds.attempts
            FROM deployment_status ds
            JOIN deployments d ON ds.deployment_id = d.id
            JOIN package_versions pv ON d.package_version_id = pv.id
            JOIN packages p ON pv.package_id = p.id
            WHERE ds.node_id = $1
              AND d.status = 'active'
              AND ds.status IN ('pending', 'failed')
              AND (ds.attempts < 3 OR ds.attempts IS NULL)
              AND (d.scheduled_start IS NULL OR d.scheduled_start <= NOW())
              AND (d.scheduled_end IS NULL OR d.scheduled_end >= NOW())
            ORDER BY d.created_at
        """, node["id"])
        
        return [dict(d) for d in deployments]


@app.post("/api/v1/nodes/{node_id}/deployments/{deployment_id}/status", dependencies=[Depends(verify_api_key)])
async def update_node_deployment_status(node_id: str, deployment_id: str, data: dict):
    """Agent reports deployment status for this node"""
    status = data.get("status")
    if status not in ("pending", "downloading", "installing", "success", "failed", "skipped"):
        raise HTTPException(400, "Invalid status")
    
    async with db_pool.acquire() as conn:
        node = await conn.fetchrow("SELECT id FROM nodes WHERE node_id = $1 OR id::text = $1 OR id::text = $1", node_id)
        if not node:
            raise HTTPException(404, "Node not found")
        
        exit_code = data.get("exitCode")
        output = data.get("output")
        error_message = data.get("errorMessage")
        
        # Update status - use separate statements to avoid type ambiguity
        await conn.execute("""
            UPDATE deployment_status 
            SET status = $1::text, 
                exit_code = $2, 
                output = $3, 
                error_message = $4
            WHERE deployment_id = $5::uuid AND node_id = $6::uuid
        """, status, exit_code, output, error_message, deployment_id, node["id"])
        
        # Update timestamps separately
        if status == "downloading":
            await conn.execute("""
                UPDATE deployment_status 
                SET started_at = COALESCE(started_at, NOW())
                WHERE deployment_id = $1::uuid AND node_id = $2::uuid
            """, deployment_id, node["id"])
        elif status in ("success", "failed", "skipped"):
            await conn.execute("""
                UPDATE deployment_status 
                SET completed_at = NOW(),
                    attempts = attempts + 1,
                    last_attempt_at = NOW()
                WHERE deployment_id = $1::uuid AND node_id = $2::uuid
            """, deployment_id, node["id"])
        
        # Check if all nodes completed -> mark deployment as completed
        stats = await conn.fetchrow("""
            SELECT COUNT(*) as total,
                   COUNT(*) FILTER (WHERE status IN ('success', 'failed', 'skipped')) as completed
            FROM deployment_status WHERE deployment_id = $1
        """, deployment_id)
        
        if stats["total"] == stats["completed"]:
            await conn.execute("UPDATE deployments SET status = 'completed', updated_at = NOW() WHERE id = $1", deployment_id)
        
        return {"status": "updated"}


# Get package versions for dropdown


# ============================================================================
# E7: ALERTING & NOTIFICATIONS
# ============================================================================

# --- Alert Rules (Legacy - Replaced by E19) ---
# These endpoints use old schema, replaced by /api/v1/alert-* endpoints

# @app.get("/api/v1/alerts/rules", dependencies=[Depends(verify_api_key)])
# async def list_alert_rules_legacy():
#     """Legacy: List all alert rules"""
#     pass


# --- Notification Channels (Legacy - Replaced by E19 /api/v1/alert-channels) ---
# Old endpoints that use notification_channels table - replaced by alert_channels

# Legacy endpoints commented out - use /api/v1/alert-channels, /api/v1/alert-rules instead


@app.get("/api/v1/alerts/test-legacy")
async def test_legacy_endpoint():
    """Placeholder to maintain route prefix"""
    return {"message": "Use /api/v1/alert-channels and /api/v1/alert-rules instead"}


# --- Link Rules to Channels (Legacy) ---
# Removed old link_rule_to_channel endpoints


# --- Alert History (Legacy) ---
# Old alert history endpoint - replaced by /api/v1/alert-history

@app.get("/api/v1/alerts", dependencies=[Depends(verify_api_key)])
async def list_alerts_legacy():
    """Legacy: Redirects to new alert-history endpoint"""
    return {"message": "Use /api/v1/alert-history instead"}


@app.get("/api/v1/alerts/stats", dependencies=[Depends(verify_api_key)])
async def get_alert_stats_legacy():
    """Legacy: Alert statistics placeholder"""
    return {"message": "Use new E19 alert system"}


# Legacy alert endpoints removed - use /api/v1/alert-* endpoints


# ============================================================================
# Feature 2: Eventlog Charts - Trends over time
# ============================================================================

@app.get("/api/v1/eventlog/trends", dependencies=[Depends(verify_api_key)])
async def get_eventlog_trends(days: int = 7, db: asyncpg.Pool = Depends(get_db)):
    """Get eventlog trends by day for charts"""
    async with db.acquire() as conn:
        rows = await conn.fetch("""
            SELECT 
                DATE(event_time) as day,
                COUNT(*) FILTER (WHERE level <= 2) as errors,
                COUNT(*) FILTER (WHERE level = 3) as warnings,
                COUNT(*) as total
            FROM windows_eventlog
            WHERE event_time > NOW() - INTERVAL '%s days'
            GROUP BY DATE(event_time)
            ORDER BY day
        """ % days)
        
        return {
            "days": days,
            "trends": [
                {
                    "day": str(row["day"]),
                    "errors": row["errors"],
                    "warnings": row["warnings"],
                    "total": row["total"]
                }
                for row in rows
            ]
        }


# ============================================================================
# Feature 3: Software Comparison - Compare versions across nodes
# ============================================================================

@app.get("/api/v1/software/compare", dependencies=[Depends(verify_api_key)])
async def compare_software(software_name: str = None, db: asyncpg.Pool = Depends(get_db)):
    """Compare software versions across nodes"""
    async with db.acquire() as conn:
        if software_name:
            # Find specific software across all nodes
            rows = await conn.fetch("""
                SELECT 
                    n.id as node_id,
                    n.hostname,
                    s.name,
                    s.version,
                    s.publisher
                FROM nodes n
                JOIN software_current s ON n.id = s.node_id
                WHERE LOWER(s.name) LIKE $1
                ORDER BY s.name, n.hostname
            """, f"%{software_name.lower()}%")
            
            results = []
            for row in rows:
                results.append({
                    "nodeId": str(row["node_id"]),
                    "hostname": row["hostname"],
                    "name": row["name"],
                    "version": row["version"],
                    "publisher": row["publisher"]
                })
            
            # Group by version
            versions = {}
            for r in results:
                v = r["version"] or "Unknown"
                if v not in versions:
                    versions[v] = []
                versions[v].append({"nodeId": r["nodeId"], "hostname": r["hostname"]})
            
            return {
                "software": software_name,
                "totalNodes": len(set(r["nodeId"] for r in results)),
                "versions": versions,
                "results": results
            }
        else:
            # Get top installed software
            rows = await conn.fetch("""
                SELECT 
                    name,
                    COUNT(DISTINCT node_id) as node_count
                FROM software_current
                WHERE name IS NOT NULL AND name != ''
                GROUP BY name
                ORDER BY node_count DESC, name
                LIMIT 50
            """)
            
            return {
                "topSoftware": [{"name": r["name"], "count": r["node_count"]} for r in rows]
            }


# ============================================================================
# Feature 4: Compliance Dashboard - Security status at a glance
# ============================================================================


# ============================================================================
# Feature 5: OS Distribution - moved to line ~500 (before /nodes/{node_id})
# ============================================================================


# ============================================================================
# Feature 6: Export Functions - CSV/JSON export
# ============================================================================

from fastapi.responses import StreamingResponse
import io
import csv


# --- Gateway Status Endpoint (E11-05) ---

import aiohttp
from datetime import datetime

# Track gateway start time for uptime calculation
GATEWAY_START_TIME = datetime.utcnow()

@app.get("/api/v1/gateway/status", dependencies=[Depends(verify_api_key)])
async def get_gateway_status():
    """Get Octofleet Gateway health status for the dashboard widget"""
    gateway_url = "http://192.168.0.5:18789"
    
    try:
        async with aiohttp.ClientSession() as session:
            # Try to get gateway status
            async with session.get(f"{gateway_url}/status", timeout=aiohttp.ClientTimeout(total=5)) as resp:
                if resp.status == 200:
                    data = await resp.json()
                    return {
                        "online": True,
                        "version": data.get("version", "unknown"),
                        "uptime": data.get("uptime", "unknown"),
                        "connectedNodes": data.get("connectedNodes", 0),
                        "pendingJobs": data.get("pendingJobs", 0),
                        "lastCheck": datetime.utcnow().isoformat() + "Z"
                    }
    except Exception as e:
        pass
    
    # Fallback: Try basic connectivity check
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(f"{gateway_url}/", timeout=aiohttp.ClientTimeout(total=3)) as resp:
                return {
                    "online": resp.status < 500,
                    "version": "unknown",
                    "uptime": "unknown", 
                    "connectedNodes": 0,
                    "pendingJobs": 0,
                    "lastCheck": datetime.utcnow().isoformat() + "Z",
                    "note": "Basic connectivity only - detailed status unavailable"
                }
    except Exception as e:
        return {
            "online": False,
            "version": "unknown",
            "uptime": "unknown",
            "connectedNodes": 0,
            "pendingJobs": 0,
            "lastCheck": datetime.utcnow().isoformat() + "Z",
            "error": str(e)
        }


# --- Gateway Logs Endpoint (E11-06) ---

# In-memory log buffer for demo purposes
# In production, this would read from actual gateway logs
import collections
LOG_BUFFER = collections.deque(maxlen=500)

def add_log_entry(level: str, message: str, node_id: str = None, job_id: str = None):
    """Add a log entry to the buffer"""
    LOG_BUFFER.append({
        "timestamp": datetime.utcnow().isoformat() + "Z",
        "level": level,
        "message": message,
        "node_id": node_id,
        "job_id": job_id
    })

# Seed some demo logs
add_log_entry("info", "Gateway started", None, None)
add_log_entry("info", "Database connection established", None, None)

@app.get("/api/v1/gateway/logs", dependencies=[Depends(verify_api_key)])
async def get_gateway_logs(
    level: Optional[str] = None,
    node_id: Optional[str] = None,
    job_id: Optional[str] = None,
    limit: int = 100
):
    """Get gateway logs with optional filters"""
    logs = list(LOG_BUFFER)
    
    # Apply filters
    if level:
        logs = [l for l in logs if l["level"] == level]
    if node_id:
        logs = [l for l in logs if l.get("node_id") == node_id]
    if job_id:
        logs = [l for l in logs if l.get("job_id") == job_id]
    
    # Return most recent logs
    logs = logs[-limit:]
    
    return {
        "logs": logs,
        "total": len(logs),
        "filters": {"level": level, "node_id": node_id, "job_id": job_id}
    }


# ============================================
# Auth & User Management (Moved to routers/auth.py)
# ============================================


# ============================================
# Feature 2: Eventlog Charts - Trends over time
# ============================================


@app.post("/api/v1/users/{user_id}/roles/{role_name}")
async def assign_role(
    user_id: str,
    role_name: str,
    user: CurrentUser = Depends(require_permission("users:write"))
):
    """Assign role to user"""
    async with db_pool.acquire() as conn:
        role = await conn.fetchrow("SELECT id FROM roles WHERE name = $1", role_name)
        if not role:
            raise not_found("Role", role_id)
        
        await conn.execute(
            """INSERT INTO user_roles (user_id, role_id, assigned_by)
               VALUES ($1, $2, $3)
               ON CONFLICT DO NOTHING""",
            UUID(user_id),
            role["id"],
            UUID(user.id) if user.id != "system" else None
        )
        return {"status": "role assigned"}


@app.delete("/api/v1/users/{user_id}/roles/{role_name}")
async def remove_role(
    user_id: str,
    role_name: str,
    user: CurrentUser = Depends(require_permission("users:write"))
):
    """Remove role from user"""
    async with db_pool.acquire() as conn:
        role = await conn.fetchrow("SELECT id FROM roles WHERE name = $1", role_name)
        if not role:
            raise not_found("Role", role_id)
        
        await conn.execute(
            "DELETE FROM user_roles WHERE user_id = $1 AND role_id = $2",
            UUID(user_id),
            role["id"]
        )
        return {"status": "role removed"}


# --- Role Management ---

@app.get("/api/v1/roles")
async def list_roles(user: CurrentUser = Depends(require_permission("users:read"))):
    """List all roles"""
    async with db_pool.acquire() as conn:
        roles = await conn.fetch(
            "SELECT id, name, description, permissions, is_system, created_at FROM roles ORDER BY name"
        )
        return {
            "roles": [
                {
                    "id": str(r["id"]),
                    "name": r["name"],
                    "description": r["description"],
                    "permissions": r["permissions"],
                    "is_system": r["is_system"],
                    "created_at": r["created_at"].isoformat() if r["created_at"] else None
                }
                for r in roles
            ]
        }


@app.post("/api/v1/roles")
async def create_role(
    data: RoleCreate,
    user: CurrentUser = Depends(require_permission("roles:write"))
):
    """Create custom role"""
    async with db_pool.acquire() as conn:
        role = await conn.fetchrow(
            """INSERT INTO roles (name, description, permissions, is_system)
               VALUES ($1, $2, $3, false)
               RETURNING id, name, description, permissions, is_system, created_at""",
            data.name,
            data.description,
            data.permissions
        )
        return {
            "id": str(role["id"]),
            "name": role["name"],
            "description": role["description"],
            "permissions": role["permissions"],
            "is_system": role["is_system"],
            "created_at": role["created_at"].isoformat()
        }


@app.delete("/api/v1/roles/{role_id}")
async def delete_role(
    role_id: str,
    user: CurrentUser = Depends(require_permission("roles:write"))
):
    """Delete custom role (not system roles)"""
    async with db_pool.acquire() as conn:
        role = await conn.fetchrow(
            "SELECT is_system FROM roles WHERE id = $1",
            UUID(role_id)
        )
        if not role:
            raise not_found("Role", role_id)
        if role["is_system"]:
            raise HTTPException(status_code=400, detail="Cannot delete system role")
        
        await conn.execute("DELETE FROM roles WHERE id = $1", UUID(role_id))
        return {"status": "deleted"}


# --- Initial Admin Setup ---

@app.post("/api/v1/auth/setup")
async def setup_admin(data: UserCreate):
    """
    One-time admin setup. Only works if no users exist.
    """
    async with db_pool.acquire() as conn:
        count = await conn.fetchval("SELECT COUNT(*) FROM users")
        if count > 0:
            raise HTTPException(status_code=400, detail="Setup already completed")
        
        # Create admin user
        user = await conn.fetchrow(
            """INSERT INTO users (username, email, password_hash, display_name, is_superuser)
               VALUES ($1, $2, $3, $4, true)
               RETURNING id""",
            data.username,
            data.email,
            hash_password(data.password),
            data.display_name or data.username
        )
        
        # Assign admin role
        admin_role = await conn.fetchval("SELECT id FROM roles WHERE name = 'admin'")
        if admin_role:
            await conn.execute(
                "INSERT INTO user_roles (user_id, role_id) VALUES ($1, $2)",
                user["id"],
                admin_role
            )
        
        return {"status": "Admin created", "username": data.username}


# --- Audit Log Endpoints ---

@app.get("/api/v1/audit")
async def get_audit_log(
    limit: int = 100,
    offset: int = 0,
    user_id: Optional[str] = None,
    action: Optional[str] = None,
    resource_type: Optional[str] = None,
    user: CurrentUser = Depends(require_permission("audit:read"))
):
    """Get audit log entries"""
    async with db_pool.acquire() as conn:
        query = """
            SELECT a.id, a.timestamp, a.user_id, u.username, a.action, 
                   a.resource_type, a.resource_id, a.details, a.ip_address
            FROM audit_log a
            LEFT JOIN users u ON a.user_id = u.id
            WHERE 1=1
        """
        params = []
        param_idx = 1
        
        if user_id:
            query += f" AND a.user_id = ${param_idx}"
            params.append(UUID(user_id))
            param_idx += 1
        if action:
            query += f" AND a.action ILIKE ${param_idx}"
            params.append(f"%{action}%")
            param_idx += 1
        if resource_type:
            query += f" AND a.resource_type = ${param_idx}"
            params.append(resource_type)
            param_idx += 1
        
        query += f" ORDER BY a.timestamp DESC LIMIT ${param_idx} OFFSET ${param_idx + 1}"
        params.extend([limit, offset])
        
        rows = await conn.fetch(query, *params)
        total = await conn.fetchval("SELECT COUNT(*) FROM audit_log")
        
        return {
            "entries": [
                {
                    "id": r["id"],
                    "timestamp": r["timestamp"].isoformat() if r["timestamp"] else None,
                    "user_id": str(r["user_id"]) if r["user_id"] else None,
                    "username": r["username"],
                    "action": r["action"],
                    "resource_type": r["resource_type"],
                    "resource_id": r["resource_id"],
                    "details": r["details"],
                    "ip_address": str(r["ip_address"]) if r["ip_address"] else None
                }
                for r in rows
            ],
            "total": total,
            "limit": limit,
            "offset": offset
        }


async def log_audit(
    conn,
    user_id: Optional[str],
    action: str,
    resource_type: Optional[str] = None,
    resource_id: Optional[str] = None,
    details: Optional[dict] = None,
    ip_address: Optional[str] = None
):
    """Helper to insert audit log entry"""
    await conn.execute(
        """INSERT INTO audit_log (user_id, action, resource_type, resource_id, details, ip_address)
           VALUES ($1, $2, $3, $4, $5, $6)""",
        UUID(user_id) if user_id and user_id != "system" else None,
        action,
        resource_type,
        resource_id,
        json.dumps(details) if details else None,
        ip_address
    )


# --- API Key Management ---

@app.get("/api/v1/api-keys")
async def list_api_keys(user: CurrentUser = Depends(require_auth)):
    """List API keys for current user (or all if admin)"""
    async with db_pool.acquire() as conn:
        if user.is_superuser or user.has_permission("users:read"):
            # Admin sees all keys
            keys = await conn.fetch(
                """SELECT k.id, k.user_id, u.username, k.name, k.permissions, k.scopes,
                          k.expires_at, k.last_used, k.created_at, k.is_active
                   FROM api_keys k
                   LEFT JOIN users u ON k.user_id = u.id
                   ORDER BY k.created_at DESC"""
            )
        else:
            # User sees own keys
            keys = await conn.fetch(
                """SELECT id, user_id, name, permissions, scopes, expires_at, last_used, created_at, is_active
                   FROM api_keys WHERE user_id = $1
                   ORDER BY created_at DESC""",
                UUID(user.id)
            )
        
        return {
            "keys": [
                {
                    "id": str(k["id"]),
                    "user_id": str(k["user_id"]) if k["user_id"] else None,
                    "username": k.get("username"),
                    "name": k["name"],
                    "permissions": k["permissions"] or [],
                    "scopes": list(k["scopes"]) if k["scopes"] else ["agent"],
                    "expires_at": k["expires_at"].isoformat() if k["expires_at"] else None,
                    "last_used": k["last_used"].isoformat() if k["last_used"] else None,
                    "created_at": k["created_at"].isoformat() if k["created_at"] else None,
                    "is_active": k["is_active"]
                }
                for k in keys
            ]
        }


@app.post("/api/v1/api-keys")
async def create_api_key(
    data: APIKeyCreate,
    user: CurrentUser = Depends(require_auth)
):
    """Create new API key for current user"""
    import secrets
    
    # Generate key
    raw_key = f"oci_{secrets.token_hex(24)}"  # oci = octofleet inventory
    key_hash = hash_api_key(raw_key)
    
    expires_at = None
    if data.expires_days:
        expires_at = datetime.utcnow() + timedelta(days=data.expires_days)
    
    async with db_pool.acquire() as conn:
        scopes = data.scopes if data.scopes else ["agent"]
        key = await conn.fetchrow(
            """INSERT INTO api_keys (user_id, key_hash, name, scopes, expires_at)
               VALUES ($1, $2, $3, $4, $5)
               RETURNING id, name, created_at, expires_at""",
            UUID(user.id) if user.id != "system" else None,
            key_hash,
            data.name,
            scopes,
            expires_at
        )
        
        # Return key only once (never stored in plain text)
        return {
            "id": str(key["id"]),
            "name": key["name"],
            "key": raw_key,  # Only shown once!
            "scopes": scopes,
            "created_at": key["created_at"].isoformat(),
            "expires_at": key["expires_at"].isoformat() if key["expires_at"] else None,
            "warning": "Save this key now! It won't be shown again."
        }


@app.delete("/api/v1/api-keys/{key_id}")
async def revoke_api_key(
    key_id: str,
    user: CurrentUser = Depends(require_auth)
):
    """Revoke an API key"""
    async with db_pool.acquire() as conn:
        # Check ownership (unless admin)
        if not user.is_superuser and not user.has_permission("users:write"):
            key = await conn.fetchrow(
                "SELECT user_id FROM api_keys WHERE id = $1",
                UUID(key_id)
            )
            if not key or str(key["user_id"]) != user.id:
                raise HTTPException(status_code=403, detail="Cannot revoke this key")
        
        await conn.execute(
            "UPDATE api_keys SET is_active = false WHERE id = $1",
            UUID(key_id)
        )
        return {"status": "revoked"}


@app.delete("/api/v1/api-keys/{key_id}/permanent")
async def delete_api_key(
    key_id: str,
    user: CurrentUser = Depends(require_permission("users:write"))
):
    """Permanently delete an API key (admin only)"""
    async with db_pool.acquire() as conn:
        await conn.execute("DELETE FROM api_keys WHERE id = $1", UUID(key_id))
        return {"status": "deleted"}


# ========== E9: Maintenance Windows ==========

@app.get("/api/v1/maintenance-windows", dependencies=[Depends(verify_api_key)])
async def list_maintenance_windows():
    """List all maintenance windows"""
    async with db_pool.acquire() as conn:
        # Create table if not exists
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS maintenance_windows (
                id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
                name TEXT NOT NULL,
                description TEXT,
                start_time TIME NOT NULL,
                end_time TIME NOT NULL,
                days_of_week INTEGER[] NOT NULL DEFAULT '{1,2,3,4,5}',
                timezone TEXT DEFAULT 'Europe/Berlin',
                is_active BOOLEAN DEFAULT true,
                target_type TEXT CHECK (target_type IN ('all', 'group', 'node')),
                target_id UUID,
                created_at TIMESTAMPTZ DEFAULT NOW(),
                updated_at TIMESTAMPTZ DEFAULT NOW()
            )
        """)
        rows = await conn.fetch("SELECT * FROM maintenance_windows ORDER BY name")
        return {"windows": [dict(r) for r in rows]}


@app.post("/api/v1/maintenance-windows", dependencies=[Depends(require_auth)])
async def create_maintenance_window(data: dict):
    """Create a maintenance window"""
    required = ["name", "startTime", "endTime"]
    for f in required:
        if f not in data:
            raise HTTPException(400, f"Missing required field: {f}")
    
    async with db_pool.acquire() as conn:
        window_id = await conn.fetchval("""
            INSERT INTO maintenance_windows (name, description, start_time, end_time, days_of_week, 
                                             timezone, target_type, target_id)
            VALUES ($1, $2, $3, $4, $5, $6, $7, $8)
            RETURNING id
        """, data["name"], data.get("description"),
            data["startTime"], data["endTime"],
            data.get("daysOfWeek", [1, 2, 3, 4, 5]),
            data.get("timezone", "Europe/Berlin"),
            data.get("targetType", "all"),
            data.get("targetId"))
        return {"id": str(window_id), "status": "created"}


@app.get("/api/v1/maintenance-windows/{window_id}", dependencies=[Depends(verify_api_key)])
async def get_maintenance_window(window_id: str):
    """Get a specific maintenance window"""
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow("SELECT * FROM maintenance_windows WHERE id = $1", window_id)
        if not row:
            raise HTTPException(404, "Maintenance window not found")
        return dict(row)


@app.put("/api/v1/maintenance-windows/{window_id}", dependencies=[Depends(require_auth)])
async def update_maintenance_window(window_id: str, data: dict):
    """Update a maintenance window"""
    async with db_pool.acquire() as conn:
        await conn.execute("""
            UPDATE maintenance_windows
            SET name = COALESCE($2, name),
                description = COALESCE($3, description),
                start_time = COALESCE($4, start_time),
                end_time = COALESCE($5, end_time),
                days_of_week = COALESCE($6, days_of_week),
                timezone = COALESCE($7, timezone),
                is_active = COALESCE($8, is_active),
                target_type = COALESCE($9, target_type),
                target_id = COALESCE($10, target_id),
                updated_at = NOW()
            WHERE id = $1
        """, window_id, data.get("name"), data.get("description"),
            data.get("startTime"), data.get("endTime"),
            data.get("daysOfWeek"), data.get("timezone"),
            data.get("isActive"), data.get("targetType"), data.get("targetId"))
        return {"status": "updated"}


@app.delete("/api/v1/maintenance-windows/{window_id}", dependencies=[Depends(require_auth)])
async def delete_maintenance_window(window_id: str):
    """Delete a maintenance window"""
    async with db_pool.acquire() as conn:
        await conn.execute("DELETE FROM maintenance_windows WHERE id = $1", window_id)
        return {"status": "deleted"}


@app.get("/api/v1/maintenance-windows/check/{node_id}", dependencies=[Depends(verify_api_key)])
async def check_maintenance_window(node_id: str):
    """Check if a node is currently in a maintenance window"""
    from datetime import datetime
    import pytz
    
    async with db_pool.acquire() as conn:
        # Get node's group memberships
        groups = await conn.fetch(
            "SELECT group_id FROM device_groups WHERE node_id = $1", node_id
        )
        group_ids = [str(g["group_id"]) for g in groups]
        
        # Get all active maintenance windows
        windows = await conn.fetch("""
            SELECT * FROM maintenance_windows 
            WHERE is_active = true
        """)
        
        for w in windows:
            # Check if window applies to this node
            if w["target_type"] == "node" and str(w["target_id"]) != node_id:
                continue
            if w["target_type"] == "group" and str(w["target_id"]) not in group_ids:
                continue
            
            # Check if current time is within window
            tz = pytz.timezone(w["timezone"] or "Europe/Berlin")
            now = datetime.now(tz)
            
            # Check day of week (1=Monday, 7=Sunday)
            if now.isoweekday() not in (w["days_of_week"] or [1, 2, 3, 4, 5]):
                continue
            
            # Check time range
            current_time = now.time()
            if w["start_time"] <= current_time <= w["end_time"]:
                return {
                    "in_maintenance_window": True,
                    "window_id": str(w["id"]),
                    "window_name": w["name"],
                    "ends_at": w["end_time"].isoformat()
                }
        
        return {"in_maintenance_window": False}


# ========== E9: Rollout Strategies ==========

@app.get("/api/v1/rollout-strategies", dependencies=[Depends(verify_api_key)])
async def list_rollout_strategies():
    """List available rollout strategies"""
    return {
        "strategies": [
            {
                "id": "immediate",
                "name": "Sofort (Immediate)",
                "description": "Alle Zielgeräte gleichzeitig",
                "config_schema": {}
            },
            {
                "id": "staged",
                "name": "Gestaffelt (Staged)",
                "description": "Rollout in Wellen mit Wartezeit zwischen Gruppen",
                "config_schema": {
                    "wave_size": {"type": "integer", "default": 10, "description": "Geräte pro Welle"},
                    "wave_delay_minutes": {"type": "integer", "default": 60, "description": "Wartezeit zwischen Wellen (Minuten)"},
                    "success_threshold_percent": {"type": "integer", "default": 90, "description": "Min. Erfolgsrate um fortzufahren (%)"}
                }
            },
            {
                "id": "canary",
                "name": "Canary",
                "description": "Erst kleine Testgruppe, dann voller Rollout",
                "config_schema": {
                    "canary_count": {"type": "integer", "default": 1, "description": "Anzahl Canary-Geräte"},
                    "canary_duration_hours": {"type": "integer", "default": 24, "description": "Canary-Beobachtungszeit (Stunden)"},
                    "auto_proceed": {"type": "boolean", "default": False, "description": "Automatisch fortfahren wenn Canary erfolgreich"}
                }
            },
            {
                "id": "percentage",
                "name": "Prozentual",
                "description": "Schrittweise Erhöhung der Zielgruppe",
                "config_schema": {
                    "initial_percent": {"type": "integer", "default": 10, "description": "Startprozent"},
                    "increment_percent": {"type": "integer", "default": 20, "description": "Erhöhung pro Schritt"},
                    "step_delay_hours": {"type": "integer", "default": 4, "description": "Wartezeit zwischen Schritten (Stunden)"}
                }
            }
        ]
    }


@app.post("/api/v1/deployments/{deployment_id}/rollout", dependencies=[Depends(verify_api_key)])
async def configure_rollout_strategy(deployment_id: str, data: dict):
    """Configure rollout strategy for a deployment"""
    strategy = data.get("strategy", "immediate")
    config = data.get("config", {})
    
    valid_strategies = ["immediate", "staged", "canary", "percentage"]
    if strategy not in valid_strategies:
        raise HTTPException(400, f"Invalid strategy. Must be one of: {valid_strategies}")
    
    async with db_pool.acquire() as conn:
        # Ensure rollout_config column exists
        await conn.execute("""
            ALTER TABLE deployments 
            ADD COLUMN IF NOT EXISTS rollout_strategy TEXT DEFAULT 'immediate',
            ADD COLUMN IF NOT EXISTS rollout_config JSONB DEFAULT '{}',
            ADD COLUMN IF NOT EXISTS rollout_state JSONB DEFAULT '{}'
        """)
        
        await conn.execute("""
            UPDATE deployments 
            SET rollout_strategy = $2, rollout_config = $3, updated_at = NOW()
            WHERE id = $1
        """, deployment_id, strategy, json.dumps(config))
        
        return {"status": "configured", "strategy": strategy}


@app.get("/api/v1/deployments/{deployment_id}/rollout", dependencies=[Depends(verify_api_key)])
async def get_rollout_status(deployment_id: str):
    """Get rollout strategy status for a deployment"""
    async with db_pool.acquire() as conn:
        dep = await conn.fetchrow("""
            SELECT rollout_strategy, rollout_config, rollout_state,
                   (SELECT COUNT(*) FROM deployment_status WHERE deployment_id = $1) as total,
                   (SELECT COUNT(*) FROM deployment_status WHERE deployment_id = $1 AND status = 'success') as success,
                   (SELECT COUNT(*) FROM deployment_status WHERE deployment_id = $1 AND status = 'failed') as failed,
                   (SELECT COUNT(*) FROM deployment_status WHERE deployment_id = $1 AND status = 'pending') as pending
            FROM deployments WHERE id = $1
        """, deployment_id)
        
        if not dep:
            raise HTTPException(404, "Deployment not found")
        
        return {
            "strategy": dep["rollout_strategy"] or "immediate",
            "config": json.loads(dep["rollout_config"] or "{}"),
            "state": json.loads(dep["rollout_state"] or "{}"),
            "progress": {
                "total": dep["total"],
                "success": dep["success"],
                "failed": dep["failed"],
                "pending": dep["pending"],
                "success_rate": round(dep["success"] / dep["total"] * 100, 1) if dep["total"] > 0 else 0
            }
        }


@app.post("/api/v1/deployments/{deployment_id}/rollout/advance", dependencies=[Depends(verify_api_key)])
async def advance_rollout(deployment_id: str):
    """Manually advance a staged/canary rollout to the next phase"""
    async with db_pool.acquire() as conn:
        dep = await conn.fetchrow("""
            SELECT rollout_strategy, rollout_config, rollout_state 
            FROM deployments WHERE id = $1
        """, deployment_id)
        
        if not dep:
            raise HTTPException(404, "Deployment not found")
        
        strategy = dep["rollout_strategy"] or "immediate"
        config = json.loads(dep["rollout_config"] or "{}")
        state = json.loads(dep["rollout_state"] or "{}")
        
        if strategy == "immediate":
            return {"message": "Immediate rollout - no phases to advance"}
        
        # Get current wave/phase
        current_wave = state.get("current_wave", 0)
        
        if strategy == "canary":
            if not state.get("canary_complete"):
                # Mark canary as complete, enable full rollout
                state["canary_complete"] = True
                state["full_rollout_started"] = datetime.now().isoformat()
                
                # Enable all remaining pending nodes
                await conn.execute("""
                    UPDATE deployment_status 
                    SET status = 'pending', updated_at = NOW()
                    WHERE deployment_id = $1 AND status = 'pending'
                """, deployment_id)
                
                await conn.execute("""
                    UPDATE deployments SET rollout_state = $2, updated_at = NOW() WHERE id = $1
                """, deployment_id, json.dumps(state))
                
                return {"message": "Canary complete - full rollout started", "state": state}
            else:
                return {"message": "Full rollout already in progress"}
        
        elif strategy == "staged":
            wave_size = config.get("wave_size", 10)
            
            # Activate next wave
            await conn.execute("""
                UPDATE deployment_status 
                SET status = 'pending', updated_at = NOW()
                WHERE deployment_id = $1 
                AND status = 'waiting'
                AND id IN (
                    SELECT id FROM deployment_status 
                    WHERE deployment_id = $1 AND status = 'waiting'
                    LIMIT $2
                )
            """, deployment_id, wave_size)
            
            state["current_wave"] = current_wave + 1
            state["wave_started_at"] = datetime.now().isoformat()
            
            await conn.execute("""
                UPDATE deployments SET rollout_state = $2, updated_at = NOW() WHERE id = $1
            """, deployment_id, json.dumps(state))
            
            return {"message": f"Advanced to wave {state['current_wave']}", "state": state}
        
        elif strategy == "percentage":
            current_percent = state.get("current_percent", config.get("initial_percent", 10))
            increment = config.get("increment_percent", 20)
            new_percent = min(100, current_percent + increment)
            
            # Calculate how many more nodes to activate
            total = await conn.fetchval(
                "SELECT COUNT(*) FROM deployment_status WHERE deployment_id = $1", deployment_id
            )
            target_count = int(total * new_percent / 100)
            current_active = await conn.fetchval(
                "SELECT COUNT(*) FROM deployment_status WHERE deployment_id = $1 AND status != 'waiting'",
                deployment_id
            )
            to_activate = target_count - current_active
            
            if to_activate > 0:
                await conn.execute("""
                    UPDATE deployment_status 
                    SET status = 'pending', updated_at = NOW()
                    WHERE deployment_id = $1 
                    AND status = 'waiting'
                    AND id IN (
                        SELECT id FROM deployment_status 
                        WHERE deployment_id = $1 AND status = 'waiting'
                        LIMIT $2
                    )
                """, deployment_id, to_activate)
            
            state["current_percent"] = new_percent
            state["step_started_at"] = datetime.now().isoformat()
            
            await conn.execute("""
                UPDATE deployments SET rollout_state = $2, updated_at = NOW() WHERE id = $1
            """, deployment_id, json.dumps(state))
            
            return {"message": f"Advanced to {new_percent}%", "state": state}
        
        return {"message": "Unknown strategy"}

# ============================================
# E13: Vulnerability Tracking API
# ============================================

from vulnerability import VulnerabilityScanner, get_vulnerability_summary, get_node_vulnerabilities
import logging
logger = logging.getLogger(__name__)


# ============================================
# System Settings API
# ============================================

@app.get("/api/v1/settings")
async def get_settings():
    """Get all system settings (values masked for secrets)."""
    async with db_pool.acquire() as conn:
        rows = await conn.fetch("SELECT key, value, updated_at FROM system_settings")
        settings = {}
        for row in rows:
            key = row["key"]
            value = row["value"]
            # Mask sensitive values
            if "key" in key.lower() or "secret" in key.lower() or "password" in key.lower():
                settings[key] = {"value": "***" + (value[-4:] if value and len(value) > 4 else ""), "masked": True, "updated_at": row["updated_at"].isoformat() if row["updated_at"] else None}
            else:
                settings[key] = {"value": value, "masked": False, "updated_at": row["updated_at"].isoformat() if row["updated_at"] else None}
        return settings

@app.get("/api/v1/settings/{key}")
async def get_setting(key: str):
    """Get a specific setting."""
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow("SELECT value, updated_at FROM system_settings WHERE key = $1", key)
        if not row:
            raise HTTPException(status_code=404, detail="Setting not found")
        return {"key": key, "value": row["value"], "updated_at": row["updated_at"].isoformat() if row["updated_at"] else None}

@app.put("/api/v1/settings/{key}")
async def update_setting(key: str, value: str = Body(..., embed=True)):
    """Update or create a setting."""
    async with db_pool.acquire() as conn:
        await conn.execute("""
            INSERT INTO system_settings (key, value, updated_at, updated_by)
            VALUES ($1, $2, NOW(), 'admin')
            ON CONFLICT (key) DO UPDATE SET
                value = EXCLUDED.value,
                updated_at = NOW(),
                updated_by = EXCLUDED.updated_by
        """, key, value)
    
    # Special handling: if NVD_API_KEY changed, update the scanner
    if key == "nvd_api_key":
        os.environ["NVD_API_KEY"] = value
    
    return {"status": "updated", "key": key}

@app.delete("/api/v1/settings/{key}")
async def delete_setting(key: str):
    """Delete a setting."""
    async with db_pool.acquire() as conn:
        result = await conn.execute("DELETE FROM system_settings WHERE key = $1", key)
        if result == "DELETE 0":
            raise HTTPException(status_code=404, detail="Setting not found")
    return {"status": "deleted", "key": key}

@app.get("/api/v1/test/nvd")
async def test_nvd():
    """Test NVD API connectivity."""
    import httpx
    from urllib.parse import quote
    
    keyword = "Google Chrome"
    url = f"https://services.nvd.nist.gov/rest/json/cves/2.0?keywordSearch={quote(keyword)}&resultsPerPage=1"
    headers = {"User-Agent": "Octofleet-Inventory/1.0"}
    
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            response = await client.get(url, headers=headers)
            return {
                "status_code": response.status_code,
                "response_length": len(response.text),
                "first_100": response.text[:100]
            }
    except Exception as e:
        return {"error": str(e)}


# ============================================
# E14: Auto-Remediation API
# ============================================

from remediation import (
    RemediationPackageCreate, RemediationPackageUpdate,
    RemediationRuleCreate, RemediationRuleUpdate,
    MaintenanceWindowCreate, TriggerRemediationRequest, ApproveRemediationRequest,
    get_remediation_packages, get_remediation_package, create_remediation_package,
    update_remediation_package, delete_remediation_package,
    get_remediation_rules, get_remediation_rule, create_remediation_rule,
    update_remediation_rule, delete_remediation_rule,
    get_maintenance_windows, create_maintenance_window, is_in_maintenance_window,
    get_remediation_jobs, get_remediation_job, approve_remediation_jobs,
    update_remediation_job_status, get_remediation_summary,
    RemediationEngine
)


# --- Remediation Packages (Fix Mappings) ---


# --- Remediation Rules ---


# --- Maintenance Windows ---


# --- Remediation Jobs ---


# --- Remediation Engine ---

class TriggerRemediationRequest(BaseModel):
    severity_filter: List[str] = ["CRITICAL", "HIGH"]
    software_filter: Optional[str] = None
    node_ids: Optional[List[str]] = None
    dry_run: bool = True

class RemediationEngine:
    def __init__(self, pool):
        self.pool = pool

    async def scan_and_create_jobs(self, severity_filter=None, software_filter=None, node_ids=None, dry_run=True):
        async with self.pool.acquire() as conn:
            # Find vulnerabilities with available fixes
            where = "WHERE 1=1"
            params = []
            idx = 1
            if severity_filter:
                placeholders = ", ".join(f"${i}" for i in range(idx, idx + len(severity_filter)))
                where += f" AND UPPER(v.severity) IN ({placeholders})"
                params.extend([s.upper() for s in severity_filter])
                idx += len(severity_filter)
            if software_filter:
                where += f" AND v.software_name ILIKE ${idx}"
                params.append(f"%{software_filter}%")
                idx += 1

            vulns = await conn.fetch(f"""
                SELECT v.id, v.software_name, v.software_version, v.cve_id, v.severity, v.cvss_score,
                       nv.node_id
                FROM vulnerabilities v
                JOIN node_vulnerabilities nv ON nv.vulnerability_id = v.id
                {where}
                ORDER BY v.cvss_score DESC NULLS LAST
                LIMIT 500
            """, *params)

            if node_ids:
                vulns = [v for v in vulns if str(v["node_id"]) in node_ids]

            # Match to remediation packages
            packages = await conn.fetch("SELECT * FROM remediation_packages WHERE enabled = true")
            pkg_map = {}
            for p in packages:
                key = (p["target_software"] or "").lower()
                if key not in pkg_map:
                    pkg_map[key] = p

            jobs_created = []
            jobs_preview = []
            for v in vulns:
                sw = (v["software_name"] or "").lower()
                pkg = pkg_map.get(sw)
                fix_method = pkg["fix_method"] if pkg else "winget"
                fix_cmd = pkg["fix_command"] if pkg else f"winget upgrade {v['software_name']}"

                job_info = {
                    "vulnerabilityId": v["id"],
                    "nodeId": str(v["node_id"]),
                    "softwareName": v["software_name"],
                    "softwareVersion": v["software_version"],
                    "cveId": v["cve_id"],
                    "severity": v["severity"],
                    "fixMethod": fix_method,
                    "fixCommand": fix_cmd,
                    "packageId": str(pkg["id"]) if pkg else None,
                }

                if dry_run:
                    jobs_preview.append(job_info)
                else:
                    row = await conn.fetchrow("""
                        INSERT INTO remediation_jobs (vulnerability_id, remediation_package_id, node_id,
                            software_name, software_version, cve_id, status, requires_approval)
                        VALUES ($1, $2, $3::uuid, $4, $5, $6, 'pending', false) RETURNING id
                    """, v["id"], pkg["id"] if pkg else None, str(v["node_id"]),
                        v["software_name"], v["software_version"], v["cve_id"])
                    job_info["id"] = str(row["id"])
                    job_info["status"] = "pending"
                    jobs_created.append(job_info)

            return {
                "dryRun": dry_run,
                "totalVulnerabilities": len(vulns),
                "jobsCreated": len(jobs_created),
                "jobsPreview": len(jobs_preview),
                "jobs": jobs_created if not dry_run else jobs_preview,
            }

async def get_remediation_summary(pool):
    async with pool.acquire() as conn:
        fixed = await conn.fetchval("SELECT COUNT(*) FROM remediation_jobs WHERE status IN ('completed', 'success')") or 0
        pending = await conn.fetchval("SELECT COUNT(*) FROM remediation_jobs WHERE status IN ('pending', 'approved', 'running')") or 0
        failed = await conn.fetchval("SELECT COUNT(*) FROM remediation_jobs WHERE status = 'failed'") or 0
        running = await conn.fetchval("SELECT COUNT(*) FROM remediation_jobs WHERE status = 'running'") or 0
        fixable = await conn.fetchval("""
            SELECT COUNT(DISTINCT v.id) FROM vulnerabilities v
            JOIN node_vulnerabilities nv ON nv.vulnerability_id = v.id
            WHERE UPPER(v.severity) IN ('CRITICAL', 'HIGH')
        """) or 0
        active_packages = await conn.fetchval("SELECT COUNT(*) FROM remediation_packages WHERE enabled = true") or 0
        active_rules = await conn.fetchval("SELECT COUNT(*) FROM remediation_rules WHERE enabled = true") or 0
        recent = await conn.fetch("""
            SELECT rj.id, rj.node_id, rj.software_name, rj.software_version, rj.cve_id, rj.status, 
                   rp.fix_method, rj.created_at, rj.completed_at, rj.exit_code, rj.error_message
            FROM remediation_jobs rj
            LEFT JOIN remediation_packages rp ON rp.id = rj.remediation_package_id
            ORDER BY rj.created_at DESC LIMIT 20
        """)
        return {
            "fixed": fixed,
            "pending": pending,
            "failed": failed,
            "fixableCves": fixable,
            "fixable_vulnerabilities": fixable,
            "active_packages": active_packages,
            "active_rules": active_rules,
            "job_counts": {"success": fixed, "completed": fixed, "approved": pending, "pending": pending, "failed": failed, "running": running},
            "recent_jobs": [{
                "id": r["id"], "node_id": str(r["node_id"]), "software_name": r["software_name"],
                "software_version": r["software_version"], "cve_id": r["cve_id"], "status": r["status"],
                "fix_method": r["fix_method"], "exit_code": r["exit_code"], "error_message": r["error_message"],
                "created_at": r["created_at"].isoformat() if r["created_at"] else None,
                "completed_at": r["completed_at"].isoformat() if r["completed_at"] else None
            } for r in recent],
            "recentJobs": [{
                "id": str(r["id"]), "nodeId": str(r["node_id"]), "softwareName": r["software_name"],
                "cveId": r["cve_id"], "status": r["status"],
                "createdAt": str(r["created_at"]), "completedAt": str(r["completed_at"]) if r["completed_at"] else None
            } for r in recent]
        }


# --- Agent Endpoint for Remediation Jobs ---


# --- Dashboard Summary ---


# ============================================================================
# E16: Live View - Real-time Node Monitoring
# ============================================================================

import asyncio
from datetime import datetime as dt

# Store active live sessions
live_sessions: Dict[str, Dict] = {}

# Cache for live network data (node_id -> network data)
live_network_cache: Dict[str, Dict] = {}
live_agent_logs_cache: Dict[str, Dict] = {}  # Cache for agent service logs
live_eventlog_cache: Dict[str, Dict] = {}  # Cache for Windows Event Logs (System, Security, Application)

async def live_data_generator(node_id: str, session_id: str, pool):
    """Generator for SSE live data stream"""
    global live_sessions
    
    # Initial connection event
    yield f"event: connected\ndata: {json.dumps({'nodeId': node_id, 'sessionId': session_id})}\n\n"
    
    last_metrics_time = 0
    last_processes_time = 0
    last_logs_time = 0
    last_network_time = 0
    last_log_id = 0  # Track last seen log ID
    
    
    try:
        while session_id in live_sessions:
            now = time.time()
            
            # Send metrics every 2 seconds
            if now - last_metrics_time >= 2:
                async with db_pool.acquire() as conn:
                    # Get latest metrics from timescale
                    metrics = await conn.fetchrow("""
                        SELECT cpu_percent, ram_percent, disk_percent,
                               network_in_mb, network_out_mb, time as timestamp
                        FROM node_metrics
                        WHERE node_id = (SELECT id FROM nodes WHERE node_id = $1 OR id::text = $1)
                        ORDER BY time DESC LIMIT 1
                    """, node_id)
                    
                    if metrics:
                        data = {
                            "type": "metrics",
                            "data": {
                                "cpu": float(metrics['cpu_percent']) if metrics['cpu_percent'] else 0,
                                "memory": float(metrics['ram_percent']) if metrics['ram_percent'] else 0,
                                "disk": float(metrics['disk_percent']) if metrics['disk_percent'] else 0,
                                "netIn": float(metrics['network_in_mb']) if metrics['network_in_mb'] else None,
                                "netOut": float(metrics['network_out_mb']) if metrics['network_out_mb'] else None,
                                "timestamp": metrics['timestamp'].isoformat() if metrics['timestamp'] else None
                            }
                        }
                        yield f"event: metrics\ndata: {json.dumps(data)}\n\n"
                
                last_metrics_time = now
            
            # Send processes every 5 seconds
            if now - last_processes_time >= 5:
                async with db_pool.acquire() as conn:
                    procs = await conn.fetch("""
                        SELECT process_name, pid, cpu_percent, memory_mb, user_name
                        FROM node_processes
                        WHERE node_id = (SELECT id FROM nodes WHERE node_id = $1 OR id::text = $1)
                          AND collected_at > NOW() - INTERVAL '30 seconds'
                        ORDER BY cpu_percent DESC NULLS LAST LIMIT 20
                    """, node_id)
                    
                    if procs:
                        data = {
                            "type": "processes",
                            "data": [
                                {
                                    "name": p['process_name'],
                                    "pid": p['pid'],
                                    "cpu": float(p['cpu_percent']) if p['cpu_percent'] else 0,
                                    "memoryMb": float(p['memory_mb']) if p['memory_mb'] else 0,
                                    "user": p['user_name']
                                }
                                for p in procs
                            ]
                        }
                        yield f"event: processes\ndata: {json.dumps(data)}\n\n"
                
                last_processes_time = now
            
            # Send new logs every 3 seconds
            if now - last_logs_time >= 3:
                async with db_pool.acquire() as conn:
                    # Get new logs since last check
                    if last_log_id == 0:
                        # First fetch - get last 50 logs
                        logs = await conn.fetch("""
                            SELECT id, log_name, event_id, level, level_name, source, 
                                   LEFT(message, 500) as message, event_time
                            FROM eventlog_entries
                            WHERE node_id = (SELECT id::text FROM nodes WHERE node_id = $1 OR id::text = $1)
                            ORDER BY id DESC LIMIT 50
                        """, node_id)
                        logs = list(reversed(logs))  # Oldest first
                    else:
                        # Incremental - only new logs
                        logs = await conn.fetch("""
                            SELECT id, log_name, event_id, level, level_name, source,
                                   LEFT(message, 500) as message, event_time
                            FROM eventlog_entries
                            WHERE node_id = (SELECT id::text FROM nodes WHERE node_id = $1 OR id::text = $1) AND id > $2
                            ORDER BY id ASC LIMIT 100
                        """, node_id, last_log_id)
                    
                    if logs:
                        last_log_id = logs[-1]['id']
                        data = {
                            "type": "logs",
                            "data": [
                                {
                                    "id": log['id'],
                                    "logName": log['log_name'],
                                    "eventId": int(log['event_id']) if log['event_id'] else 0,
                                    "level": int(log['level']) if log['level'] else 0,
                                    "levelName": log['level_name'],
                                    "source": log['source'],
                                    "message": log['message'],
                                    "timestamp": log['event_time'].isoformat() if log['event_time'] else None
                                }
                                for log in logs
                            ]
                        }
                        yield f"event: logs\ndata: {json.dumps(data)}\n\n"
                
                last_logs_time = now
            
            # Send network data every 5 seconds (from cache)
            # NOTE: Check BEFORE updating last_processes_time below
            if node_id in live_network_cache and now - last_network_time >= 5:
                net_data = live_network_cache[node_id]
                data = {
                    "type": "network",
                    "data": net_data
                }
                yield f"event: network\ndata: {json.dumps(data)}\n\n"
                last_network_time = now
            
            # Send agent logs every 5 seconds (from cache)
            if node_id in live_agent_logs_cache and now - last_network_time >= 5:
                agent_logs_data = live_agent_logs_cache[node_id]
                data = {
                    "type": "agentLogs",
                    "data": agent_logs_data
                }
                yield f"event: agentLogs\ndata: {json.dumps(data)}\n\n"
            
            # Send Windows Event Logs every 5 seconds (from cache)
            if node_id in live_eventlog_cache and now - last_network_time >= 5:
                eventlog_data = live_eventlog_cache[node_id]
                data = {
                    "type": "eventLogs",
                    "data": eventlog_data
                }
                yield f"event: eventLogs\ndata: {json.dumps(data)}\n\n"
            
            # Heartbeat every 10 seconds
            yield f"event: heartbeat\ndata: {json.dumps({'ts': int(now * 1000)})}\n\n"
            
            await asyncio.sleep(1)
            
    except asyncio.CancelledError:
        pass
    except Exception as e:
        pass
    finally:
        if session_id in live_sessions:
            del live_sessions[session_id]
        yield f"event: disconnected\ndata: {json.dumps({'reason': 'session_ended'})}\n\n"


@app.get("/api/v1/live/{node_id}")
async def live_stream(node_id: str, _: str = Depends(verify_api_key_or_query)):
    """
    SSE endpoint for live node data streaming.
    
    Returns Server-Sent Events with:
    - metrics: CPU, RAM, Disk, Network (every 2s)
    - processes: Top 20 processes by CPU (every 5s)
    - heartbeat: Connection keepalive (every 10s)
    """
    # Verify node exists
    async with db_pool.acquire() as conn:
        node = await conn.fetchrow("SELECT id FROM nodes WHERE node_id = $1 OR id::text = $1", node_id)
        if not node:
            raise not_found("Node", node_id)
    
    # Create session
    session_id = str(uuid.uuid4())
    live_sessions[session_id] = {
        "node_id": node_id,
        "started_at": dt.utcnow(),
        "last_activity": dt.utcnow()
    }
    
    return StreamingResponse(
        live_data_generator(node_id, session_id, db_pool),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no"
        }
    )


@app.get("/api/v1/live-sessions")
async def list_live_sessions(_: str = Depends(verify_api_key)):
    """List active live monitoring sessions"""
    return {
        "sessions": [
            {
                "sessionId": sid,
                "nodeId": data["node_id"],
                "startedAt": data["started_at"].isoformat(),
                "lastActivity": data["last_activity"].isoformat()
            }
            for sid, data in live_sessions.items()
        ],
        "count": len(live_sessions)
    }


@app.delete("/api/v1/live/{session_id}")
async def stop_live_session(session_id: str, _: str = Depends(verify_api_key)):
    """Stop a live monitoring session"""
    if session_id in live_sessions:
        del live_sessions[session_id]
        return {"status": "stopped", "sessionId": session_id}
    raise HTTPException(status_code=404, detail="Session not found")


@app.post("/api/v1/live-data")
async def receive_live_data(data: Dict[str, Any], db: asyncpg.Pool = Depends(get_db)):
    """
    Receives live monitoring data from agents.
    Stores metrics in node_metrics and processes in node_processes.
    """
    node_id_text = data.get("nodeId")
    if not node_id_text:
        raise bad_request("nodeId is required", "nodeId")
    
    # Debug: log network data presence
    network_data = data.get("network", [])
    if network_data:
        logger.info(f"[LIVE-DATA] {node_id_text}: {len(network_data)} network interfaces received")
    
    async with db.acquire() as conn:
        # Get node UUID
        node = await conn.fetchrow("SELECT id FROM nodes WHERE node_id = $1 OR id::text = $1", node_id_text)
        if not node:
            raise not_found("Node", node_id_text)
        
        node_uuid = node['id']
        now = dt.utcnow()
        
        # Update last_seen - live-data proves the node is alive
        await conn.execute("UPDATE nodes SET last_seen = NOW() WHERE id = $1", node_uuid)
        
        # Store metrics
        metrics = data.get("metrics", {})
        if metrics:
            await conn.execute("""
                INSERT INTO node_metrics (time, node_id, cpu_percent, ram_percent, disk_percent)
                VALUES ($1, $2, $3, $4, $5)
            """, now, node_uuid, 
                metrics.get("cpuPercent"),
                metrics.get("memoryPercent"),
                metrics.get("diskPercent")
            )
        
        # Store processes (replace old data)
        processes = data.get("processes", [])
        if processes:
            # Delete old process data for this node
            await conn.execute(
                "DELETE FROM node_processes WHERE node_id = $1 AND collected_at < NOW() - INTERVAL '30 seconds'",
                node_uuid
            )
            
            # Insert new process data
            for proc in processes:
                await conn.execute("""
                    INSERT INTO node_processes (node_id, process_name, pid, cpu_percent, memory_mb, user_name, collected_at)
                    VALUES ($1, $2, $3, $4, $5, $6, $7)
                """, node_uuid, 
                    proc.get("name", "unknown"),
                    proc.get("pid", 0),
                    proc.get("cpuPercent"),
                    proc.get("memoryMb"),
                    proc.get("userName"),
                    now
                )
        
        # Cache network data for SSE streaming
        network = data.get("network", [])
        if network:
            live_network_cache[node_id_text] = {
                "timestamp": now.isoformat(),
                "interfaces": network
            }
        
        # Cache agent logs for SSE streaming
        agent_logs = data.get("agentLogs", [])
        if agent_logs:
            live_agent_logs_cache[node_id_text] = {
                "timestamp": now.isoformat(),
                "logs": agent_logs
            }
        
        # Cache Windows Event Logs for SSE streaming
        event_logs = data.get("eventLogs", [])
        if event_logs:
            live_eventlog_cache[node_id_text] = {
                "timestamp": now.isoformat(),
                "logs": event_logs
            }
        
        return {
            "status": "ok",
            "metricsStored": bool(metrics),
            "processesStored": len(processes),
            "networkCached": len(network),
            "agentLogsCached": len(agent_logs)
        }


@app.get("/api/v1/nodes/{node_id}/metrics/history", dependencies=[Depends(verify_api_key)])
async def get_metrics_history(
    node_id: str, 
    hours: int = 24,
    interval: str = "5m",
    _: str = Depends(verify_api_key),
    db: asyncpg.Pool = Depends(get_db)
):
    """
    Get historical metrics for a node with time bucketing.
    
    Args:
        node_id: Node identifier
        hours: How many hours back (default 24)
        interval: Bucket size - 1m, 5m, 15m, 1h (default 5m)
    
    Returns time-series data for charts.
    """
    # Map interval to TimescaleDB time_bucket
    interval_map = {
        "1m": "1 minute",
        "5m": "5 minutes",
        "15m": "15 minutes",
        "30m": "30 minutes",
        "1h": "1 hour",
        "6h": "6 hours",
        "1d": "1 day"
    }
    bucket = interval_map.get(interval, "5 minutes")
    
    async with db.acquire() as conn:
        # Get node UUID
        node = await conn.fetchrow("SELECT id FROM nodes WHERE node_id = $1 OR id::text = $1", node_id)
        if not node:
            raise not_found("Node", node_id)
        
        node_uuid = node['id']
        
        # Use time_bucket for aggregation (TimescaleDB)
        rows = await conn.fetch(f"""
            SELECT 
                time_bucket('{bucket}', time) as bucket,
                AVG(cpu_percent) as cpu,
                AVG(ram_percent) as memory,
                AVG(disk_percent) as disk,
                AVG(network_in_mb) as net_in,
                AVG(network_out_mb) as net_out
            FROM node_metrics
            WHERE node_id = $1 
              AND time > NOW() - INTERVAL '{hours} hours'
            GROUP BY bucket
            ORDER BY bucket ASC
        """, node_uuid)
        
        return {
            "nodeId": node_id,
            "hours": hours,
            "interval": interval,
            "dataPoints": len(rows),
            "data": [
                {
                    "timestamp": row['bucket'].isoformat(),
                    "cpu": round(row['cpu'], 1) if row['cpu'] else None,
                    "memory": round(row['memory'], 1) if row['memory'] else None,
                    "disk": round(row['disk'], 1) if row['disk'] else None,
                    "netIn": round(row['net_in'], 2) if row['net_in'] else None,
                    "netOut": round(row['net_out'], 2) if row['net_out'] else None
                }
                for row in rows
            ]
        }


# ============================================================================
# E15-06: Hardware Export Endpoints
# ============================================================================


# =============================================================================
# E17: Screen Mirroring Endpoints
# =============================================================================

from screen_session import screen_session_manager, ScreenSessionState

@app.on_event("startup")
async def start_screen_manager():
    await screen_session_manager.start()

@app.on_event("shutdown")
async def stop_screen_manager():
    await screen_session_manager.stop()


# =============================================================================
# REMOTE SHELL (E19)
# =============================================================================

from shell_session import shell_session_manager, ShellSessionState

@app.on_event("startup")
async def start_shell_manager():
    await shell_session_manager.start()

@app.on_event("shutdown")
async def stop_shell_manager():
    await shell_session_manager.stop()


# === Remediation Live SSE ===


# ============================================================================
# E18: Service Orchestration API
# ============================================================================

# --- Service Classes ---


# --- Services ---


# --- Service Node Assignments ---


# --- Service Reconciliation Log ---


# ============================================================================
# E18-03: Service Reconciliation Engine
# ============================================================================


# ============================================================================
# E19: Alert System
# ============================================================================

import aiohttp

async def send_discord_webhook(webhook_url: str, embed: dict) -> bool:
    """Send a Discord webhook message."""
    try:
        async with aiohttp.ClientSession() as session:
            payload = {"embeds": [embed]}
            async with session.post(webhook_url, json=payload) as resp:
                return resp.status in (200, 204)
    except Exception as e:
        print(f"Discord webhook error: {e}")
        return False

async def trigger_alert(event_type: str, event_data: dict):
    """Check alert rules and send notifications."""
    async with db_pool.acquire() as conn:
        # Find matching enabled rules
        rules = await conn.fetch("""
            SELECT r.*, c.channel_type, c.config, c.name as channel_name
            FROM alert_rules r
            JOIN alert_channels c ON r.channel_id = c.id
            WHERE r.event_type = $1 
            AND r.enabled = true 
            AND c.enabled = true
        """, event_type)
        
        for rule in rules:
            # Check cooldown
            last_alert = await conn.fetchval("""
                SELECT created_at FROM alert_history
                WHERE rule_id = $1 AND status = 'sent'
                ORDER BY created_at DESC LIMIT 1
            """, rule['id'])
            
            if last_alert:
                from datetime import timedelta
                cooldown = timedelta(minutes=rule['cooldown_minutes'])
                if datetime.utcnow().replace(tzinfo=None) - last_alert.replace(tzinfo=None) < cooldown:
                    # Throttled
                    await conn.execute("""
                        INSERT INTO alert_history (rule_id, channel_id, event_type, event_data, status)
                        VALUES ($1, $2, $3, $4, 'throttled')
                    """, rule['id'], rule['channel_id'], event_type, json.dumps(event_data))
                    continue
            
            # Send alert based on channel type
            success = False
            error_msg = None
            
            if rule['channel_type'] == 'discord':
                webhook_url = rule['config'].get('webhook_url')
                if webhook_url:
                    # Build Discord embed
                    color = {
                        'node_offline': 0xFF0000,  # Red
                        'node_online': 0x00FF00,   # Green
                        'job_failed': 0xFF6600,    # Orange
                        'job_success': 0x00FF00,   # Green
                        'disk_warning': 0xFFFF00,  # Yellow
                        'vulnerability_critical': 0xFF0000,  # Red
                    }.get(event_type, 0x0099FF)
                    
                    embed = {
                        "title": f"🔔 {event_type.replace('_', ' ').title()}",
                        "description": event_data.get('message', str(event_data)),
                        "color": color,
                        "timestamp": datetime.utcnow().isoformat() + "Z",
                        "footer": {"text": "Octofleet Inventory"},
                        "fields": []
                    }
                    
                    # Add fields from event_data
                    for key, value in event_data.items():
                        if key != 'message' and value:
                            embed["fields"].append({
                                "name": key.replace('_', ' ').title(),
                                "value": str(value)[:1024],
                                "inline": True
                            })
                    
                    success = await send_discord_webhook(webhook_url, embed)
                    if not success:
                        error_msg = "Webhook request failed"
            
            # Log alert
            await conn.execute("""
                INSERT INTO alert_history (rule_id, channel_id, event_type, event_data, status, error_message)
                VALUES ($1, $2, $3, $4, $5, $6)
            """, rule['id'], rule['channel_id'], event_type, json.dumps(event_data),
                'sent' if success else 'failed', error_msg)


# Alert Channels CRUD
@app.get("/api/v1/alert-channels")
async def list_alert_channels(_: str = Depends(verify_api_key)):
    """List all alert channels."""
    async with db_pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT id, name, channel_type, config, enabled, created_at, updated_at
            FROM alert_channels ORDER BY created_at DESC
        """)
        return [dict(r) for r in rows]

@app.post("/api/v1/alert-channels")
async def create_alert_channel(request: Request, _: str = Depends(verify_api_key)):
    """Create a new alert channel."""
    data = await request.json()
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow("""
            INSERT INTO alert_channels (name, channel_type, config, enabled)
            VALUES ($1, $2, $3, $4)
            RETURNING id, name, channel_type, config, enabled, created_at
        """, data['name'], data['channel_type'], json.dumps(data.get('config', {})), 
            data.get('enabled', True))
        return dict(row)

@app.delete("/api/v1/alert-channels/{channel_id}")
async def delete_alert_channel(channel_id: str, _: str = Depends(verify_api_key)):
    """Delete an alert channel."""
    async with db_pool.acquire() as conn:
        await conn.execute("DELETE FROM alert_channels WHERE id = $1", channel_id)
        return {"status": "deleted"}

# Alert Rules CRUD
@app.get("/api/v1/alert-rules")
async def list_alert_rules(_: str = Depends(verify_api_key)):
    """List all alert rules."""
    async with db_pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT r.*, c.name as channel_name
            FROM alert_rules r
            LEFT JOIN alert_channels c ON r.channel_id = c.id
            ORDER BY r.created_at DESC
        """)
        return [dict(r) for r in rows]

@app.post("/api/v1/alert-rules")
async def create_alert_rule(request: Request, _: str = Depends(verify_api_key)):
    """Create a new alert rule."""
    data = await request.json()
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow("""
            INSERT INTO alert_rules (name, event_type, condition, channel_id, cooldown_minutes, enabled)
            VALUES ($1, $2, $3, $4, $5, $6)
            RETURNING *
        """, data['name'], data['event_type'], json.dumps(data.get('condition', {})),
            data['channel_id'], data.get('cooldown_minutes', 15), data.get('enabled', True))
        return dict(row)

@app.delete("/api/v1/alert-rules/{rule_id}")
async def delete_alert_rule(rule_id: str, _: str = Depends(verify_api_key)):
    """Delete an alert rule."""
    async with db_pool.acquire() as conn:
        await conn.execute("DELETE FROM alert_rules WHERE id = $1", rule_id)
        return {"status": "deleted"}

# Alert History
@app.get("/api/v1/alert-history")
async def list_alert_history(limit: int = 50, _: str = Depends(verify_api_key)):
    """Get recent alert history."""
    async with db_pool.acquire() as conn:
        rows = await conn.fetch("""
            SELECT h.*, r.name as rule_name, c.name as channel_name
            FROM alert_history h
            LEFT JOIN alert_rules r ON h.rule_id = r.id
            LEFT JOIN alert_channels c ON h.channel_id = c.id
            ORDER BY h.created_at DESC
            LIMIT $1
        """, limit)
        return [dict(r) for r in rows]

# Test alert endpoint
@app.post("/api/v1/alert-channels/{channel_id}/test")
async def test_alert_channel(channel_id: str, _: str = Depends(verify_api_key)):
    """Send a test alert to a channel."""
    async with db_pool.acquire() as conn:
        channel = await conn.fetchrow(
            "SELECT * FROM alert_channels WHERE id = $1", channel_id)
        
        if not channel:
            raise HTTPException(404, "Channel not found")
        
        # Parse config if it's a string
        config = channel['config']
        if isinstance(config, str):
            config = json.loads(config)
        
        if channel['channel_type'] == 'discord':
            webhook_url = config.get('webhook_url')
            if webhook_url:
                embed = {
                    "title": "🔔 Test Alert",
                    "description": "This is a test alert from Octofleet Inventory.",
                    "color": 0x0099FF,
                    "timestamp": datetime.utcnow().isoformat() + "Z",
                    "footer": {"text": "Octofleet Inventory"}
                }
                success = await send_discord_webhook(webhook_url, embed)
                return {"status": "sent" if success else "failed"}
        
        return {"status": "unsupported_channel_type"}

# ============================================================================
# Node Health Monitor - Background Task
# ============================================================================

import asyncio
from datetime import timedelta

# Track last known status to detect changes
_node_status_cache: dict = {}

async def check_node_health_and_alert():
    """Background task to check node health and trigger alerts."""
    global _node_status_cache
    
    try:
        async with db_pool.acquire() as conn:
            rows = await conn.fetch("""
                SELECT node_id, hostname, last_seen, is_online
                FROM nodes
            """)
            
            for row in rows:
                node_id = row['node_id']
                hostname = row['hostname']
                last_seen = row['last_seen']
                
                # Calculate current status
                if last_seen:
                    diff = datetime.utcnow() - last_seen.replace(tzinfo=None)
                    if diff < timedelta(minutes=5):
                        current_status = "online"
                    elif diff < timedelta(minutes=60):
                        current_status = "away"
                    else:
                        current_status = "offline"
                else:
                    current_status = "offline"
                
                # Check if status changed
                prev_status = _node_status_cache.get(node_id)
                
                if prev_status and prev_status != current_status:
                    # Status changed!
                    if current_status == "offline" and prev_status in ("online", "away"):
                        # Node went offline - trigger alert
                        await trigger_alert('node_offline', {
                            'message': f"Node '{hostname}' went offline",
                            'hostname': hostname,
                            'node_id': node_id,
                            'previous_status': prev_status,
                            'last_seen': last_seen.isoformat() if last_seen else 'Never'
                        })
                    elif current_status == "online" and prev_status == "offline":
                        # Node came back online
                        await trigger_alert('node_online', {
                            'message': f"Node '{hostname}' is back online",
                            'hostname': hostname,
                            'node_id': node_id,
                            'previous_status': prev_status
                        })
                
                # Update cache
                _node_status_cache[node_id] = current_status
                
    except Exception as e:
        print(f"Node health check error: {e}")

async def node_health_monitor_task():
    """Background task that runs every 2 minutes."""
    await asyncio.sleep(30)  # Initial delay
    while True:
        await check_node_health_and_alert()
        await asyncio.sleep(120)  # Check every 2 minutes

@app.on_event("startup")
async def start_node_health_monitor():
    """Start the node health monitor on app startup."""
    asyncio.create_task(node_health_monitor_task())

# ============================================================================
# E20: Remote Terminal
# ============================================================================

# Terminal sessions storage
_terminal_sessions: dict = {}

class TerminalSession:
    def __init__(self, session_id: str, node_id: str, shell: str = "powershell"):
        self.session_id = session_id
        self.node_id = node_id
        self.shell = shell
        self.created_at = datetime.utcnow()
        self.output_buffer = []
        self.pending_commands = []
        self.connected = False


# Agent polling endpoint for terminal commands

# Agent posts output here

# WebSocket for real-time terminal (optional)


# =============================================================================
# Auto-Update Endpoint
# =============================================================================

@app.get("/api/v1/agent/version")
async def get_agent_version():
    """
    Returns the latest agent version from GitHub Releases.
    Agents poll this to check for updates.
    """
    import aiohttp
    
    cache_key = "agent_version"
    cache_ttl = 300  # 5 minutes cache
    
    # Check cache
    if cache_key in _cache:
        cached, timestamp = _cache[cache_key]
        if time.time() - timestamp < cache_ttl:
            return cached
    
    try:
        async with aiohttp.ClientSession() as session:
            url = "https://api.github.com/repos/BenediktSchackenberg/octofleet/releases/latest"
            headers = {"Accept": "application/vnd.github.v3+json"}
            
            async with session.get(url, headers=headers) as resp:
                if resp.status != 200:
                    raise HTTPException(status_code=502, detail="Failed to fetch release info")
                
                data = await resp.json()
                
                version = data["tag_name"].lstrip("v")
                
                # Find ZIP asset
                zip_asset = next((a for a in data["assets"] if a["name"].endswith(".zip")), None)
                sha_asset = next((a for a in data["assets"] if a["name"].endswith(".sha256")), None)
                
                if not zip_asset:
                    raise HTTPException(status_code=502, detail="No ZIP asset found")
                
                # Get SHA256 if available
                sha256 = None
                if sha_asset:
                    async with session.get(sha_asset["browser_download_url"]) as sha_resp:
                        if sha_resp.status == 200:
                            sha256 = (await sha_resp.text()).strip()
                
                # Keep backward-compatible keys used by existing tests/clients.
                result = {
                    "latest": version,
                    "version": version,
                    "latestVersion": version,
                    "downloadUrl": zip_asset["browser_download_url"],
                    "sha256": sha256,
                    "releaseDate": data["published_at"],
                    "releaseNotes": data.get("body", "")[:500]
                }
                
                # Cache result
                _cache[cache_key] = (result, time.time())
                
                return result
                
    except aiohttp.ClientError as e:
        logger.error(f"Failed to fetch GitHub release: {e}")
        raise HTTPException(status_code=502, detail="Failed to connect to GitHub")

# Simple cache dict
_cache: Dict[str, tuple] = {}


# =============================================================================
# Software Repository Endpoints (Epic #57)
# =============================================================================
import aiofiles
import hashlib

REPO_BASE_PATH = os.environ.get("OCTOFLEET_REPO_PATH", os.path.expanduser("~/.openclaw/repo"))
MAX_REPO_FILE_SIZE = int(os.environ.get("OCTOFLEET_MAX_FILE_SIZE", 5 * 1024 * 1024 * 1024))  # 5GB


def get_repo_storage_path(file_type: str, filename: str) -> str:
    """Get the full storage path for a file based on its type."""
    type_dirs = {'msi': 'msi', 'exe': 'exe', 'sql-cu': 'sql-cu', 'script': 'scripts', 'ps1': 'scripts', 'other': 'other'}
    subdir = type_dirs.get(file_type, 'other')
    return os.path.join(REPO_BASE_PATH, subdir, filename)


def detect_repo_file_type(filename: str) -> str:
    """Detect file type from extension."""
    ext = filename.lower().rsplit('.', 1)[-1] if '.' in filename else ''
    return {'msi': 'msi', 'exe': 'exe', 'zip': 'zip', 'ps1': 'ps1', 'cab': 'cab'}.get(ext, 'other')


async def compute_file_sha256(filepath: str) -> str:
    """Compute SHA256 hash of a file."""
    sha256_hash = hashlib.sha256()
    async with aiofiles.open(filepath, 'rb') as f:
        while chunk := await f.read(8192):
            sha256_hash.update(chunk)
    return sha256_hash.hexdigest()


# =============================================================================
# Export Endpoints (Epic #19 - Reporting & Exports)
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

def style_excel_header(ws, row=1):
    """Apply header styling to first row."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for cell in ws[row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

def auto_column_width(ws):
    """Auto-adjust column widths."""
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width


# =============================================================================
# E19: PDF Report Generation
# =============================================================================

def create_pdf_styles():
    """Create custom styles for PDF reports"""
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name='ReportTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        textColor=colors.HexColor('#1a1a2e')
    ))
    styles.add(ParagraphStyle(
        name='SectionTitle',
        parent=styles['Heading2'],
        fontSize=16,
        spaceAfter=12,
        spaceBefore=20,
        textColor=colors.HexColor('#16213e')
    ))
    styles.add(ParagraphStyle(
        name='SubSection',
        parent=styles['Heading3'],
        fontSize=12,
        spaceAfter=8,
        textColor=colors.HexColor('#0f3460')
    ))
    return styles


def create_header_footer(canvas, doc, title: str):
    """Add header and footer to PDF pages"""
    canvas.saveState()
    
    # Header
    canvas.setFont('Helvetica-Bold', 10)
    canvas.setFillColor(colors.HexColor('#1a1a2e'))
    canvas.drawString(50, A4[1] - 30, f"Octofleet - {title}")
    canvas.drawRightString(A4[0] - 50, A4[1] - 30, datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC'))
    
    # Footer
    canvas.setFont('Helvetica', 8)
    canvas.setFillColor(colors.grey)
    canvas.drawString(50, 30, "Generated by Octofleet Endpoint Management Platform")
    canvas.drawRightString(A4[0] - 50, 30, f"Page {doc.page}")
    
    canvas.restoreState()


def create_status_table(data: List[List], col_widths: List[float] = None) -> Table:
    """Create a styled table for reports"""
    table = Table(data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a2e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
    ]))
    return table


# =============================================================================
# E22: Provisioning API - Zero-Touch OS Deployment
# =============================================================================

from provisioning import generate_autounattend, generate_ipxe_script

# Provisioning Config
PROVISIONING_ANSWERS_PATH = os.getenv("PROVISIONING_ANSWERS_PATH", "/home/benedikt/.openclaw/workspace/octofleet-work/provisioning/answers")
PXE_SERVER_URL = os.getenv("PXE_SERVER_URL", "http://192.168.0.5:9080")


class ProvisioningTaskCreate(BaseModel):
    """Create a new provisioning task"""
    hostname: str = Field(..., min_length=1, max_length=63, pattern=r'^[a-zA-Z0-9\-]+$')
    mac_address: str = Field(..., pattern=r'^([0-9A-Fa-f]{2}[:-]){5}([0-9A-Fa-f]{2})$')
    
    # Network
    use_dhcp: bool = True
    ip_address: Optional[str] = None
    subnet_mask: str = "255.255.255.0"
    gateway: Optional[str] = None
    dns_servers: Optional[List[str]] = None
    
    # Credentials
    admin_password: str = Field(..., min_length=8)
    
    # Domain Join
    join_domain: bool = False
    domain_name: Optional[str] = None
    domain_ou: Optional[str] = None
    domain_user: Optional[str] = None
    domain_password: Optional[str] = None
    
    # Options
    windows_edition: str = "Windows Server 2025 SERVERDATACENTER"
    language: str = "en-US"
    timezone: str = "W. Europe Standard Time"
    
    # Post-Install
    install_octofleet_agent: bool = True
    enable_rdp: bool = True
    disable_firewall: bool = False


class ProvisioningTaskResponse(BaseModel):
    """Response for created provisioning task"""
    id: int
    hostname: str
    mac_address: str
    status: str
    ipxe_url: str
    autounattend_url: str
    created_at: datetime


@app.post("/api/v1/provisioning/tasks", response_model=ProvisioningTaskResponse, tags=["Provisioning"])
async def create_provisioning_task(
    task: ProvisioningTaskCreate,
    db: asyncpg.Connection = Depends(get_db),
    api_key: str = Depends(verify_api_key)
):
    """
    Create a new provisioning task for zero-touch Windows deployment.
    
    This generates:
    - MAC-specific iPXE boot script
    - Autounattend.xml with all configuration
    
    The target machine will automatically install Windows on next PXE boot.
    """
    # Normalize MAC address
    mac_normalized = task.mac_address.lower().replace("-", ":")
    mac_hyp = mac_normalized.replace(":", "-")
    
    # Check for existing task with same MAC
    existing = await db.fetchrow(
        "SELECT id FROM provisioning_tasks WHERE mac_address = $1 AND status NOT IN ('completed', 'failed', 'cancelled')",
        mac_normalized
    )
    if existing:
        raise conflict(f"Active provisioning task already exists for MAC {mac_normalized}")
    
    # Generate Autounattend.xml
    autounattend_xml = generate_autounattend(
        hostname=task.hostname,
        admin_password=task.admin_password,
        use_dhcp=task.use_dhcp,
        ip_address=task.ip_address,
        subnet_mask=task.subnet_mask,
        gateway=task.gateway,
        dns_servers=task.dns_servers,
        join_domain=task.join_domain,
        domain_name=task.domain_name,
        domain_ou=task.domain_ou,
        domain_user=task.domain_user,
        domain_password=task.domain_password,
        windows_edition=task.windows_edition,
        language=task.language,
        timezone=task.timezone,
        install_octofleet_agent=task.install_octofleet_agent,
        enable_rdp=task.enable_rdp,
        disable_firewall=task.disable_firewall,
        pxe_server=PXE_SERVER_URL,
    )
    
    # Generate iPXE boot script
    ipxe_script = generate_ipxe_script(
        mac_address=mac_normalized,
        hostname=task.hostname,
        pxe_server=PXE_SERVER_URL,
    )
    
    # Ensure answers directory exists
    os.makedirs(PROVISIONING_ANSWERS_PATH, exist_ok=True)
    
    # Write files
    xml_path = os.path.join(PROVISIONING_ANSWERS_PATH, f"{mac_hyp}.xml")
    ipxe_path = os.path.join(PROVISIONING_ANSWERS_PATH, f"{mac_hyp}.ipxe")
    
    with open(xml_path, "w") as f:
        f.write(autounattend_xml)
    
    with open(ipxe_path, "w") as f:
        f.write(ipxe_script)
    
    # Insert into database
    row = await db.fetchrow("""
        INSERT INTO provisioning_tasks (
            hostname, mac_address, status, config,
            created_at, updated_at
        ) VALUES (
            $1, $2, 'pending', $3::jsonb,
            now(), now()
        )
        RETURNING id, hostname, mac_address, status, created_at
    """, task.hostname, mac_normalized, json.dumps({
        "use_dhcp": task.use_dhcp,
        "ip_address": task.ip_address,
        "join_domain": task.join_domain,
        "domain_name": task.domain_name,
        "install_agent": task.install_octofleet_agent,
        "enable_rdp": task.enable_rdp,
    }))
    
    return ProvisioningTaskResponse(
        id=row["id"],
        hostname=row["hostname"],
        mac_address=row["mac_address"],
        status=row["status"],
        ipxe_url=f"{PXE_SERVER_URL}/boot/{mac_hyp}.ipxe",
        autounattend_url=f"{PXE_SERVER_URL}/answers/{mac_hyp}.xml",
        created_at=row["created_at"],
    )


@app.get("/api/v1/provisioning/tasks", tags=["Provisioning"])
async def list_provisioning_tasks(
    status: Optional[str] = None,
    limit: int = 50,
    db: asyncpg.Connection = Depends(get_db),
    api_key: str = Depends(verify_api_key)
):
    """List all provisioning tasks"""
    query = "SELECT * FROM provisioning_tasks"
    params = []
    
    if status:
        query += " WHERE status = $1"
        params.append(status)
    
    query += " ORDER BY created_at DESC LIMIT $" + str(len(params) + 1)
    params.append(limit)
    
    rows = await db.fetch(query, *params)
    return [dict(r) for r in rows]


@app.get("/api/v1/provisioning/tasks/{task_id}", tags=["Provisioning"])
async def get_provisioning_task(
    task_id: int,
    db: asyncpg.Connection = Depends(get_db),
    api_key: str = Depends(verify_api_key)
):
    """Get provisioning task details"""
    row = await db.fetchrow("SELECT * FROM provisioning_tasks WHERE id = $1", task_id)
    if not row:
        raise not_found("Provisioning task not found")
    return dict(row)


@app.delete("/api/v1/provisioning/tasks/{task_id}", tags=["Provisioning"])
async def cancel_provisioning_task(
    task_id: int,
    db: asyncpg.Connection = Depends(get_db),
    api_key: str = Depends(verify_api_key)
):
    """Cancel a provisioning task and remove boot files"""
    row = await db.fetchrow("SELECT mac_address, status FROM provisioning_tasks WHERE id = $1", task_id)
    if not row:
        raise not_found("Provisioning task not found")
    
    if row["status"] in ("completed", "failed"):
        raise bad_request(f"Cannot cancel task in status '{row['status']}'")
    
    mac_hyp = row["mac_address"].replace(":", "-")
    
    # Remove boot files
    xml_path = os.path.join(PROVISIONING_ANSWERS_PATH, f"{mac_hyp}.xml")
    ipxe_path = os.path.join(PROVISIONING_ANSWERS_PATH, f"{mac_hyp}.ipxe")
    
    for path in [xml_path, ipxe_path]:
        if os.path.exists(path):
            os.remove(path)
    
    # Update status
    await db.execute(
        "UPDATE provisioning_tasks SET status = 'cancelled', updated_at = now() WHERE id = $1",
        task_id
    )
    
    return {"message": "Provisioning task cancelled", "id": task_id}


@app.get("/api/v1/provisioning/images", tags=["Provisioning"])
async def list_available_images(api_key: str = Depends(verify_api_key)):
    """List available OS images for provisioning"""
    images_path = os.getenv("PROVISIONING_IMAGES_PATH", "/home/benedikt/.openclaw/workspace/octofleet-work/provisioning/images")
    
    images = []
    
    # Check win2025
    win2025_wim = os.path.join(images_path, "win2025", "install.wim")
    if os.path.exists(win2025_wim):
        size = os.path.getsize(win2025_wim)
        images.append({
            "id": "win2025",
            "name": "Windows Server 2025",
            "editions": [
                "Windows Server 2025 SERVERSTANDARDCORE",
                "Windows Server 2025 SERVERSTANDARD",
                "Windows Server 2025 SERVERDATACENTERCORE",
                "Windows Server 2025 SERVERDATACENTER",
            ],
            "size_bytes": size,
            "size_human": f"{size / 1024 / 1024 / 1024:.1f} GB",
        })
    
    # Check winpe
    winpe_dir = os.path.join(images_path, "winpe")
    if os.path.isdir(winpe_dir):
        winpe_files = os.listdir(winpe_dir)
        required = {"wimboot", "BCD", "boot.sdi", "boot.wim"}
        if required.issubset(set(winpe_files)):
            images.append({
                "id": "winpe",
                "name": "WinPE Boot Environment",
                "status": "ready",
                "files": winpe_files,
            })
    
    return {"images": images}


@app.get("/api/v1/provisioning/windows-editions", tags=["Provisioning"])
async def list_windows_editions(api_key: str = Depends(verify_api_key)):
    """List available Windows editions"""
    return {
        "editions": [
            {"id": "Windows Server 2025 SERVERSTANDARDCORE", "name": "Windows Server 2025 Standard (Core)"},
            {"id": "Windows Server 2025 SERVERSTANDARD", "name": "Windows Server 2025 Standard (Desktop)"},
            {"id": "Windows Server 2025 SERVERDATACENTERCORE", "name": "Windows Server 2025 Datacenter (Core)"},
            {"id": "Windows Server 2025 SERVERDATACENTER", "name": "Windows Server 2025 Datacenter (Desktop)"},
        ]
    }


# ============================================
# PXE API - Dynamic iPXE Script Generation (#72)
# ============================================

@app.get("/api/v1/pxe/{mac}", tags=["PXE"])
async def get_pxe_script(mac: str, db: asyncpg.Pool = Depends(get_db)):
    """
    Generate iPXE script for a specific MAC address.
    Called by iPXE bootloader during PXE boot.
    """
    mac = mac.upper().replace("-", ":").replace(".", ":")
    
    # Find pending task for this MAC
    task = await db.fetchrow("""
        SELECT t.*, i.wim_path, i.wim_index, i.os_type, p.ipxe_template
        FROM provisioning_tasks t
        JOIN provisioning_images i ON t.image_id = i.id
        JOIN provisioning_templates p ON t.platform::text = p.platform::text
        WHERE t.mac_address = $1 AND t.status = 'pending'
    """, mac)
    
    pxe_server = os.environ.get("PXE_SERVER", "http://192.168.0.5:9080")
    api_server = os.environ.get("API_SERVER", "http://192.168.0.5:8080")
    
    if not task:
        # No task - return fallback/info screen
        from fastapi.responses import PlainTextResponse
        fallback_script = f"""#!ipxe
# Octofleet PXE - No provisioning task for this MAC
# MAC: {mac}

echo
echo ========================================
echo   Octofleet PXE Boot
echo ========================================
echo
echo MAC Address: {mac}
echo No provisioning task found.
echo
echo To provision this machine:
echo   1. Go to Octofleet Web UI
echo   2. Create a new provisioning task
echo   3. Enter this MAC address
echo   4. Reboot this machine
echo
echo Booting from local disk in 30 seconds...
sleep 30
exit
"""
        return PlainTextResponse(content=fallback_script)
    
    # Mark task as booting
    await db.execute("""
        UPDATE provisioning_tasks 
        SET status = 'booting', started_at = NOW(), status_message = 'iPXE script delivered'
        WHERE mac_address = $1
    """, mac)
    
    # Get template and substitute variables
    script = task["ipxe_template"]
    
    # Variable substitutions
    substitutions = {
        "${PXE_SERVER}": pxe_server,
        "${API_SERVER}": api_server,
        "${MAC}": mac.lower().replace(":", "-"),
        "${MAC_COLON}": mac,
        "${IMAGE_PATH}": task["wim_path"],
        "${IMAGE_INDEX}": str(task["wim_index"]),
        "${TASK_ID}": str(task["id"]),
        "${HOSTNAME}": task["hostname"] or "localhost",
        "${OS_TYPE}": task.get("os_type") or "windows",
    }
    
    for var, value in substitutions.items():
        script = script.replace(var, value)
    
    from fastapi.responses import PlainTextResponse
    return PlainTextResponse(content=script)


@app.get("/api/v1/pxe", tags=["PXE"])
async def get_default_pxe():
    """Default iPXE script - chainload to MAC-specific script"""
    pxe_server = os.environ.get("PXE_SERVER", "http://192.168.0.5:9080")
    api_server = os.environ.get("API_SERVER", "http://192.168.0.5:8080")
    
    from fastapi.responses import PlainTextResponse
    script = f"""#!ipxe
# Octofleet PXE Bootloader
# Chainloads to MAC-specific provisioning script

echo Octofleet PXE Boot
echo MAC: ${{net0/mac}}

# Fetch MAC-specific script from API
chain {api_server}/api/v1/pxe/${{net0/mac}} || goto fallback

:fallback
echo
echo Failed to fetch provisioning script.
echo Retrying in 10 seconds...
sleep 10
goto start
"""
    return PlainTextResponse(content=script)


# ============================================
# Status Callbacks API (#71)
# ============================================

class TaskEventCreate(BaseModel):
    """Status callback from provisioning client"""
    event: str = Field(..., description="Event type: booted, downloading, partitioning, applying, configuring, firstboot, done, failed")
    message: Optional[str] = None
    progress: Optional[int] = Field(None, ge=0, le=100)
    error: Optional[str] = None


@app.post("/api/v1/provisioning/tasks/{task_id}/events", tags=["Provisioning"])
async def create_task_event(task_id: str, event: TaskEventCreate, db: asyncpg.Pool = Depends(get_db)):
    """
    Receive status callback from provisioning client.
    Called by WinPE/installer scripts during deployment.
    """
    # Verify task exists
    task = await db.fetchrow("SELECT id, status FROM provisioning_tasks WHERE id = $1", task_id)
    if not task:
        raise not_found("Task not found")
    
    # Map event to task status
    event_to_status = {
        "booted": "booting",
        "downloading": "installing",
        "partitioning": "installing",
        "applying": "installing",
        "configuring": "installing",
        "firstboot": "installing",
        "done": "completed",
        "failed": "failed",
    }
    
    # Map event to default progress
    event_to_progress = {
        "booted": 5,
        "downloading": 15,
        "partitioning": 25,
        "applying": 50,
        "configuring": 85,
        "firstboot": 95,
        "done": 100,
        "failed": None,
    }
    
    new_status = event_to_status.get(event.event, str(task["status"]))
    progress = event.progress if event.progress is not None else event_to_progress.get(event.event)
    
    # Create event record
    event_id = str(uuid.uuid4())
    await db.execute("""
        INSERT INTO provisioning_task_events (id, task_id, event, message, progress)
        VALUES ($1, $2, $3, $4, $5)
    """, event_id, task_id, event.event, event.message or event.error, progress)
    
    # Update task status and progress
    update_parts = ["status = $2", "status_message = $3"]
    update_params = [task_id, new_status, event.message or event.error or event.event]
    
    if progress is not None:
        update_parts.append(f"progress_percent = ${len(update_params) + 1}")
        update_params.append(progress)
    
    if event.event in ("done", "failed"):
        update_parts.append("completed_at = NOW()")
    elif event.event == "booted" and str(task["status"]) == "pending":
        update_parts.append("started_at = NOW()")
    
    await db.execute(f"""
        UPDATE provisioning_tasks 
        SET {", ".join(update_parts)}, updated_at = NOW()
        WHERE id = $1
    """, *update_params)
    
    # Fetch created event
    row = await db.fetchrow("""
        SELECT id, task_id, event, message, progress, created_at
        FROM provisioning_task_events WHERE id = $1
    """, event_id)
    
    return {
        "id": str(row["id"]),
        "task_id": str(row["task_id"]),
        "event": row["event"],
        "message": row.get("message"),
        "progress": row.get("progress"),
        "created_at": row["created_at"].isoformat() if row["created_at"] else None,
    }


@app.get("/api/v1/provisioning/tasks/{task_id}/events", tags=["Provisioning"])
async def list_task_events(task_id: str, db: asyncpg.Pool = Depends(get_db)):
    """Get all events/callbacks for a task"""
    # Verify task exists
    task = await db.fetchrow("SELECT id FROM provisioning_tasks WHERE id = $1", task_id)
    if not task:
        raise not_found("Task not found")
    
    rows = await db.fetch("""
        SELECT id, task_id, event, message, progress, created_at
        FROM provisioning_task_events
        WHERE task_id = $1
        ORDER BY created_at ASC
    """, task_id)
    
    return [{
        "id": str(row["id"]),
        "task_id": str(row["task_id"]),
        "event": row["event"],
        "message": row.get("message"),
        "progress": row.get("progress"),
        "created_at": row["created_at"].isoformat() if row["created_at"] else None,
    } for row in rows]


# ============================================================================
# E20: PDF Report Generation
# ============================================================================

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')  # Non-interactive backend


def create_pie_chart(data: dict, title: str, filename: str) -> str:
    """Create a pie chart and save to temp file."""
    fig, ax = plt.subplots(figsize=(4, 4))
    
    # Filter out zero values
    filtered = {k: v for k, v in data.items() if v > 0}
    if not filtered:
        filtered = {"No Data": 1}
    
    colors_map = {
        "online": "#22c55e",
        "away": "#f59e0b", 
        "offline": "#ef4444",
        "critical": "#dc2626",
        "high": "#f97316",
        "medium": "#eab308",
        "low": "#22c55e",
        "success": "#22c55e",
        "failed": "#ef4444",
        "running": "#3b82f6",
        "pending": "#6b7280"
    }
    
    chart_colors = [colors_map.get(k.lower(), "#6b7280") for k in filtered.keys()]
    
    wedges, texts, autotexts = ax.pie(
        filtered.values(), 
        labels=filtered.keys(),
        autopct='%1.1f%%',
        colors=chart_colors,
        startangle=90
    )
    ax.set_title(title, fontsize=12, fontweight='bold')
    
    plt.tight_layout()
    filepath = f"/tmp/{filename}"
    plt.savefig(filepath, dpi=150, bbox_inches='tight', facecolor='white')
    plt.close()
    return filepath


def create_bar_chart(data: dict, title: str, filename: str, xlabel: str = "", ylabel: str = "") -> str:
    """Create a bar chart and save to temp file."""
    fig, ax = plt.subplots(figsize=(6, 3))
    
    bars = ax.bar(data.keys(), data.values(), color='#3b82f6')
    ax.set_title(title, fontsize=12, fontweight='bold')
    ax.set_xlabel(xlabel)
    ax.set_ylabel(ylabel)
    
    # Add value labels on bars
    for bar, val in zip(bars, data.values()):
        ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.5, 
                str(val), ha='center', va='bottom', fontsize=9)
    
    plt.tight_layout()
    filepath = f"/tmp/{filename}"
    plt.savefig(filepath, dpi=150, bbox_inches='tight', facecolor='white')
    plt.close()
    return filepath


# ============================================
# E19-06: Systems Registry API
# ============================================

@app.get("/api/v1/discovered-systems", tags=["Provisioning"])
async def list_discovered_systems(
    status: Optional[str] = None,
    limit: int = 100,
    offset: int = 0,
    db: asyncpg.Pool = Depends(get_db)
):
    """List all discovered systems from PXE boot registry"""
    async with db.acquire() as conn:
        query = "SELECT * FROM discovered_systems WHERE 1=1"
        params = []
        param_idx = 1
        
        if status:
            query += f" AND status = ${param_idx}"
            params.append(status)
            param_idx += 1
        
        query += f" ORDER BY last_seen DESC LIMIT ${param_idx} OFFSET ${param_idx + 1}"
        params.extend([limit, offset])
        
        rows = await conn.fetch(query, *params)
        total = await conn.fetchval("SELECT COUNT(*) FROM discovered_systems" + 
                                    (f" WHERE status = $1" if status else ""),
                                    *([status] if status else []))
        
        return {
            "systems": [dict(row) for row in rows],
            "total": total,
            "limit": limit,
            "offset": offset
        }

@app.get("/api/v1/discovered-systems/{system_id}", tags=["Provisioning"])
async def get_discovered_system(system_id: str, db: asyncpg.Pool = Depends(get_db)):
    """Get a specific discovered system by ID or MAC address"""
    async with db.acquire() as conn:
        # Try by ID first, then by MAC
        row = await conn.fetchrow(
            "SELECT * FROM discovered_systems WHERE id::text = $1 OR mac_address = $1",
            system_id
        )
        if not row:
            raise HTTPException(status_code=404, detail="System not found")
        return dict(row)

@app.post("/api/v1/discovered-systems/register", tags=["Provisioning"])
async def register_discovered_system(data: Dict[str, Any], db: asyncpg.Pool = Depends(get_db)):
    """
    Register or update a discovered system (called by iPXE during boot).
    If MAC already exists, updates last_seen and increments boot_count.
    """
    mac = data.get("mac_address", "").upper().replace("-", ":").strip()
    if not mac:
        raise HTTPException(status_code=400, detail="mac_address is required")
    
    async with db.acquire() as conn:
        # Check if exists
        existing = await conn.fetchrow(
            "SELECT id, boot_count FROM discovered_systems WHERE mac_address = $1", mac
        )
        
        if existing:
            # Update existing
            await conn.execute("""
                UPDATE discovered_systems SET
                    ip_address = COALESCE($2, ip_address),
                    hostname = COALESCE($3, hostname),
                    manufacturer = COALESCE($4, manufacturer),
                    model = COALESCE($5, model),
                    serial_number = COALESCE($6, serial_number),
                    bios_version = COALESCE($7, bios_version),
                    cpu_info = COALESCE($8, cpu_info),
                    memory_mb = COALESCE($9, memory_mb),
                    disk_info = COALESCE($10::jsonb, disk_info),
                    boot_mode = COALESCE($11, boot_mode),
                    last_seen = NOW(),
                    boot_count = boot_count + 1
                WHERE mac_address = $1
            """, mac, data.get("ip_address"), data.get("hostname"),
                data.get("manufacturer"), data.get("model"), data.get("serial_number"),
                data.get("bios_version"), data.get("cpu_info"), data.get("memory_mb"),
                json.dumps(data.get("disk_info")) if data.get("disk_info") else None,
                data.get("boot_mode"))
            
            return {"status": "updated", "id": str(existing["id"]), "boot_count": existing["boot_count"] + 1}
        else:
            # Insert new
            row = await conn.fetchrow("""
                INSERT INTO discovered_systems (
                    mac_address, ip_address, hostname, manufacturer, model,
                    serial_number, bios_version, cpu_info, memory_mb, disk_info, boot_mode
                ) VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10::jsonb, $11)
                RETURNING id
            """, mac, data.get("ip_address"), data.get("hostname"),
                data.get("manufacturer"), data.get("model"), data.get("serial_number"),
                data.get("bios_version"), data.get("cpu_info"), data.get("memory_mb"),
                json.dumps(data.get("disk_info")) if data.get("disk_info") else None,
                data.get("boot_mode"))
            
            return {"status": "created", "id": str(row["id"]), "boot_count": 1}

@app.patch("/api/v1/discovered-systems/{system_id}", tags=["Provisioning"])
async def update_discovered_system(
    system_id: str, 
    data: Dict[str, Any], 
    db: asyncpg.Pool = Depends(get_db)
):
    """Update a discovered system (status, notes, tags)"""
    async with db.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT id FROM discovered_systems WHERE id::text = $1 OR mac_address = $1",
            system_id
        )
        if not row:
            raise HTTPException(status_code=404, detail="System not found")
        
        updates = []
        params = [str(row["id"])]
        param_idx = 2
        
        for field in ["status", "notes", "hostname"]:
            if field in data:
                updates.append(f"{field} = ${param_idx}")
                params.append(data[field])
                param_idx += 1
        
        if "tags" in data:
            updates.append(f"tags = ${param_idx}::text[]")
            params.append(data["tags"])
            param_idx += 1
        
        if not updates:
            raise HTTPException(status_code=400, detail="No fields to update")
        
        await conn.execute(
            f"UPDATE discovered_systems SET {', '.join(updates)} WHERE id = $1::uuid",
            *params
        )
        
        return {"status": "updated", "id": str(row["id"])}

@app.delete("/api/v1/discovered-systems/{system_id}", tags=["Provisioning"])
async def delete_discovered_system(system_id: str, db: asyncpg.Pool = Depends(get_db)):
    """Delete a discovered system from the registry"""
    async with db.acquire() as conn:
        result = await conn.execute(
            "DELETE FROM discovered_systems WHERE id::text = $1 OR mac_address = $1",
            system_id
        )
        if result == "DELETE 0":
            raise HTTPException(status_code=404, detail="System not found")
        return {"status": "deleted"}

@app.post("/api/v1/discovered-systems/{system_id}/provision", tags=["Provisioning"])
async def provision_discovered_system(
    system_id: str,
    data: Dict[str, Any],
    db: asyncpg.Pool = Depends(get_db)
):
    """
    Create a provisioning task for a discovered system.
    Requires: hostname, image_id or template
    """
    async with db.acquire() as conn:
        system = await conn.fetchrow(
            "SELECT * FROM discovered_systems WHERE id::text = $1 OR mac_address = $1",
            system_id
        )
        if not system:
            raise HTTPException(status_code=404, detail="System not found")
        
        hostname = data.get("hostname") or system["hostname"]
        if not hostname:
            raise HTTPException(status_code=400, detail="hostname is required")
        
        # Update system status
        await conn.execute(
            "UPDATE discovered_systems SET status = 'pending', hostname = $2 WHERE id = $1",
            system["id"], hostname
        )
        
        # Return info for creating provisioning task
        return {
            "status": "ready",
            "system_id": str(system["id"]),
            "mac_address": system["mac_address"],
            "hostname": hostname,
            "message": "System marked for provisioning. Create a provisioning task with this MAC address."
        }

# ============================================
# E19-09: Linux Agent Download Endpoint
# ============================================

LINUX_AGENT_SCRIPT = '''#!/bin/bash
# Octofleet Linux Agent v1.0.0
# Auto-generated from provisioning endpoint

set -e

CONFIG_FILE="${CONFIG_FILE:-/opt/octofleet-agent/config.env}"

# Load config
if [ -f "$CONFIG_FILE" ]; then
    source "$CONFIG_FILE"
fi

API_URL="${API_URL:-http://192.168.0.5:8080}"
API_KEY="${API_KEY:?API_KEY must be set}"
NODE_ID="${NODE_ID:-$(hostname)}"
POLL_INTERVAL="${POLL_INTERVAL:-30}"

log() {
    echo "[$(date '+%Y-%m-%d %H:%M:%S')] $1"
}

# Collect inventory
collect_inventory() {
    local os_name=$(lsb_release -d 2>/dev/null | cut -f2 || cat /etc/os-release | grep PRETTY_NAME | cut -d'"' -f2)
    local os_version=$(lsb_release -r 2>/dev/null | cut -f2 || cat /etc/os-release | grep VERSION_ID | cut -d'"' -f2)
    local kernel=$(uname -r)
    local hostname=$(hostname -f 2>/dev/null || hostname)
    local cpu_model=$(grep "model name" /proc/cpuinfo | head -1 | cut -d':' -f2 | xargs)
    local cpu_cores=$(nproc)
    local mem_total_kb=$(grep MemTotal /proc/meminfo | awk '{print $2}')
    local mem_total_mb=$((mem_total_kb / 1024))
    local uptime_seconds=$(cat /proc/uptime | cut -d' ' -f1 | cut -d'.' -f1)
    
    cat << EOF
{
    "nodeId": "$NODE_ID",
    "hostname": "$hostname",
    "osFamily": "Linux",
    "osName": "$os_name",
    "osVersion": "$os_version",
    "kernel": "$kernel",
    "cpuModel": "$cpu_model",
    "cpuCores": $cpu_cores,
    "memoryMb": $mem_total_mb,
    "uptimeSeconds": $uptime_seconds,
    "agentVersion": "1.0.0"
}
EOF
}

# Push inventory to backend
push_inventory() {
    local data=$(collect_inventory)
    curl -s -X POST "$API_URL/api/v1/inventory" \\
        -H "Content-Type: application/json" \\
        -H "X-API-Key: $API_KEY" \\
        -d "$data" > /dev/null 2>&1 && log "Inventory pushed" || log "Inventory push failed"
}

# Poll for jobs
poll_jobs() {
    local response=$(curl -s "$API_URL/api/v1/jobs/pending/$NODE_ID" -H "X-API-Key: $API_KEY" 2>/dev/null)
    local count=$(echo "$response" | jq -r '.count // 0' 2>/dev/null)
    
    if [ "$count" -gt 0 ]; then
        log "Found $count pending jobs"
        echo "$response" | jq -c '.jobs[]' 2>/dev/null | while read job; do
            execute_job "$job"
        done
    fi
}

# Execute a job
execute_job() {
    local job="$1"
    local instance_id=$(echo "$job" | jq -r '.instanceId')
    local command_type=$(echo "$job" | jq -r '.commandType')
    local payload=$(echo "$job" | jq -r '.commandPayload')
    
    log "Executing job $instance_id (type: $command_type)"
    
    # Mark as started
    curl -s -X POST "$API_URL/api/v1/jobs/instances/$instance_id/start" \\
        -H "X-API-Key: $API_KEY" > /dev/null 2>&1
    
    local exit_code=0
    local stdout=""
    local stderr=""
    
    case "$command_type" in
        script|run)
            local script=$(echo "$payload" | jq -r '.script // .command // ""')
            stdout=$(bash -c "$script" 2>&1) || exit_code=$?
            ;;
        *)
            stderr="Unknown command type: $command_type"
            exit_code=1
            ;;
    esac
    
    # Report result
    curl -s -X POST "$API_URL/api/v1/jobs/instances/$instance_id/result" \\
        -H "Content-Type: application/json" \\
        -H "X-API-Key: $API_KEY" \\
        -d "{\\"exitCode\\": $exit_code, \\"stdout\\": \\"$(echo "$stdout" | head -c 10000 | sed 's/"/\\\\"/g' | tr '\\n' ' ')\\", \\"stderr\\": \\"$(echo "$stderr" | sed 's/"/\\\\"/g')\\"}" > /dev/null 2>&1
    
    log "Job $instance_id completed with exit code $exit_code"
}

# Main loop
log "Octofleet Linux Agent starting..."
log "API: $API_URL, Node: $NODE_ID"

push_inventory

while true; do
    poll_jobs
    sleep "$POLL_INTERVAL"
done
'''

@app.get("/api/v1/provisioning/linux-agent", tags=["Provisioning"])
async def get_linux_agent():
    """Download the Linux agent shell script for auto-installation"""
    from fastapi.responses import PlainTextResponse
    return PlainTextResponse(
        content=LINUX_AGENT_SCRIPT,
        media_type="text/x-shellscript",
        headers={"Content-Disposition": "attachment; filename=agent.sh"}
    )

@app.get("/api/v1/provisioning/autoinstall/{hostname}", tags=["Provisioning"])
async def get_autoinstall_config(
    hostname: str,
    api_url: str = "http://192.168.0.5:8080",
    api_key: str = "",
    task_id: Optional[str] = None
):
    """Generate Ubuntu autoinstall config with Octofleet agent pre-installed"""
    
    template = """#cloud-config
autoinstall:
  version: 1
  locale: en_US.UTF-8
  keyboard:
    layout: de
  identity:
    hostname: {hostname}
    username: octofleet
    password: "$6$rounds=4096$salt$iqcgL5FX/a5Zq8e8Vz7x3FJsYJNBjrNhZNpzNgS1XHnWz/vZF8PUy3gH8rN5X3bN7qY4fJ5nM2kP9sL1hT6wK."
  ssh:
    install-server: true
    allow-pw: true
  storage:
    layout:
      name: lvm
  packages:
    - curl
    - jq
    - net-tools
  late-commands:
    - curtin in-target --target=/target -- mkdir -p /opt/octofleet-agent
    - |
      curl -fsSL "{api_url}/api/v1/provisioning/linux-agent" -o /target/opt/octofleet-agent/agent.sh
      chmod +x /target/opt/octofleet-agent/agent.sh
      cat > /target/opt/octofleet-agent/config.env << CONF
      API_URL={api_url}
      API_KEY={api_key}
      NODE_ID={hostname}
      POLL_INTERVAL=30
      CONF
      cat > /target/etc/systemd/system/octofleet-agent.service << SVC
      [Unit]
      Description=Octofleet Linux Agent
      After=network-online.target
      Wants=network-online.target
      [Service]
      Type=simple
      ExecStart=/opt/octofleet-agent/agent.sh
      Restart=always
      RestartSec=30
      WorkingDirectory=/opt/octofleet-agent
      [Install]
      WantedBy=multi-user.target
      SVC
    - curtin in-target --target=/target -- systemctl enable octofleet-agent
"""
    
    config = template.format(
        hostname=hostname,
        api_url=api_url,
        api_key=api_key
    )
    
    from fastapi.responses import PlainTextResponse
    return PlainTextResponse(content=config, media_type="text/yaml")


# ============================================================
# E21: Security Monitoring & Compliance
# ============================================================

# --- Monitoring Profiles ---


# --- Monitoring Assignments ---


# --- Event Ingest ---


# --- Events Query ---


# --- File Events Query ---


# --- Findings ---


# --- Security Policies ---

@app.get("/api/v1/security/policies", dependencies=[Depends(verify_api_key)])
async def list_security_policies():
    async with db_pool.acquire() as conn:
        rows = await conn.fetch("SELECT * FROM security_policies ORDER BY created_at DESC")
        return {"policies": [dict(r) for r in rows]}

@app.post("/api/v1/security/policies", dependencies=[Depends(verify_api_key)])
async def create_security_policy(req: Request):
    body = await req.json()
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow("""
            INSERT INTO security_policies (name, description, definition, enabled, severity, created_by)
            VALUES ($1, $2, $3::jsonb, $4, $5, $6)
            RETURNING *
        """, body["name"], body.get("description"), json.dumps(body.get("definition", {})),
            body.get("enabled", True), body.get("severity", "medium"), body.get("created_by"))
        return dict(row)

@app.put("/api/v1/security/policies/{policy_id}", dependencies=[Depends(verify_api_key)])
async def update_security_policy(policy_id: str, req: Request):
    body = await req.json()
    async with db_pool.acquire() as conn:
        row = await conn.fetchrow("""
            UPDATE security_policies SET
                name = COALESCE($2, name),
                description = COALESCE($3, description),
                definition = COALESCE($4::jsonb, definition),
                enabled = COALESCE($5, enabled),
                severity = COALESCE($6, severity),
                version = version + 1,
                updated_at = now()
            WHERE id = $1::uuid RETURNING *
        """, policy_id, body.get("name"), body.get("description"),
            json.dumps(body.get("definition")) if "definition" in body else None,
            body.get("enabled"), body.get("severity"))
        if not row:
            raise HTTPException(status_code=404, detail="Policy not found")
        return dict(row)

@app.delete("/api/v1/security/policies/{policy_id}", dependencies=[Depends(verify_api_key)])
async def delete_security_policy(policy_id: str):
    async with db_pool.acquire() as conn:
        await conn.execute("DELETE FROM security_policies WHERE id = $1::uuid", policy_id)
        return {"status": "deleted"}

# --- Retention Config ---


# --- Evidence Exports ---


# --- UI Audit Events ---


# --- Security Dashboard Stats ---

@app.get("/api/v1/security/dashboard", dependencies=[Depends(verify_api_key)])
async def security_dashboard():
    async with db_pool.acquire() as conn:
        # Findings by severity
        findings_by_severity = await conn.fetch("""
            SELECT severity, status, count(*) as count
            FROM findings
            GROUP BY severity, status
        """)
        # Recent events count (24h)
        event_count_24h = await conn.fetchval("""
            SELECT count(*) FROM events_normalized WHERE ts > now() - interval '24 hours'
        """)
        file_event_count_24h = await conn.fetchval("""
            SELECT count(*) FROM file_events WHERE ts > now() - interval '24 hours'
        """)
        # Active monitoring assignments
        active_assignments = await conn.fetchval("""
            SELECT count(*) FROM monitoring_assignments WHERE status = 'active'
        """)
        # Top event types (24h)
        top_event_types = await conn.fetch("""
            SELECT event_type, count(*) as count
            FROM events_normalized WHERE ts > now() - interval '24 hours'
            GROUP BY event_type ORDER BY count DESC LIMIT 10
        """)
        return {
            "findings_by_severity": [dict(r) for r in findings_by_severity],
            "events_24h": event_count_24h,
            "file_events_24h": file_event_count_24h,
            "active_monitoring": active_assignments,
            "top_event_types": [dict(r) for r in top_event_types]
        }


# ============================================================
# E21 Story #83: Agent Capability & Health Reporting
# ============================================================


# ============================================================
# E21 Story #86: Unified Event Schema (Normalizer)
# ============================================================


# ============================================================
# E21 Story #83: Monitoring Health Panel (per Node)
# ============================================================


# ============================================================
# E21 Story #84: Baseline Inventory & Config Posture Snapshots
# ============================================================


def _compute_posture_diff(baseline: dict, current: dict) -> dict:
    """Compute structured diff between two posture snapshots"""
    changes = []
    
    # OS info changes
    b_os = baseline.get("osInfo") or {}
    c_os = current.get("osInfo") or {}
    if isinstance(b_os, str):
        try: b_os = json.loads(b_os)
        except: b_os = {}
    if isinstance(c_os, str):
        try: c_os = json.loads(c_os)
        except: c_os = {}
    for key in set(list(b_os.keys()) + list(c_os.keys())):
        if b_os.get(key) != c_os.get(key):
            changes.append({"category": "os", "field": key, "old": b_os.get(key), "new": c_os.get(key), "severity": "medium"})
    
    # Installed packages diff
    b_pkgs = set()
    c_pkgs = set()
    b_pkg_list = baseline.get("installedPackages") or []
    c_pkg_list = current.get("installedPackages") or []
    if isinstance(b_pkg_list, str):
        try: b_pkg_list = json.loads(b_pkg_list)
        except: b_pkg_list = []
    if isinstance(c_pkg_list, str):
        try: c_pkg_list = json.loads(c_pkg_list)
        except: c_pkg_list = []
    for p in b_pkg_list:
        name = p.get("name", p) if isinstance(p, dict) else str(p)
        b_pkgs.add(name)
    for p in c_pkg_list:
        name = p.get("name", p) if isinstance(p, dict) else str(p)
        c_pkgs.add(name)
    
    for pkg in c_pkgs - b_pkgs:
        changes.append({"category": "packages", "change": "added", "name": pkg, "severity": "low"})
    for pkg in b_pkgs - c_pkgs:
        changes.append({"category": "packages", "change": "removed", "name": pkg, "severity": "medium"})
    
    # Services diff
    b_svcs = set()
    c_svcs = set()
    b_svc_list = baseline.get("runningServices") or []
    c_svc_list = current.get("runningServices") or []
    if isinstance(b_svc_list, str):
        try: b_svc_list = json.loads(b_svc_list)
        except: b_svc_list = []
    if isinstance(c_svc_list, str):
        try: c_svc_list = json.loads(c_svc_list)
        except: c_svc_list = []
    for s in b_svc_list:
        name = s.get("name", s) if isinstance(s, dict) else str(s)
        b_svcs.add(name)
    for s in c_svc_list:
        name = s.get("name", s) if isinstance(s, dict) else str(s)
        c_svcs.add(name)
    
    for svc in c_svcs - b_svcs:
        changes.append({"category": "services", "change": "started", "name": svc, "severity": "low"})
    for svc in b_svcs - c_svcs:
        changes.append({"category": "services", "change": "stopped", "name": svc, "severity": "medium"})
    
    # Config setting changes (SSH, RDP, SMB, firewall, local admins)
    b_cfg = baseline.get("configSettings") or {}
    c_cfg = current.get("configSettings") or {}
    if isinstance(b_cfg, str):
        try: b_cfg = json.loads(b_cfg)
        except: b_cfg = {}
    if isinstance(c_cfg, str):
        try: c_cfg = json.loads(c_cfg)
        except: c_cfg = {}
    
    security_critical = {"rdpEnabled", "sshEnabled", "firewallEnabled", "guestAccount", "autoLogin"}
    for key in set(list(b_cfg.keys()) + list(c_cfg.keys())):
        if b_cfg.get(key) != c_cfg.get(key):
            sev = "high" if key in security_critical else "medium"
            changes.append({"category": "config", "field": key, "old": b_cfg.get(key), "new": c_cfg.get(key), "severity": sev})
    
    # Open ports diff
    b_ports = set()
    c_ports = set()
    b_port_list = baseline.get("openPorts") or []
    c_port_list = current.get("openPorts") or []
    if isinstance(b_port_list, str):
        try: b_port_list = json.loads(b_port_list)
        except: b_port_list = []
    if isinstance(c_port_list, str):
        try: c_port_list = json.loads(c_port_list)
        except: c_port_list = []
    for p in b_port_list:
        port_key = f"{p.get('port','?')}/{p.get('protocol','tcp')}" if isinstance(p, dict) else str(p)
        b_ports.add(port_key)
    for p in c_port_list:
        port_key = f"{p.get('port','?')}/{p.get('protocol','tcp')}" if isinstance(p, dict) else str(p)
        c_ports.add(port_key)
    
    for port in c_ports - b_ports:
        changes.append({"category": "ports", "change": "opened", "port": port, "severity": "high"})
    for port in b_ports - c_ports:
        changes.append({"category": "ports", "change": "closed", "port": port, "severity": "info"})
    
    return {"changes": changes, "totalChanges": len(changes)}


# ============================================================
# E21 Story #85: Vulnerability Scan & CVE Mapping (Fleet View)
# ============================================================

@app.get("/api/v1/security/vulnerabilities/fleet")
async def fleet_vulnerability_summary(db: asyncpg.Pool = Depends(get_db)):
    """Fleet-wide vulnerability aggregation"""
    async with db.acquire() as conn:
        # By severity
        by_severity = await conn.fetch("""
            SELECT v.severity, COUNT(*) as count 
            FROM node_vulnerabilities nv
            JOIN vulnerabilities v ON v.id = nv.vulnerability_id
            GROUP BY v.severity ORDER BY 
                CASE v.severity WHEN 'critical' THEN 1 WHEN 'high' THEN 2 WHEN 'medium' THEN 3 WHEN 'low' THEN 4 ELSE 5 END
        """)
        
        # By node
        by_node = await conn.fetch("""
            SELECT nv.node_id, 
                   COUNT(*) as total,
                   COUNT(*) FILTER (WHERE v.severity = 'critical') as critical,
                   COUNT(*) FILTER (WHERE v.severity = 'high') as high
            FROM node_vulnerabilities nv
            JOIN vulnerabilities v ON v.id = nv.vulnerability_id
            GROUP BY nv.node_id ORDER BY critical DESC, high DESC, total DESC
        """)
        
        # Top CVEs across fleet
        top_cves = await conn.fetch("""
            SELECT v.cve_id, v.severity, v.software_name as package_name, 
                   COUNT(DISTINCT nv.node_id) as affected_nodes,
                   MAX(v.cvss_score) as cvss_score,
                   MAX(v.description) as description
            FROM node_vulnerabilities nv
            JOIN vulnerabilities v ON v.id = nv.vulnerability_id
            WHERE v.cve_id IS NOT NULL
            GROUP BY v.cve_id, v.severity, v.software_name
            ORDER BY 
                CASE v.severity WHEN 'critical' THEN 1 WHEN 'high' THEN 2 WHEN 'medium' THEN 3 ELSE 4 END,
                affected_nodes DESC
            LIMIT 50
        """)
        
        # By package
        by_package = await conn.fetch("""
            SELECT v.software_name as package_name, COUNT(*) as vuln_count,
                   COUNT(DISTINCT nv.node_id) as affected_nodes,
                   MAX(v.severity) as max_severity
            FROM node_vulnerabilities nv
            JOIN vulnerabilities v ON v.id = nv.vulnerability_id
            GROUP BY v.software_name
            ORDER BY vuln_count DESC LIMIT 30
        """)
        
        total = await conn.fetchval("SELECT COUNT(*) FROM node_vulnerabilities")
        total_nodes = await conn.fetchval("SELECT COUNT(DISTINCT node_id) FROM node_vulnerabilities")
        
        return {
            "total": total,
            "affectedNodes": total_nodes,
            "bySeverity": [{"severity": r["severity"], "count": r["count"]} for r in by_severity],
            "byNode": [{"nodeId": r["node_id"], "total": r["total"], "critical": r["critical"], "high": r["high"]} for r in by_node],
            "topCves": [{
                "cveId": r["cve_id"], "severity": r["severity"], "packageName": r["package_name"],
                "affectedNodes": r["affected_nodes"], "cvssScore": float(r["cvss_score"]) if r["cvss_score"] else None,
                "description": r["description"]
            } for r in top_cves],
            "byPackage": [{"packageName": r["package_name"], "vulnCount": r["vuln_count"], 
                          "affectedNodes": r["affected_nodes"], "maxSeverity": r["max_severity"]} for r in by_package]
        }

# ============================================================
# E21: Monitoring Profiles & Assignments
# ============================================================

@app.get("/api/v1/security/monitoring/profiles", dependencies=[Depends(verify_api_key)])
async def list_monitoring_profiles(db: asyncpg.Pool = Depends(get_db)):
    async with db.acquire() as conn:
        rows = await conn.fetch("SELECT * FROM monitoring_profiles ORDER BY created_at DESC")
        return {"profiles": [dict(r) for r in rows]}

@app.post("/api/v1/security/monitoring/profiles", dependencies=[Depends(verify_api_key)])
async def create_monitoring_profile(req: Request, db: asyncpg.Pool = Depends(get_db)):
    data = await req.json()
    async with db.acquire() as conn:
        row = await conn.fetchrow("""
            INSERT INTO monitoring_profiles (name, description, version, sensors, sampling, include_paths, exclude_paths, created_by)
            VALUES ($1, $2, $3, $4::jsonb, $5::jsonb, $6::jsonb, $7::jsonb, $8) RETURNING *
        """, data.get("name"), data.get("description"), data.get("version", 1),
            json.dumps(data.get("sensors", {})), json.dumps(data.get("sampling", {})),
            json.dumps(data.get("includePaths", [])), json.dumps(data.get("excludePaths", [])),
            data.get("createdBy"))
        return dict(row)

@app.put("/api/v1/security/monitoring/profiles/{profile_id}", dependencies=[Depends(verify_api_key)])
async def update_monitoring_profile(profile_id: str, req: Request, db: asyncpg.Pool = Depends(get_db)):
    data = await req.json()
    async with db.acquire() as conn:
        updates, params, idx = [], [profile_id], 2
        for field, col in [("name","name"),("description","description"),("version","version")]:
            if field in data:
                updates.append(f"{col} = ${idx}")
                params.append(data[field])
                idx += 1
        for field, col in [("sensors","sensors"),("sampling","sampling"),("includePaths","include_paths"),("excludePaths","exclude_paths")]:
            if field in data:
                updates.append(f"{col} = ${idx}::jsonb")
                params.append(json.dumps(data[field]))
                idx += 1
        if updates:
            updates.append("updated_at = NOW()")
            await conn.execute(f"UPDATE monitoring_profiles SET {', '.join(updates)} WHERE id = $1::uuid", *params)
        return {"updated": True}

@app.delete("/api/v1/security/monitoring/profiles/{profile_id}", dependencies=[Depends(verify_api_key)])
async def delete_monitoring_profile(profile_id: str, db: asyncpg.Pool = Depends(get_db)):
    async with db.acquire() as conn:
        await conn.execute("DELETE FROM monitoring_profiles WHERE id = $1::uuid", profile_id)
        return {"deleted": True}

@app.get("/api/v1/security/monitoring/assignments", dependencies=[Depends(verify_api_key)])
async def list_monitoring_assignments(db: asyncpg.Pool = Depends(get_db)):
    async with db.acquire() as conn:
        rows = await conn.fetch("""
            SELECT ma.*, mp.name as profile_name FROM monitoring_assignments ma
            LEFT JOIN monitoring_profiles mp ON mp.id = ma.profile_id
            ORDER BY ma.created_at DESC
        """)
        return {"assignments": [dict(r) for r in rows]}

@app.post("/api/v1/security/monitoring/assignments", dependencies=[Depends(verify_api_key)])
async def create_monitoring_assignment(req: Request, db: asyncpg.Pool = Depends(get_db)):
    data = await req.json()
    async with db.acquire() as conn:
        row = await conn.fetchrow("""
            INSERT INTO monitoring_assignments (target_type, target_id, profile_id, profile_version, priority, status)
            VALUES ($1, $2, $3::uuid, $4, $5, $6) RETURNING *
        """, data.get("targetType", "node"), data.get("targetId"),
            data.get("profileId"), data.get("profileVersion", 1),
            data.get("priority", 0), data.get("status", "active"))
        return dict(row)

@app.delete("/api/v1/security/monitoring/assignments/{assignment_id}", dependencies=[Depends(verify_api_key)])
async def delete_monitoring_assignment(assignment_id: str, db: asyncpg.Pool = Depends(get_db)):
    async with db.acquire() as conn:
        await conn.execute("DELETE FROM monitoring_assignments WHERE id = $1::uuid", assignment_id)
        return {"deleted": True}

# ============================================================
# E21: Posture Snapshots & Diffs
# ============================================================

@app.get("/api/v1/security/posture/snapshots", dependencies=[Depends(verify_api_key)])
async def list_posture_snapshots(node_id: str = None, limit: int = 50, db: asyncpg.Pool = Depends(get_db)):
    async with db.acquire() as conn:
        if node_id:
            rows = await conn.fetch("SELECT * FROM config_posture_snapshots WHERE node_id = $1 ORDER BY created_at DESC LIMIT $2", node_id, limit)
        else:
            rows = await conn.fetch("SELECT * FROM config_posture_snapshots ORDER BY created_at DESC LIMIT $1", limit)
        return {"snapshots": [dict(r) for r in rows]}

@app.post("/api/v1/security/posture/snapshots", dependencies=[Depends(verify_api_key)])
async def create_posture_snapshot(req: Request, db: asyncpg.Pool = Depends(get_db)):
    data = await req.json()
    async with db.acquire() as conn:
        row = await conn.fetchrow("""
            INSERT INTO config_posture_snapshots (node_id, snapshot_type, os_info, installed_packages, running_services, config_settings, open_ports)
            VALUES ($1, $2, $3::jsonb, $4::jsonb, $5::jsonb, $6::jsonb, $7::jsonb) RETURNING *
        """, data.get("nodeId"), data.get("snapshotType", "full"),
            json.dumps(data.get("osInfo", {})), json.dumps(data.get("installedPackages", [])),
            json.dumps(data.get("runningServices", [])), json.dumps(data.get("configSettings", {})),
            json.dumps(data.get("openPorts", [])))
        return dict(row)

@app.get("/api/v1/security/posture/diffs", dependencies=[Depends(verify_api_key)])
async def list_posture_diffs(node_id: str = None, limit: int = 50, db: asyncpg.Pool = Depends(get_db)):
    async with db.acquire() as conn:
        if node_id:
            rows = await conn.fetch("SELECT * FROM config_posture_diffs WHERE node_id = $1 ORDER BY created_at DESC LIMIT $2", node_id, limit)
        else:
            rows = await conn.fetch("SELECT * FROM config_posture_diffs ORDER BY created_at DESC LIMIT $1", limit)
        return {"diffs": [dict(r) for r in rows]}

# ============================================================
# E21: Events (normalized) & File Events
# ============================================================

@app.get("/api/v1/security/events", dependencies=[Depends(verify_api_key)])
async def list_security_events(node_id: str = None, event_type: str = None, severity: str = None,
                                hours: int = 24, limit: int = 100, db: asyncpg.Pool = Depends(get_db)):
    async with db.acquire() as conn:
        where = "WHERE ts > NOW() - INTERVAL '1 hour' * $1"
        params = [hours]
        idx = 2
        if node_id:
            where += f" AND node_id = ${idx}"; params.append(node_id); idx += 1
        if event_type:
            where += f" AND event_type = ${idx}"; params.append(event_type); idx += 1
        if severity:
            where += f" AND severity = ${idx}"; params.append(severity); idx += 1
        rows = await conn.fetch(f"SELECT * FROM events_normalized {where} ORDER BY ts DESC LIMIT ${idx}", *params, limit)
        total = await conn.fetchval(f"SELECT COUNT(*) FROM events_normalized {where}", *params)
        return {"events": [dict(r) for r in rows], "total": total}

@app.get("/api/v1/security/file-events", dependencies=[Depends(verify_api_key)])
async def list_file_events(node_id: str = None, op: str = None, path_filter: str = None,
                           hours: int = 24, limit: int = 100, db: asyncpg.Pool = Depends(get_db)):
    async with db.acquire() as conn:
        where = "WHERE ts > NOW() - INTERVAL '1 hour' * $1"
        params = [hours]
        idx = 2
        if node_id:
            where += f" AND node_id = ${idx}"; params.append(node_id); idx += 1
        if op:
            where += f" AND op = ${idx}"; params.append(op); idx += 1
        if path_filter:
            where += f" AND path ILIKE ${idx}"; params.append(f"%{path_filter}%"); idx += 1
        rows = await conn.fetch(f"SELECT * FROM file_events {where} ORDER BY ts DESC LIMIT ${idx}", *params, limit)
        total = await conn.fetchval(f"SELECT COUNT(*) FROM file_events {where}", *params)
        return {"events": [dict(r) for r in rows], "total": total}

# ============================================================
# E21: Evidence Export
# ============================================================

@app.get("/api/v1/security/evidence/export", dependencies=[Depends(verify_api_key)])
async def export_evidence(node_id: str = None, hours: int = 24, db: asyncpg.Pool = Depends(get_db)):
    """Export security evidence package (events + file events + findings) as JSON"""
    async with db.acquire() as conn:
        params_ev = [hours]
        filter_ev = "WHERE ts > NOW() - INTERVAL '1 hour' * $1"
        if node_id:
            filter_ev += " AND node_id = $2"; params_ev.append(node_id)

        events = await conn.fetch(f"SELECT * FROM events_normalized {filter_ev} ORDER BY ts DESC LIMIT 10000", *params_ev)
        file_evts = await conn.fetch(f"SELECT * FROM file_events {filter_ev} ORDER BY ts DESC LIMIT 10000", *params_ev)

        params_f = []
        filter_f = "WHERE 1=1"
        if node_id:
            filter_f += " AND node_id = $1"; params_f.append(node_id)
        findings = await conn.fetch(f"SELECT * FROM security_findings {filter_f} ORDER BY created_at DESC LIMIT 1000", *params_f)

        return {
            "exportedAt": str(datetime.now(timezone.utc)),
            "hours": hours,
            "nodeId": node_id,
            "events": [dict(r) for r in events],
            "fileEvents": [dict(r) for r in file_evts],
            "findings": [dict(r) for r in findings]
        }

# ============================================================
# E21: Agent Capabilities
# ============================================================

@app.get("/api/v1/security/agent/capabilities", dependencies=[Depends(verify_api_key)])
async def list_agent_capabilities(db: asyncpg.Pool = Depends(get_db)):
    async with db.acquire() as conn:
        rows = await conn.fetch("SELECT * FROM agent_capabilities ORDER BY last_seen DESC NULLS LAST")
        return {"capabilities": [dict(r) for r in rows]}

@app.post("/api/v1/security/agent/capabilities", dependencies=[Depends(verify_api_key)])
async def upsert_agent_capabilities(req: Request, db: asyncpg.Pool = Depends(get_db)):
    data = await req.json()
    async with db.acquire() as conn:
        await conn.execute("""
            INSERT INTO agent_capabilities (node_id, sensors, agent_version, os_type, os_version, kernel_build, permissions, last_seen)
            VALUES ($1, $2::jsonb, $3, $4, $5, $6, $7::jsonb, NOW())
            ON CONFLICT (node_id) DO UPDATE SET
                sensors = EXCLUDED.sensors, agent_version = EXCLUDED.agent_version,
                os_type = EXCLUDED.os_type, os_version = EXCLUDED.os_version,
                kernel_build = EXCLUDED.kernel_build, permissions = EXCLUDED.permissions, last_seen = NOW()
        """, data.get("nodeId"), json.dumps(data.get("sensors", {})),
            data.get("agentVersion"), data.get("osType"), data.get("osVersion"),
            data.get("kernelBuild"), json.dumps(data.get("permissions", {})))
        return {"updated": True}

# ============================================================
# E21: UI Audit Log
# ============================================================

@app.get("/api/v1/security/audit-log", dependencies=[Depends(verify_api_key)])
async def list_ui_audit_events(actor: str = None, action: str = None,
                                hours: int = 168, limit: int = 100, db: asyncpg.Pool = Depends(get_db)):
    async with db.acquire() as conn:
        where = "WHERE ts > NOW() - INTERVAL '1 hour' * $1"
        params = [hours]
        idx = 2
        if actor:
            where += f" AND actor_user_id = ${idx}"; params.append(actor); idx += 1
        if action:
            where += f" AND action = ${idx}"; params.append(action); idx += 1
        rows = await conn.fetch(f"SELECT * FROM ui_audit_events {where} ORDER BY ts DESC LIMIT ${idx}", *params, limit)
        return {"events": [dict(r) for r in rows]}

@app.post("/api/v1/security/audit-log", dependencies=[Depends(verify_api_key)])
async def create_ui_audit_event(req: Request, db: asyncpg.Pool = Depends(get_db)):
    data = await req.json()
    async with db.acquire() as conn:
        row = await conn.fetchrow("""
            INSERT INTO ui_audit_events (actor_user_id, action, object_type, object_id, details)
            VALUES ($1, $2, $3, $4, $5::jsonb) RETURNING *
        """, data.get("actor"), data.get("action"), data.get("objectType"),
            data.get("objectId"), json.dumps(data.get("details", {})))
        return dict(row)

# ============================================================
# E21: Detection Rules
# ============================================================

@app.get("/api/v1/security/detection-rules", dependencies=[Depends(verify_api_key)])
async def list_detection_rules(db: asyncpg.Pool = Depends(get_db)):
    async with db.acquire() as conn:
        rows = await conn.fetch("SELECT * FROM detection_rules ORDER BY rule_order, created_at DESC")
        return {"rules": [dict(r) for r in rows]}

@app.post("/api/v1/security/detection-rules", dependencies=[Depends(verify_api_key)])
async def create_detection_rule(req: Request, db: asyncpg.Pool = Depends(get_db)):
    data = await req.json()
    async with db.acquire() as conn:
        row = await conn.fetchrow("""
            INSERT INTO detection_rules (package_version_id, rule_order, rule_type, config, operator)
            VALUES ($1::uuid, $2, $3, $4::jsonb, $5) RETURNING *
        """, data.get("packageVersionId"), data.get("ruleOrder", 0),
            data.get("ruleType"), json.dumps(data.get("config", {})),
            data.get("operator", "AND"))
        return dict(row)

# ============================================================
# E21: Retention Policies & Legal Holds
# ============================================================

@app.get("/api/v1/security/retention/policies", dependencies=[Depends(verify_api_key)])
async def list_retention_policies(db: asyncpg.Pool = Depends(get_db)):
    async with db.acquire() as conn:
        try:
            rows = await conn.fetch("SELECT * FROM retention_policies ORDER BY created_at DESC")
            return {"policies": [dict(r) for r in rows]}
        except Exception:
            return {"policies": [], "note": "retention_policies table not yet created"}

@app.post("/api/v1/security/retention/policies", dependencies=[Depends(verify_api_key)])
async def create_retention_policy(req: Request, db: asyncpg.Pool = Depends(get_db)):
    data = await req.json()
    async with db.acquire() as conn:
        try:
            row = await conn.fetchrow("""
                INSERT INTO retention_policies (name, table_name, retention_days, enabled, created_by)
                VALUES ($1, $2, $3, $4, $5) RETURNING *
            """, data.get("name"), data.get("tableName"), data.get("retentionDays", 90),
                data.get("enabled", True), data.get("createdBy"))
            return dict(row)
        except Exception as e:
            return {"error": str(e), "note": "retention_policies table may not exist yet"}

@app.get("/api/v1/security/retention/holds", dependencies=[Depends(verify_api_key)])
async def list_legal_holds(db: asyncpg.Pool = Depends(get_db)):
    async with db.acquire() as conn:
        try:
            rows = await conn.fetch("SELECT * FROM legal_holds ORDER BY created_at DESC")
            return {"holds": [dict(r) for r in rows]}
        except Exception:
            return {"holds": [], "note": "legal_holds table not yet created"}

@app.post("/api/v1/security/retention/holds", dependencies=[Depends(verify_api_key)])
async def create_legal_hold(req: Request, db: asyncpg.Pool = Depends(get_db)):
    data = await req.json()
    async with db.acquire() as conn:
        try:
            row = await conn.fetchrow("""
                INSERT INTO legal_holds (name, reason, table_names, node_ids, date_from, date_to, created_by)
                VALUES ($1, $2, $3::jsonb, $4::jsonb, $5, $6, $7) RETURNING *
            """, data.get("name"), data.get("reason"),
                json.dumps(data.get("tableNames", [])), json.dumps(data.get("nodeIds", [])),
                data.get("dateFrom"), data.get("dateTo"), data.get("createdBy"))
            return dict(row)
        except Exception as e:
            return {"error": str(e), "note": "legal_holds table may not exist yet"}

# ============================================================
# E21 Story #89: Behavior Rules - Policy Engine (MVP)
# ============================================================

@app.get("/api/v1/security/rules")
async def list_behavior_rules(db: asyncpg.Pool = Depends(get_db)):
    async with db.acquire() as conn:
        rows = await conn.fetch("SELECT * FROM behavior_rules ORDER BY created_at DESC")
        return {"rules": [{
            "id": str(r["id"]), "name": r["name"], "description": r["description"],
            "ruleType": r["rule_type"], "conditions": json.loads(r["conditions"]) if isinstance(r["conditions"], str) else r["conditions"],
            "actions": json.loads(r["actions"]) if isinstance(r["actions"], str) else r["actions"],
            "severity": r["severity"], "enabled": r["enabled"],
            "cooldownSeconds": r["cooldown_seconds"], "createdAt": str(r["created_at"])
        } for r in rows]}

@app.post("/api/v1/security/rules")
async def create_behavior_rule(data: Dict[str, Any], db: asyncpg.Pool = Depends(get_db)):
    async with db.acquire() as conn:
        row = await conn.fetchrow("""
            INSERT INTO behavior_rules (name, description, rule_type, conditions, actions, severity, enabled, cooldown_seconds, created_by)
            VALUES ($1, $2, $3, $4::jsonb, $5::jsonb, $6, $7, $8, $9)
            RETURNING id, created_at
        """, data.get("name", ""), data.get("description"),
            data.get("ruleType", "threshold"),
            json.dumps(data.get("conditions", {})),
            json.dumps(data.get("actions", [])),
            data.get("severity", "medium"),
            data.get("enabled", True),
            data.get("cooldownSeconds", 300),
            data.get("createdBy", "api"))
        return {"id": str(row["id"]), "createdAt": str(row["created_at"])}

@app.put("/api/v1/security/rules/{rule_id}")
async def update_behavior_rule(rule_id: str, data: Dict[str, Any], db: asyncpg.Pool = Depends(get_db)):
    async with db.acquire() as conn:
        await conn.execute("""
            UPDATE behavior_rules SET name=$2, description=$3, rule_type=$4, conditions=$5::jsonb,
                actions=$6::jsonb, severity=$7, enabled=$8, cooldown_seconds=$9, updated_at=NOW()
            WHERE id=$1::uuid
        """, rule_id, data.get("name"), data.get("description"),
            data.get("ruleType", "threshold"),
            json.dumps(data.get("conditions", {})),
            json.dumps(data.get("actions", [])),
            data.get("severity", "medium"),
            data.get("enabled", True),
            data.get("cooldownSeconds", 300))
        return {"updated": True}

@app.delete("/api/v1/security/rules/{rule_id}")
async def delete_behavior_rule(rule_id: str, db: asyncpg.Pool = Depends(get_db)):
    async with db.acquire() as conn:
        await conn.execute("DELETE FROM behavior_rules WHERE id = $1::uuid", rule_id)
        return {"deleted": True}

@app.post("/api/v1/security/rules/evaluate")
async def evaluate_rules(data: Dict[str, Any], db: asyncpg.Pool = Depends(get_db)):
    """Evaluate all enabled behavior rules against recent events. Called periodically or on-demand."""
    async with db.acquire() as conn:
        rules = await conn.fetch("SELECT * FROM behavior_rules WHERE enabled = true")
        findings_created = 0
        
        for rule in rules:
            conditions = json.loads(rule["conditions"]) if isinstance(rule["conditions"], str) else rule["conditions"]
            actions = json.loads(rule["actions"]) if isinstance(rule["actions"], str) else rule["actions"]
            rule_type = rule["rule_type"]
            
            # Check cooldown
            last_finding = await conn.fetchval("""
                SELECT MAX(created_at) FROM security_findings WHERE rule_id = $1::uuid
            """, str(rule["id"]))
            if last_finding:
                from datetime import timezone
                cooldown = rule["cooldown_seconds"] or 300
                if last_finding.replace(tzinfo=timezone.utc) > datetime.now(timezone.utc) - timedelta(seconds=cooldown):
                    continue
            
            triggered = False
            matched_events = []
            affected_node = None
            affected_user = None
            
            if rule_type == "threshold":
                # Count events matching criteria in time window
                event_type = conditions.get("eventType", "%")
                time_window = conditions.get("timeWindowSeconds", 600)
                threshold = conditions.get("threshold", 100)
                path_pattern = conditions.get("pathPattern")
                
                query = """
                    SELECT COUNT(*), MAX(node_id) as node_id
                    FROM events_normalized
                    WHERE ts > NOW() - INTERVAL '1 second' * $1
                      AND event_type LIKE $2
                """
                params = [time_window, event_type]
                if path_pattern:
                    query += " AND payload::text ~* $3"
                    params.append(path_pattern)
                
                row = await conn.fetchrow(query, *params)
                if row and row["count"] >= threshold:
                    triggered = True
                    affected_node = row["node_id"]
            
            elif rule_type == "pattern":
                # Match specific patterns in recent events
                pattern = conditions.get("pattern", "")
                lookback = conditions.get("lookbackSeconds", 300)
                
                rows = await conn.fetch("""
                    SELECT id, node_id, payload FROM events_normalized
                    WHERE ts > NOW() - INTERVAL '1 second' * $1
                      AND (payload::text ~* $2 OR event_type ~* $2)
                    LIMIT 10
                """, lookback, pattern)
                
                if rows:
                    triggered = True
                    affected_node = rows[0]["node_id"]
                    matched_events = [str(r["id"]) for r in rows]
            
            elif rule_type == "time":
                # Events outside business hours
                start_hour = conditions.get("businessHoursStart", 8)
                end_hour = conditions.get("businessHoursEnd", 18)
                event_type = conditions.get("eventType", "file.%")
                
                rows = await conn.fetch("""
                    SELECT id, node_id FROM events_normalized
                    WHERE ts > NOW() - INTERVAL '5 minutes'
                      AND event_type LIKE $1
                      AND (EXTRACT(HOUR FROM ts) < $2 OR EXTRACT(HOUR FROM ts) >= $3)
                    LIMIT 10
                """, event_type, start_hour, end_hour)
                
                if rows:
                    triggered = True
                    affected_node = rows[0]["node_id"]
                    matched_events = [str(r["id"]) for r in rows]
            
            if triggered:
                # Create finding
                title = f"Rule triggered: {rule['name']}"
                for action in actions:
                    if action.get("type") == "create_finding":
                        title = action.get("title", title)
                
                await conn.execute("""
                    INSERT INTO security_findings (title, finding_type, severity, node_id, user_name, 
                        rule_id, description, risk_score)
                    VALUES ($1, 'policy_violation', $2, $3, $4, $5::uuid, $6, $7)
                """, title, rule["severity"], affected_node, affected_user,
                    str(rule["id"]),
                    f"Behavior rule '{rule['name']}' ({rule_type}) triggered",
                    _calculate_risk_score(rule["severity"], len(matched_events)))
                findings_created += 1
        
        return {"evaluated": len(rules), "findingsCreated": findings_created}


def _calculate_risk_score(severity: str, event_count: int = 1) -> float:
    """Calculate risk score based on severity and frequency"""
    base = {"critical": 9.0, "high": 7.0, "medium": 5.0, "low": 3.0, "info": 1.0}.get(severity, 5.0)
    frequency_factor = min(1.0 + (event_count * 0.1), 2.0)
    return round(min(base * frequency_factor, 10.0), 1)


# ============================================================
# E21 Story #90: Findings & Risk Scoring
# ============================================================

@app.get("/api/v1/security/findings")
async def list_findings(
    status: str = None, severity: str = None, node_id: str = None,
    limit: int = 50, offset: int = 0,
    db: asyncpg.Pool = Depends(get_db)
):
    async with db.acquire() as conn:
        query = "SELECT * FROM security_findings WHERE 1=1"
        params = []
        idx = 1
        
        if status:
            query += f" AND status = ${idx}"
            params.append(status)
            idx += 1
        if severity:
            query += f" AND severity = ${idx}"
            params.append(severity)
            idx += 1
        if node_id:
            query += f" AND UPPER(node_id) = UPPER(${idx})"
            params.append(node_id)
            idx += 1
        
        count = await conn.fetchval(query.replace("SELECT *", "SELECT COUNT(*)"), *params)
        
        query += f" ORDER BY risk_score DESC, created_at DESC LIMIT ${idx} OFFSET ${idx+1}"
        params.extend([limit, offset])
        
        rows = await conn.fetch(query, *params)
        return {
            "findings": [{
                "id": str(r["id"]), "title": r["title"], "findingType": r["finding_type"],
                "severity": r["severity"], "riskScore": r["risk_score"], "status": r["status"],
                "nodeId": r["node_id"], "userName": r["user_name"],
                "ruleId": str(r["rule_id"]) if r["rule_id"] else None,
                "description": r["description"], "remediation": r["remediation"],
                "firstSeen": str(r["first_seen"]), "lastSeen": str(r["last_seen"]),
                "closedAt": str(r["closed_at"]) if r["closed_at"] else None,
                "createdAt": str(r["created_at"])
            } for r in rows],
            "total": count
        }

@app.get("/api/v1/security/findings/summary")
async def findings_summary(db: asyncpg.Pool = Depends(get_db)):
    async with db.acquire() as conn:
        by_status = await conn.fetch("SELECT status, COUNT(*) as count FROM security_findings GROUP BY status")
        by_severity = await conn.fetch("SELECT severity, COUNT(*) as count FROM security_findings GROUP BY severity")
        by_type = await conn.fetch("SELECT finding_type, COUNT(*) as count FROM security_findings GROUP BY finding_type")
        top_nodes = await conn.fetch("""
            SELECT node_id, COUNT(*) as count, AVG(risk_score) as avg_risk
            FROM security_findings WHERE status != 'closed'
            GROUP BY node_id ORDER BY count DESC LIMIT 10
        """)
        trend = await conn.fetch("""
            SELECT DATE(created_at) as day, COUNT(*) as count
            FROM security_findings WHERE created_at > NOW() - INTERVAL '30 days'
            GROUP BY DATE(created_at) ORDER BY day
        """)
        return {
            "byStatus": [{"status": r["status"], "count": r["count"]} for r in by_status],
            "bySeverity": [{"severity": r["severity"], "count": r["count"]} for r in by_severity],
            "byType": [{"type": r["finding_type"], "count": r["count"]} for r in by_type],
            "topNodes": [{"nodeId": r["node_id"], "count": r["count"], "avgRisk": round(float(r["avg_risk"] or 0), 1)} for r in top_nodes],
            "trend": [{"day": str(r["day"]), "count": r["count"]} for r in trend]
        }

@app.put("/api/v1/security/findings/{finding_id}")
async def update_finding(finding_id: str, data: Dict[str, Any], db: asyncpg.Pool = Depends(get_db)):
    async with db.acquire() as conn:
        updates = []
        params = [finding_id]
        idx = 2
        
        for field in ["status", "severity", "remediation", "description"]:
            if field in data:
                updates.append(f"{field} = ${idx}")
                params.append(data[field])
                idx += 1
        
        if "status" in data and data["status"] in ("closed", "false_positive"):
            updates.append(f"closed_at = NOW()")
            updates.append(f"closed_by = ${idx}")
            params.append(data.get("closedBy", "api"))
            idx += 1
        
        if updates:
            await conn.execute(f"UPDATE security_findings SET {', '.join(updates)} WHERE id = $1::uuid", *params)
        return {"updated": True}

@app.post("/api/v1/security/findings")
async def create_finding(data: Dict[str, Any], db: asyncpg.Pool = Depends(get_db)):
    async with db.acquire() as conn:
        severity = data.get("severity", "medium")
        row = await conn.fetchrow("""
            INSERT INTO security_findings (title, finding_type, severity, risk_score, node_id, user_name, description, remediation)
            VALUES ($1, $2, $3, $4, $5, $6, $7, $8) RETURNING id
        """, data.get("title", ""), data.get("findingType", "alert"),
            severity, _calculate_risk_score(severity),
            data.get("nodeId"), data.get("userName"),
            data.get("description"), data.get("remediation"))
        return {"id": str(row["id"])}


# ============================================================
# E21 Story #91: Dashboards - File Activity & User Activity
# ============================================================

@app.get("/api/v1/security/activity/files")
async def file_activity_dashboard(
    node_id: str = None, user: str = None, path_filter: str = None,
    hours: int = 24, limit: int = 100,
    db: asyncpg.Pool = Depends(get_db)
):
    """File activity dashboard data"""
    async with db.acquire() as conn:
        base_filter = "WHERE ts > NOW() - INTERVAL '1 hour' * $1"
        params = [hours]
        idx = 2
        
        if node_id:
            base_filter += f" AND node_id = ${idx}"
            params.append(node_id)
            idx += 1
        
        # Top paths
        top_paths = await conn.fetch(f"""
            SELECT path, COUNT(*) as count, 
                   array_agg(DISTINCT op) as ops
            FROM file_events {base_filter}
            AND path IS NOT NULL
            GROUP BY path
            ORDER BY count DESC LIMIT 20
        """, *params)
        
        # Operations distribution
        ops_dist = await conn.fetch(f"""
            SELECT op as event_type, COUNT(*) as count
            FROM file_events {base_filter}
            GROUP BY op ORDER BY count DESC
        """, *params)
        
        # Timeline (hourly buckets)
        timeline = await conn.fetch(f"""
            SELECT date_trunc('hour', ts) as hour, COUNT(*) as count
            FROM file_events {base_filter}
            GROUP BY date_trunc('hour', ts) ORDER BY hour
        """, *params)
        
        # Top users
        top_users = await conn.fetch(f"""
            SELECT user_id as username, COUNT(*) as count
            FROM file_events {base_filter}
            AND user_id IS NOT NULL
            GROUP BY user_id
            ORDER BY count DESC LIMIT 10
        """, *params)
        
        # Recent events
        limit_param = f"${idx}"
        recent = await conn.fetch(f"""
            SELECT event_id as id, ts, node_id, op as event_type, 
                   path, old_path, new_path, process_name, pid, user_id,
                   hash_before, hash_after, file_size
            FROM file_events {base_filter}
            ORDER BY ts DESC LIMIT {limit_param}
        """, *params, limit)
        
        return {
            "topPaths": [{"path": r["path"], "count": r["count"], "operations": list(r["ops"] or [])} for r in top_paths],
            "operationsDistribution": [{"operation": r["event_type"], "count": r["count"]} for r in ops_dist],
            "timeline": [{"hour": str(r["hour"]), "count": r["count"]} for r in timeline],
            "topUsers": [{"username": r["username"], "count": r["count"]} for r in top_users],
            "recentEvents": [{
                "id": str(r["id"]), "ts": str(r["ts"]), "nodeId": r["node_id"],
                "eventType": r["event_type"],
                "payload": {"path": r["path"], "oldPath": r["old_path"], "newPath": r["new_path"],
                            "process": r["process_name"], "pid": r["pid"], "user": r["user_id"],
                            "hashBefore": r["hash_before"], "hashAfter": r["hash_after"], "size": r["file_size"]}
            } for r in recent]
        }

@app.get("/api/v1/security/activity/users")
async def user_activity_dashboard(
    node_id: str = None, username: str = None, hours: int = 24,
    db: asyncpg.Pool = Depends(get_db)
):
    """User activity dashboard"""
    async with db.acquire() as conn:
        base_filter = "WHERE ts > NOW() - INTERVAL '1 hour' * $1"
        params = [hours]
        idx = 2
        
        if node_id:
            base_filter += f" AND node_id = ${idx}"
            params.append(node_id)
            idx += 1
        
        # User activity summary
        users = await conn.fetch(f"""
            SELECT user_id as username,
                   COUNT(*) as total_events,
                   COUNT(DISTINCT op) as unique_ops,
                   COUNT(DISTINCT path) as unique_files,
                   MIN(ts) as first_activity,
                   MAX(ts) as last_activity
            FROM file_events {base_filter}
            AND user_id IS NOT NULL
            GROUP BY user_id
            ORDER BY total_events DESC LIMIT 20
        """, *params)
        
        # Unusual hours activity (outside 8-18)
        after_hours = await conn.fetch(f"""
            SELECT user_id as username, COUNT(*) as count
            FROM file_events {base_filter}
            AND (EXTRACT(HOUR FROM ts) < 8 OR EXTRACT(HOUR FROM ts) >= 18)
            AND user_id IS NOT NULL
            GROUP BY user_id
            ORDER BY count DESC LIMIT 10
        """, *params)
        
        # Sensitive path access (e.g., /etc, Windows system dirs)
        sensitive = await conn.fetch(f"""
            SELECT user_id as username, path, COUNT(*) as count
            FROM file_events {base_filter}
            AND (path LIKE '/etc/%' OR path LIKE 'C:\\Windows\\%' 
                 OR path LIKE 'C:\\Program Files\\%')
            GROUP BY user_id, path
            ORDER BY count DESC LIMIT 20
        """, *params)
        
        return {
            "users": [{
                "username": r["username"], "totalEvents": r["total_events"],
                "uniqueOps": r["unique_ops"], "uniqueFiles": r["unique_files"],
                "firstActivity": str(r["first_activity"]), "lastActivity": str(r["last_activity"])
            } for r in users],
            "afterHoursActivity": [{"username": r["username"], "count": r["count"]} for r in after_hours],
            "sensitiveAccess": [{"username": r["username"], "path": r["path"], "count": r["count"]} for r in sensitive]
        }

@app.get("/api/v1/security/activity/export")
async def export_file_activity(
    node_id: str = None, hours: int = 24, format: str = "csv",
    db: asyncpg.Pool = Depends(get_db)
):
    """Export file activity as CSV"""
    async with db.acquire() as conn:
        query = "SELECT ts, node_id, op, path, user_id, process_name, file_size, hash_after FROM file_events WHERE ts > NOW() - INTERVAL '1 hour' * $1"
        params = [hours]
        if node_id:
            query += " AND node_id = $2"
            params.append(node_id)
        query += " ORDER BY ts DESC LIMIT 10000"
        
        rows = await conn.fetch(query, *params)
        
        if format == "csv":
            import io
            output = io.StringIO()
            output.write("timestamp,node_id,operation,path,user,file_size,hash\n")
            for r in rows:
                path = r.get("path", "") or ""
                user = r.get("user_id", "") or ""
                size = r.get("file_size", "") or ""
                fhash = r.get("hash_after", "") or ""
                output.write(f'{r["ts"]},"{r["node_id"]}","{r["op"]}","{path}","{user}",{size},"{fhash}"\n')
            
            from starlette.responses import Response
            return Response(
                content=output.getvalue(),
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename=file-activity-{datetime.now().strftime('%Y%m%d')}.csv"}
            )
        
        return {"events": [{"ts": str(r["ts"]), "nodeId": r["node_id"], "eventType": r["event_type"],
                           "payload": json.loads(r["payload"]) if isinstance(r["payload"], str) else r["payload"]} for r in rows]}


@app.get("/api/v1/admin/agent-activity")
async def get_agent_activity(limit: int = 100, _: str = Depends(verify_api_key)):
    """Get recent agent activity from in-memory buffer."""
    items = list(_agent_activity_buffer)[-limit:]
    items.reverse()
    return {"events": items, "count": len(items)}


@app.get("/api/v1/admin/agent-status")
async def get_agent_status(_: str = Depends(verify_api_key)):
    """Get live status of all agents with their recent activity."""
    async with db_pool.acquire() as conn:
        nodes = await conn.fetch("""
            SELECT n.id, n.hostname, n.os_name, n.agent_version, n.last_seen, n.is_online,
                   (SELECT COUNT(*) FROM job_instances ji WHERE ji.node_id = n.id::text AND ji.status = 'running') as running_jobs,
                   (SELECT COUNT(*) FROM job_instances ji WHERE ji.node_id = n.id::text AND ji.status = 'queued') as queued_jobs,
                   (SELECT COUNT(*) FROM remediation_jobs rj WHERE rj.node_id = n.id AND rj.status IN ('pending', 'approved')) as pending_remediation,
                   (SELECT COUNT(*) FROM remediation_jobs rj WHERE rj.node_id = n.id AND rj.status = 'running') as running_remediation
            FROM nodes n
            ORDER BY n.last_seen DESC NULLS LAST
        """)

        result = []
        for n in nodes:
            last_seen = n["last_seen"]
            now = datetime.now(timezone.utc)
            if last_seen:
                ago = (now - last_seen).total_seconds()
                if ago < 30:
                    status = "active"
                elif ago < 120:
                    status = "idle"
                elif ago < 600:
                    status = "stale"
                else:
                    status = "offline"
            else:
                status = "unknown"
                ago = None

            # Find recent activity from buffer
            recent_actions = [e for e in _agent_activity_buffer if e["hostname"] == n["hostname"]][-5:]

            result.append({
                "id": str(n["id"]),
                "hostname": n["hostname"],
                "os_name": n["os_name"],
                "agent_version": n["agent_version"],
                "last_seen": last_seen.isoformat() if last_seen else None,
                "seconds_ago": int(ago) if ago is not None else None,
                "status": status,
                "is_online": n["is_online"],
                "running_jobs": n["running_jobs"],
                "queued_jobs": n["queued_jobs"],
                "pending_remediation": n["pending_remediation"],
                "running_remediation": n["running_remediation"],
                "recent_actions": recent_actions
            })

        return {"agents": result}


@app.get("/api/v1/admin/agent-live")
async def agent_live_sse(request: Request, token: str = None):
    """SSE endpoint for live agent activity stream."""
    async def event_generator():
        last_idx = len(_agent_activity_buffer)
        last_node_check = 0

        while True:
            if await request.is_disconnected():
                break

            # Send new activity events
            current = list(_agent_activity_buffer)
            if len(current) > last_idx - (len(_agent_activity_buffer) - len(current)):
                new_events = current[-(len(current) - last_idx):] if last_idx < len(current) else current
                for evt in new_events[-10:]:
                    yield f"data: {json.dumps({'type': 'activity', 'event': evt}, default=str)}\n\n"
                last_idx = len(current)

            # Every 5 seconds, send node status update
            now = _time.time()
            if now - last_node_check > 5:
                last_node_check = now
                try:
                    async with db_pool.acquire() as conn:
                        nodes = await conn.fetch("""
                            SELECT hostname, last_seen, is_online, agent_version FROM nodes ORDER BY hostname
                        """)
                        statuses = []
                        for n in nodes:
                            ls = n["last_seen"]
                            ago = (datetime.now(timezone.utc) - ls).total_seconds() if ls else 9999
                            statuses.append({
                                "hostname": n["hostname"],
                                "last_seen": ls.isoformat() if ls else None,
                                "seconds_ago": int(ago),
                                "status": "active" if ago < 30 else "idle" if ago < 120 else "stale" if ago < 600 else "offline",
                                "agent_version": n["agent_version"]
                            })
                        yield f"data: {json.dumps({'type': 'status', 'agents': statuses}, default=str)}\n\n"
                except Exception as e:
                    logger.error(f"Agent SSE status error: {e}")

            await asyncio.sleep(2)

    return StreamingResponse(event_generator(), media_type="text/event-stream")


# ============================================================
# E30: Patch & Update Orchestration
# ============================================================


# -- Patch Rings --


# -- Patch Deployments --


# -- Agent Endpoints --


# ============================================================
# E31: Configuration Baselines & Drift Management
# ============================================================

# --- Baselines CRUD ---


# --- Compliance Trends ---


# --- CIS Benchmark Templates ---

CIS_TEMPLATES = {
    "cis-win-server-2022-l1": {
        "id": "cis-win-server-2022-l1",
        "name": "CIS Windows Server 2022/2025 L1",
        "description": "CIS Benchmark Level 1 for Windows Server 2022/2025 — essential security hardening",
        "baseline_type": "custom",
        "rules": [
            {"rule_name": "Windows Firewall Service Running", "rule_type": "service",
             "expected_value": {"service": "MpsSvc", "state": "running"},
             "severity": "critical", "remediation_action": {"type": "start_service", "service": "MpsSvc"}},
            {"rule_name": "Windows Update Service Enabled", "rule_type": "service",
             "expected_value": {"service": "wuauserv", "state": "running"},
             "severity": "high", "remediation_action": {"type": "start_service", "service": "wuauserv"}},
            {"rule_name": "Remote Registry Disabled", "rule_type": "service",
             "expected_value": {"service": "RemoteRegistry", "state": "stopped"},
             "severity": "high", "remediation_action": {"type": "run_command", "command": "Stop-Service RemoteRegistry; Set-Service RemoteRegistry -StartupType Disabled"}},
            {"rule_name": "WinRM Service Running", "rule_type": "service",
             "expected_value": {"service": "WinRM", "state": "running"},
             "severity": "medium", "remediation_action": {"type": "start_service", "service": "WinRM"}},
            {"rule_name": "Blank Password Use Limited", "rule_type": "registry",
             "expected_value": {"key_path": r"HKLM\SYSTEM\CurrentControlSet\Control\Lsa", "value_name": "LimitBlankPasswordUse", "expected": "1", "operator": "eq"},
             "severity": "critical", "remediation_action": {"type": "run_command", "command": "Set-ItemProperty -Path 'HKLM:\\SYSTEM\\CurrentControlSet\\Control\\Lsa' -Name LimitBlankPasswordUse -Value 1"}},
            {"rule_name": "Security Event Log Min Size", "rule_type": "registry",
             "expected_value": {"key_path": r"HKLM\SYSTEM\CurrentControlSet\Services\EventLog\Security", "value_name": "MaxSize", "expected": "20480", "operator": "gte"},
             "severity": "high", "remediation_action": {"type": "run_command", "command": "Set-ItemProperty -Path 'HKLM:\\SYSTEM\\CurrentControlSet\\Services\\EventLog\\Security' -Name MaxSize -Value 20971520"}},
            {"rule_name": "SMBv1 Disabled", "rule_type": "registry",
             "expected_value": {"key_path": r"HKLM\SYSTEM\CurrentControlSet\Services\LanmanServer\Parameters", "value_name": "SMB1", "expected": "0", "operator": "eq"},
             "severity": "critical", "remediation_action": {"type": "run_command", "command": "Set-ItemProperty -Path 'HKLM:\\SYSTEM\\CurrentControlSet\\Services\\LanmanServer\\Parameters' -Name SMB1 -Value 0"}},
            {"rule_name": "NTLMv2 Required", "rule_type": "registry",
             "expected_value": {"key_path": r"HKLM\SYSTEM\CurrentControlSet\Control\Lsa", "value_name": "LmCompatibilityLevel", "expected": "3", "operator": "gte"},
             "severity": "critical", "remediation_action": {"type": "run_command", "command": "Set-ItemProperty -Path 'HKLM:\\SYSTEM\\CurrentControlSet\\Control\\Lsa' -Name LmCompatibilityLevel -Value 5"}},
            {"rule_name": "Windows Defender / Antivirus Present", "rule_type": "software",
             "expected_value": {"package": "Windows Defender", "operator": "installed"},
             "severity": "critical", "remediation_action": {"type": "run_command", "command": "try { Install-WindowsFeature -Name Windows-Defender -ErrorAction Stop } catch { Add-WindowsCapability -Online -Name 'Microsoft.Windows.Defender~~~~' -ErrorAction Stop }"}},
        ]
    },
    "cis-win11-enterprise-l1": {
        "id": "cis-win11-enterprise-l1",
        "name": "CIS Windows 11 Enterprise L1",
        "description": "CIS Benchmark Level 1 for Windows 11 Enterprise — workstation security hardening",
        "baseline_type": "custom",
        "rules": [
            {"rule_name": "Windows Defender Antivirus Running", "rule_type": "service",
             "expected_value": {"service": "WinDefend", "state": "running"},
             "severity": "critical", "remediation_action": {"type": "start_service", "service": "WinDefend"}},
            {"rule_name": "BitLocker Drive Encryption Running", "rule_type": "service",
             "expected_value": {"service": "BDESVC", "state": "running"},
             "severity": "high", "remediation_action": {"type": "start_service", "service": "BDESVC"}},
            {"rule_name": "UAC Enabled", "rule_type": "registry",
             "expected_value": {"key_path": r"HKLM\SOFTWARE\Microsoft\Windows\CurrentVersion\Policies\System", "value_name": "EnableLUA", "expected": "1", "operator": "eq"},
             "severity": "critical", "remediation_action": {"type": "run_command", "command": "Set-ItemProperty -Path 'HKLM:\\SOFTWARE\\Microsoft\\Windows\\CurrentVersion\\Policies\\System' -Name EnableLUA -Value 1"}},
            {"rule_name": "Secure Boot Enabled", "rule_type": "registry",
             "expected_value": {"key_path": "SecureBoot", "value_name": "UEFISecureBootEnabled", "expected": "1", "operator": "eq"},
             "severity": "high", "remediation_action": {"type": "none"}},
            {"rule_name": "Windows Firewall Domain Profile ON", "rule_type": "registry",
             "expected_value": {"key_path": r"HKLM\SYSTEM\CurrentControlSet\Services\SharedAccess\Parameters\FirewallPolicy\DomainProfile", "value_name": "EnableFirewall", "expected": "1", "operator": "eq"},
             "severity": "critical", "remediation_action": {"type": "run_command", "command": "Set-NetFirewallProfile -Profile Domain -Enabled True"}},
            {"rule_name": "Microsoft Edge or Approved Browser Installed", "rule_type": "software",
             "expected_value": {"package": "Microsoft Edge", "operator": "installed"},
             "severity": "medium", "remediation_action": {"type": "install_package", "package": "microsoft-edge"}},
        ]
    }
}


# --- Drift Remediation ---


# --- Rules ---


# --- Assignments ---


# --- Evaluation & Drift ---

