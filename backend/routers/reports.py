"""
Octofleet API - Reports Routes
"""
import csv
import io
from datetime import datetime
from io import BytesIO
from typing import Optional

import asyncpg
from fastapi import APIRouter, Depends
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
)
from starlette.responses import StreamingResponse

from app.core.report_helpers import (
    auto_column_width,
    create_header_footer,
    create_status_table,
    style_excel_header,
)
from dependencies import get_db, verify_api_key
from routers.security import get_compliance_summary

router = APIRouter(tags=["Reports"])

def create_pdf_styles():
    """Create custom styles for PDF reports"""
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name='ReportTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        textColor=colors.HexColor('#1a1a2e')
    ))
    styles.add(ParagraphStyle(
        name='SectionTitle',
        parent=styles['Heading2'],
        fontSize=16,
        spaceAfter=12,
        spaceBefore=20,
        textColor=colors.HexColor('#16213e')
    ))
    styles.add(ParagraphStyle(
        name='SubSection',
        parent=styles['Heading3'],
        fontSize=12,
        spaceAfter=8,
        textColor=colors.HexColor('#0f3460')
    ))
    return styles





@router.get("/api/v1/export/nodes", dependencies=[Depends(verify_api_key)])
async def export_nodes(format: str = "json", db: asyncpg.Pool = Depends(get_db)):
    """Export all nodes as CSV or JSON"""
    async with db.acquire() as conn:
        rows = await conn.fetch("""
            SELECT 
                node_id, hostname, os_name, os_version, os_build,
                agent_version, first_seen, last_seen
            FROM nodes
            ORDER BY hostname
        """)
        
        data = [
            {
                "node_id": str(row["node_id"]),
                "hostname": row["hostname"],
                "os_name": row["os_name"],
                "os_version": row["os_version"],
                "os_build": row["os_build"],
                "agent_version": row["agent_version"],
                "first_seen": row["first_seen"].isoformat() if row["first_seen"] else None,
                "last_seen": row["last_seen"].isoformat() if row["last_seen"] else None,
            }
            for row in rows
        ]
        
        if format == "csv":
            output = io.StringIO()
            if data:
                writer = csv.DictWriter(output, fieldnames=data[0].keys())
                writer.writeheader()
                writer.writerows(data)
            
            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="text/csv",
                headers={"Content-Disposition": "attachment; filename=nodes.csv"}
            )
        else:
            return data


@router.get("/api/v1/export/software", dependencies=[Depends(verify_api_key)])
async def export_software(format: str = "json", db: asyncpg.Pool = Depends(get_db)):
    """Export all software as CSV or JSON"""
    async with db.acquire() as conn:
        rows = await conn.fetch("""
            SELECT 
                n.hostname,
                s.data->'installedPrograms' as programs
            FROM nodes n
            LEFT JOIN inventory_software s ON n.node_id = s.node_id
            WHERE s.data IS NOT NULL
        """)
        
        data = []
        for row in rows:
            programs = row["programs"] or []
            if isinstance(programs, str):
                import json
                programs = json.loads(programs)
            
            for prog in programs:
                data.append({
                    "hostname": row["hostname"],
                    "name": prog.get("name"),
                    "version": prog.get("version"),
                    "publisher": prog.get("publisher")
                })
        
        if format == "csv":
            output = io.StringIO()
            if data:
                writer = csv.DictWriter(output, fieldnames=["hostname", "name", "version", "publisher"])
                writer.writeheader()
                writer.writerows(data)
            
            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="text/csv",
                headers={"Content-Disposition": "attachment; filename=software.csv"}
            )
        else:
            return data


@router.get("/api/v1/export/compliance", dependencies=[Depends(verify_api_key)])
async def export_compliance(format: str = "json", db: asyncpg.Pool = Depends(get_db)):
    """Export compliance data as CSV or JSON"""
    summary = await get_compliance_summary(db)
    data = summary["nodes"]
    
    if format == "csv":
        output = io.StringIO()
        if data:
            writer = csv.DictWriter(output, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=compliance.csv"}
        )
    else:
        return data
    return {"status": "checked"}


@router.get("/api/v1/export/nodes/excel", dependencies=[Depends(verify_api_key)])
async def export_nodes_excel(db: asyncpg.Pool = Depends(get_db)):
    """Export all nodes to Excel."""
    async with db.acquire() as conn:
        rows = await conn.fetch("""
            SELECT n.id, n.hostname, n.domain, n.os_name, n.os_version, n.agent_version,
                   n.last_seen, n.is_online,
                   h.cpu->>'Name' as cpu_name,
                   (h.cpu->>'Cores')::int as cpu_cores,
                   (h.ram->>'TotalGB')::numeric as ram_total_gb,
                   h.mainboard->>'Manufacturer' as manufacturer,
                   h.mainboard->>'Product' as model,
                   h.bios->>'SerialNumber' as serial_number
            FROM nodes n
            LEFT JOIN hardware_current h ON h.node_id = n.id
            ORDER BY n.hostname
        """)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Nodes"
    
    # Headers
    headers = ["ID", "Hostname", "Domain", "OS", "OS Version", "Agent Version", "Last Seen", 
               "Online", "CPU", "Cores", "RAM (GB)",
               "Manufacturer", "Model", "Serial Number"]
    ws.append(headers)
    style_excel_header(ws)
    
    # Data
    for row in rows:
        ws.append([
            str(row["id"]), row["hostname"], row["domain"], row["os_name"], row["os_version"],
            row["agent_version"], row["last_seen"].isoformat() if row["last_seen"] else None,
            "Yes" if row["is_online"] else "No",
            row["cpu_name"], row["cpu_cores"], 
            round(float(row["ram_total_gb"]), 1) if row["ram_total_gb"] else None,
            row["manufacturer"], row["model"], row["serial_number"]
        ])
    
    auto_column_width(ws)
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"octofleet_nodes_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/api/v1/export/software/excel", dependencies=[Depends(verify_api_key)])
async def export_software_excel(db: asyncpg.Pool = Depends(get_db)):
    """Export all installed software to Excel."""
    async with db.acquire() as conn:
        rows = await conn.fetch("""
            SELECT n.hostname, s.name, s.version, s.publisher, s.install_date, s.install_path
            FROM software_current s
            JOIN nodes n ON n.id = s.node_id
            ORDER BY n.hostname, s.name
        """)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Software"
    
    headers = ["Hostname", "Software Name", "Version", "Publisher", "Install Date", "Install Location"]
    ws.append(headers)
    style_excel_header(ws)
    
    for row in rows:
        ws.append([row["hostname"], row["name"], row["version"], row["publisher"],
                   row["install_date"].isoformat() if row["install_date"] else None, 
                   row["install_path"]])
    
    auto_column_width(ws)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"octofleet_software_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/api/v1/export/vulnerabilities/excel", dependencies=[Depends(verify_api_key)])
async def export_vulnerabilities_excel(db: asyncpg.Pool = Depends(get_db)):
    """Export all vulnerabilities with affected nodes to Excel."""
    async with db.acquire() as conn:
        # Get vulnerabilities with affected node count
        rows = await conn.fetch("""
            SELECT v.cve_id, v.software_name, v.software_version,
                   v.severity, v.cvss_score, v.description, v.published_date,
                   (SELECT COUNT(DISTINCT s.node_id) 
                    FROM software_current s 
                    WHERE s.name ILIKE v.software_name 
                    AND s.version = v.software_version) as affected_nodes,
                   (SELECT STRING_AGG(DISTINCT n.hostname, ', ' ORDER BY n.hostname)
                    FROM software_current s 
                    JOIN nodes n ON n.id = s.node_id
                    WHERE s.name ILIKE v.software_name 
                    AND s.version = v.software_version
                    LIMIT 5) as node_list
            FROM vulnerabilities v
            ORDER BY v.cvss_score DESC NULLS LAST, v.software_name
        """)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Vulnerabilities"
    
    headers = ["CVE ID", "Software", "Version", "Severity", "CVSS Score",
               "Affected Nodes", "Nodes (first 5)", "Published", "Description"]
    ws.append(headers)
    style_excel_header(ws)
    
    # Color coding for severity
    severity_colors = {
        "CRITICAL": "FF0000",
        "HIGH": "FF6600", 
        "MEDIUM": "FFCC00",
        "LOW": "00CC00"
    }
    
    for i, row in enumerate(rows, start=2):
        ws.append([row["cve_id"], row["software_name"], row["software_version"],
                   row["severity"], float(row["cvss_score"]) if row["cvss_score"] else None,
                   row["affected_nodes"], row["node_list"],
                   row["published_date"].isoformat() if row["published_date"] else None,
                   row["description"][:200] if row["description"] else None])
        
        # Color the severity cell
        if row["severity"] in severity_colors:
            ws.cell(row=i, column=4).fill = PatternFill(
                start_color=severity_colors[row["severity"]], 
                end_color=severity_colors[row["severity"]], 
                fill_type="solid"
            )
    
    auto_column_width(ws)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"octofleet_vulnerabilities_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/api/v1/export/jobs/excel", dependencies=[Depends(verify_api_key)])
async def export_jobs_excel(
    days: int = 30,
    db: asyncpg.Pool = Depends(get_db)
):
    """Export job history to Excel."""
    async with db.acquire() as conn:
        rows = await conn.fetch("""
            SELECT j.id, j.name, j.command_type, j.created_at, j.created_by,
                   ji.node_id as target_node, ji.status, ji.queued_at, ji.started_at, ji.completed_at,
                   ji.exit_code, ji.error_message, ji.duration_ms
            FROM jobs j
            LEFT JOIN job_instances ji ON ji.job_id = j.id
            WHERE j.created_at > NOW() - INTERVAL '1 day' * $1
            ORDER BY j.created_at DESC
        """, days)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"
    
    headers = ["Job ID", "Name", "Type", "Created By", "Target Node", "Status", 
               "Queued", "Started", "Completed", "Duration (ms)", "Exit Code", "Error"]
    ws.append(headers)
    style_excel_header(ws)
    
    status_colors = {
        "completed": "00CC00",
        "failed": "FF0000",
        "running": "0066FF",
        "pending": "CCCCCC"
    }
    
    for i, row in enumerate(rows, start=2):
        ws.append([
            str(row["id"])[:8], row["name"], row["command_type"], row["created_by"],
            row["target_node"], row["status"],
            row["queued_at"].isoformat() if row["queued_at"] else None,
            row["started_at"].isoformat() if row["started_at"] else None,
            row["completed_at"].isoformat() if row["completed_at"] else None,
            row["duration_ms"], row["exit_code"],
            row["error_message"][:100] if row["error_message"] else None
        ])
        
        if row["status"] in status_colors:
            ws.cell(row=i, column=6).fill = PatternFill(
                start_color=status_colors[row["status"]], 
                end_color=status_colors[row["status"]], 
                fill_type="solid"
            )
    
    auto_column_width(ws)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"octofleet_jobs_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/api/v1/reports/fleet/pdf", dependencies=[Depends(verify_api_key)])
async def generate_fleet_summary_pdf(db: asyncpg.Pool = Depends(get_db)):
    """E19-05: Generate Fleet Summary PDF Report"""
    
    # Fetch data - join nodes with v_nodes_overview for all columns
    nodes = await db.fetch("""
        SELECT n.hostname, n.os_name, n.agent_version, n.is_online, n.last_seen,
               v.cpu_name, v.ram_gb
        FROM nodes n
        LEFT JOIN v_nodes_overview v ON n.id = v.id
        ORDER BY n.hostname
    """)
    
    # Online/Offline stats (no health_status in schema, use online status)
    online_stats = await db.fetch("""
        SELECT is_online, COUNT(*) as count
        FROM nodes GROUP BY is_online
    """)
    
    # Performance summary (last 24h averages) from node_metrics
    perf_stats = await db.fetchrow("""
        SELECT 
            ROUND(AVG(cpu_percent)::numeric, 1) as avg_cpu,
            ROUND(AVG(ram_percent)::numeric, 1) as avg_memory,
            ROUND(AVG(disk_percent)::numeric, 1) as avg_disk
        FROM node_metrics
        WHERE time > NOW() - INTERVAL '24 hours'
    """)
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, 
                           rightMargin=50, leftMargin=50,
                           topMargin=60, bottomMargin=50)
    
    styles = create_pdf_styles()
    story = []
    
    # Title
    story.append(Paragraph("Fleet Summary Report", styles['ReportTitle']))
    story.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Overview Section
    story.append(Paragraph("Fleet Overview", styles['SectionTitle']))
    
    total_nodes = len(nodes)
    online_nodes = sum(1 for n in nodes if n['is_online'])
    
    overview_data = [
        ["Metric", "Value"],
        ["Total Nodes", str(total_nodes)],
        ["Online Nodes", f"{online_nodes} ({round(online_nodes/total_nodes*100) if total_nodes else 0}%)"],
        ["Offline Nodes", f"{total_nodes - online_nodes}"],
        ["Agent Versions", ", ".join(set(n['agent_version'] for n in nodes if n['agent_version'])) or "-"],
    ]
    story.append(create_status_table(overview_data, [150, 300]))
    story.append(Spacer(1, 20))
    
    # Performance Section
    story.append(Paragraph("Performance Summary (24h Average)", styles['SectionTitle']))
    
    perf_data = [
        ["Metric", "Average"],
        ["CPU Usage", f"{perf_stats['avg_cpu'] or 0}%"],
        ["Memory Usage", f"{perf_stats['avg_memory'] or 0}%"],
        ["Disk Usage", f"{perf_stats['avg_disk'] or 0}%"],
    ]
    story.append(create_status_table(perf_data, [150, 150]))
    story.append(Spacer(1, 20))
    
    # Online Status Distribution
    story.append(Paragraph("Status Distribution", styles['SectionTitle']))
    
    status_data = [["Status", "Count"]]
    for s in online_stats:
        status_label = "Online" if s['is_online'] else "Offline"
        status_data.append([status_label, str(s['count'])])
    story.append(create_status_table(status_data, [150, 100]))
    story.append(Spacer(1, 20))
    
    # Node List
    story.append(Paragraph("Node Inventory", styles['SectionTitle']))
    
    node_data = [["Hostname", "OS", "Version", "Status", "CPU"]]
    for n in nodes:
        os_short = (n['os_name'] or '-')[:30]
        cpu_short = (n['cpu_name'] or '-')[:25] if n.get('cpu_name') else '-'
        node_data.append([
            n['hostname'],
            os_short,
            n['agent_version'] or '-',
            "Online" if n['is_online'] else "Offline",
            cpu_short
        ])
    story.append(create_status_table(node_data, [90, 140, 50, 55, 100]))
    
    # Build PDF
    doc.build(story, onFirstPage=lambda c, d: create_header_footer(c, d, "Fleet Summary"),
              onLaterPages=lambda c, d: create_header_footer(c, d, "Fleet Summary"))
    
    buffer.seek(0)
    filename = f"octofleet_fleet_summary_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/api/v1/reports/security/pdf", dependencies=[Depends(verify_api_key)])
async def generate_security_report_pdf(db: asyncpg.Pool = Depends(get_db)):
    """E19-06: Generate Security Report PDF (CVEs, Compliance)"""
    
    # Fetch vulnerability data
    vulns = await db.fetch("""
        SELECT v.cve_id, v.severity, v.cvss_score, v.description,
               v.software_name, v.software_version,
               (SELECT COUNT(DISTINCT s.node_id) FROM software_current s 
                WHERE s.name = v.software_name AND s.version = v.software_version) as affected_nodes
        FROM vulnerabilities v
        ORDER BY v.cvss_score DESC NULLS LAST
        LIMIT 50
    """)
    
    # Severity distribution
    severity_stats = await db.fetch("""
        SELECT severity, COUNT(*) as count
        FROM vulnerabilities
        GROUP BY severity
        ORDER BY CASE severity 
            WHEN 'critical' THEN 1 
            WHEN 'high' THEN 2 
            WHEN 'medium' THEN 3 
            WHEN 'low' THEN 4 
            ELSE 5 END
    """)
    
    # Nodes with vulnerabilities (via software match)
    node_vulns = await db.fetch("""
        SELECT n.hostname,
               COUNT(DISTINCT v.id) as vuln_count,
               SUM(CASE WHEN v.severity = 'critical' THEN 1 ELSE 0 END) as critical_count
        FROM nodes n
        JOIN software_current s ON s.node_id = n.id
        JOIN vulnerabilities v ON v.software_name = s.name AND v.software_version = s.version
        GROUP BY n.id, n.hostname
        HAVING COUNT(DISTINCT v.id) > 0
        ORDER BY critical_count DESC, vuln_count DESC
        LIMIT 20
    """)
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                           rightMargin=50, leftMargin=50,
                           topMargin=60, bottomMargin=50)
    
    styles = create_pdf_styles()
    story = []
    
    # Title
    story.append(Paragraph("Security Report", styles['ReportTitle']))
    story.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Summary
    story.append(Paragraph("Vulnerability Summary", styles['SectionTitle']))
    
    total_vulns = sum(s['count'] for s in severity_stats)
    critical = next((s['count'] for s in severity_stats if s['severity'] == 'critical'), 0)
    high = next((s['count'] for s in severity_stats if s['severity'] == 'high'), 0)
    
    summary_data = [
        ["Metric", "Value"],
        ["Total Vulnerabilities", str(total_vulns)],
        ["Critical", str(critical)],
        ["High", str(high)],
        ["Affected Nodes", str(len(node_vulns))],
    ]
    story.append(create_status_table(summary_data, [150, 150]))
    story.append(Spacer(1, 20))
    
    # Severity Distribution
    story.append(Paragraph("Severity Distribution", styles['SectionTitle']))
    sev_data = [["Severity", "Count", "Percentage"]]
    for s in severity_stats:
        pct = round(s['count'] / total_vulns * 100, 1) if total_vulns else 0
        sev_data.append([s['severity'].title(), str(s['count']), f"{pct}%"])
    story.append(create_status_table(sev_data, [100, 80, 80]))
    story.append(Spacer(1, 20))
    
    # Most Affected Nodes
    story.append(Paragraph("Most Affected Nodes", styles['SectionTitle']))
    node_data = [["Hostname", "Total CVEs", "Critical"]]
    for n in node_vulns[:10]:
        node_data.append([n['hostname'], str(n['vuln_count']), str(n['critical_count'])])
    story.append(create_status_table(node_data, [150, 80, 80]))
    story.append(Spacer(1, 20))
    
    # Top Vulnerabilities
    story.append(Paragraph("Top Vulnerabilities by CVSS Score", styles['SectionTitle']))
    vuln_data = [["CVE ID", "Severity", "CVSS", "Affected"]]
    for v in vulns[:15]:
        vuln_data.append([
            v['cve_id'],
            (v['severity'] or '-').title(),
            str(v['cvss_score'] or '-'),
            str(v['affected_nodes'])
        ])
    story.append(create_status_table(vuln_data, [120, 80, 60, 60]))
    
    # Build PDF
    doc.build(story, onFirstPage=lambda c, d: create_header_footer(c, d, "Security Report"),
              onLaterPages=lambda c, d: create_header_footer(c, d, "Security Report"))
    
    buffer.seek(0)
    filename = f"octofleet_security_report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/api/v1/reports/inventory/pdf", dependencies=[Depends(verify_api_key)])
async def generate_inventory_report_pdf(
    node_id: Optional[str] = None,
    db: asyncpg.Pool = Depends(get_db)
):
    """E19-07: Generate Inventory Report PDF (Hardware, Software)"""
    
    # Determine scope - join nodes with v_nodes_overview for cpu_name
    if node_id:
        nodes = await db.fetch("""
            SELECT n.id, n.hostname, n.os_name, n.agent_version, n.is_online, 
                   v.cpu_name, v.ram_gb 
            FROM nodes n
            LEFT JOIN v_nodes_overview v ON n.id = v.id
            WHERE n.id::text = $1 OR n.node_id = $1
        """, node_id)
        title_suffix = f" - {nodes[0]['hostname']}" if nodes else ""
    else:
        nodes = await db.fetch("""
            SELECT n.id, n.hostname, n.os_name, n.agent_version, n.is_online, 
                   v.cpu_name, v.ram_gb 
            FROM nodes n
            LEFT JOIN v_nodes_overview v ON n.id = v.id
            ORDER BY n.hostname
        """)
        title_suffix = " - All Nodes"
    
    # Software inventory
    if node_id:
        software = await db.fetch("""
            SELECT name, version, publisher, install_date,
                   COUNT(*) as install_count
            FROM software_current
            WHERE node_id = (SELECT id FROM nodes WHERE id::text = $1 OR node_id = $1 LIMIT 1)
            GROUP BY name, version, publisher, install_date
            ORDER BY name
            LIMIT 100
        """, node_id)
    else:
        software = await db.fetch("""
            SELECT name, version, publisher,
                   COUNT(DISTINCT node_id) as install_count
            FROM software_current
            GROUP BY name, version, publisher
            ORDER BY install_count DESC, name
            LIMIT 100
        """)
    
    # Hardware summary - use v_nodes_overview
    hw_summary = await db.fetch("""
        SELECT 
            cpu_name,
            COUNT(*) as count
        FROM v_nodes_overview
        WHERE cpu_name IS NOT NULL
        GROUP BY cpu_name
        ORDER BY count DESC
        LIMIT 10
    """)
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                           rightMargin=50, leftMargin=50,
                           topMargin=60, bottomMargin=50)
    
    styles = create_pdf_styles()
    story = []
    
    # Title
    story.append(Paragraph(f"Inventory Report{title_suffix}", styles['ReportTitle']))
    story.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Node Hardware
    story.append(Paragraph("Hardware Inventory", styles['SectionTitle']))
    
    node_data = [["Hostname", "OS", "CPU", "Agent"]]
    for n in nodes:
        os_short = (n['os_name'] or '-')[:25]
        cpu_short = (n['cpu_name'] or '-')[:30]
        node_data.append([
            n['hostname'],
            os_short,
            cpu_short,
            n['agent_version'] or '-'
        ])
    story.append(create_status_table(node_data, [90, 130, 180, 50]))
    story.append(Spacer(1, 20))
    
    # CPU Distribution
    if hw_summary:
        story.append(Paragraph("CPU Distribution", styles['SectionTitle']))
        cpu_data = [["CPU Model", "Count"]]
        for hw in hw_summary:
            cpu_short = (hw['cpu_name'] or '-')[:50]
            cpu_data.append([cpu_short, str(hw['count'])])
        story.append(create_status_table(cpu_data, [350, 60]))
        story.append(Spacer(1, 20))
    
    # Software Inventory
    story.append(Paragraph("Software Inventory", styles['SectionTitle']))
    
    sw_data = [["Software", "Version", "Publisher", "Installs"]]
    for s in software[:50]:
        sw_data.append([
            (s['name'] or '-')[:35],
            (s['version'] or '-')[:15],
            (s['publisher'] or '-')[:20],
            str(s['install_count'])
        ])
    story.append(create_status_table(sw_data, [180, 80, 120, 50]))
    
    # Build PDF
    doc.build(story, onFirstPage=lambda c, d: create_header_footer(c, d, "Inventory Report"),
              onLaterPages=lambda c, d: create_header_footer(c, d, "Inventory Report"))
    
    buffer.seek(0)
    filename = f"octofleet_inventory_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )



# ============================================================
# E35 - Enterprise Reporting Suite Endpoints
# ============================================================

import json
import os
import smtplib
from email.message import EmailMessage
from typing import Any
from typing import Optional as Opt

from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
from fastapi import HTTPException
from fastapi.responses import FileResponse
from pydantic import AliasChoices, BaseModel, Field

from dependencies import get_pool


class ExecuteReportRequest(BaseModel):
    report_id: str = Field(validation_alias=AliasChoices("report_id", "reportId", "id", "slug"))
    parameters: dict[str, Any] = Field(default_factory=dict)
    output_format: str = Field(default="csv", validation_alias=AliasChoices("output_format", "outputFormat", "format"))


class CreateScheduleRequest(BaseModel):
    report_id: str = Field(validation_alias=AliasChoices("report_id", "reportId", "slug", "id"))
    name: str
    cron_expression: str = Field(validation_alias=AliasChoices("cron_expression", "cronExpression"))
    parameters: dict[str, Any] = Field(default_factory=dict)
    output_format: str = Field(default="pdf", validation_alias=AliasChoices("output_format", "outputFormat"))
    delivery_method: str = Field(default="download", validation_alias=AliasChoices("delivery_method", "deliveryMethod"))
    delivery_config: dict[str, Any] = Field(default_factory=dict, validation_alias=AliasChoices("delivery_config", "deliveryConfig"))
    enabled: bool = True


class UpdateScheduleRequest(BaseModel):
    name: Opt[str] = None
    cron_expression: Opt[str] = Field(default=None, validation_alias=AliasChoices("cron_expression", "cronExpression"))
    parameters: Opt[dict[str, Any]] = None
    output_format: Opt[str] = Field(default=None, validation_alias=AliasChoices("output_format", "outputFormat"))
    delivery_method: Opt[str] = Field(default=None, validation_alias=AliasChoices("delivery_method", "deliveryMethod"))
    delivery_config: Opt[dict[str, Any]] = Field(default=None, validation_alias=AliasChoices("delivery_config", "deliveryConfig"))
    enabled: Opt[bool] = None


REPORT_DIR = "/tmp/octofleet-reports"
report_scheduler: AsyncIOScheduler | None = None


BUILTIN_REPORTS = [
    {
        "id": "patch-compliance",
        "name": "Patch Compliance Overview",
        "slug": "patch-compliance",
        "description": "Patch compliance per node including update availability.",
        "category": "Compliance",
        "query_template": """
            SELECT n.hostname,
                   COALESCE(COUNT(u.id), 0) AS updates_available,
                   COALESCE(SUM(CASE WHEN u.severity = 'critical' THEN 1 ELSE 0 END), 0) AS critical_updates,
                   CASE WHEN COUNT(u.id) = 0 THEN 'up_to_date' ELSE 'updates_pending' END AS patch_state
            FROM nodes n
            LEFT JOIN node_available_updates u ON u.node_id = n.id
            GROUP BY n.id, n.hostname
            ORDER BY critical_updates DESC, updates_available DESC, n.hostname
        """,
    },
    {
        "id": "user-activity",
        "name": "User Activity Audit",
        "slug": "user-activity",
        "description": "Login history and key user actions from audit_log.",
        "category": "Security",
        "query_template": """
            SELECT created_at, user_id, action, resource_type, resource_id, ip_address
            FROM audit_log
            ORDER BY created_at DESC
            LIMIT 1000
        """,
    },
    {
        "id": "hardware-inventory",
        "name": "Hardware Inventory",
        "slug": "hardware-inventory",
        "description": "Full hardware specs per node.",
        "category": "Fleet",
        "query_template": """
            SELECT n.hostname,
                   h.cpu->>'Name' AS cpu,
                   h.cpu->>'Cores' AS cpu_cores,
                   h.ram->>'TotalGB' AS ram_gb,
                   h.mainboard->>'Manufacturer' AS manufacturer,
                   h.mainboard->>'Product' AS model,
                   h.bios->>'SerialNumber' AS serial_number,
                   h.collected_at
            FROM nodes n
            LEFT JOIN hardware_current h ON h.node_id = n.id
            ORDER BY n.hostname
        """,
    },
    {
        "id": "service-health",
        "name": "Service Health",
        "slug": "service-health",
        "description": "Service status across fleet based on node health telemetry.",
        "category": "Operations",
        "query_template": """
            SELECT hostname, is_online, agent_version, health_status, last_seen
            FROM v_nodes_overview
            ORDER BY is_online DESC, hostname
        """,
    },
    {
        "id": "license-usage",
        "name": "License Usage",
        "slug": "license-usage",
        "description": "Software metering and license usage summary.",
        "category": "Executive",
        "query_template": """
            SELECT name AS software_name,
                   publisher,
                   COUNT(DISTINCT node_id) AS installed_nodes,
                   MIN(version) AS lowest_version,
                   MAX(version) AS highest_version
            FROM software_current
            GROUP BY name, publisher
            ORDER BY installed_nodes DESC, software_name
            LIMIT 500
        """,
    },
]


async def _ensure_builtin_reports(pool: asyncpg.Pool):
    async with pool.acquire() as conn:
        for report in BUILTIN_REPORTS:
            await conn.execute(
                """
                INSERT INTO report_catalog (id, name, slug, description, category, query_template, output_formats, is_builtin)
                VALUES ($1, $2, $3, $4, $5, $6, ARRAY['pdf','csv'], true)
                ON CONFLICT (slug) DO UPDATE
                SET name = EXCLUDED.name,
                    description = EXCLUDED.description,
                    category = EXCLUDED.category,
                    query_template = EXCLUDED.query_template,
                    updated_at = NOW()
                """,
                report["id"], report["name"], report["slug"], report["description"], report["category"], report["query_template"],
            )


def _schedule_to_job_id(schedule_id: str) -> str:
    return f"report-schedule:{schedule_id}"


async def _render_report_file(db: asyncpg.Pool, report: dict, exec_id: str, fmt: str):
    os.makedirs(REPORT_DIR, exist_ok=True)
    file_path = os.path.join(REPORT_DIR, f"{exec_id}.{fmt}")
    query = report["query_template"]
    if query:
        try:
            rows = await db.fetch(query)
        except Exception as e:
            rows = [{"warning": f"Report query fallback: {e}"}]
    else:
        rows = []

    if fmt == "csv":
        import csv as csv_mod
        columns = list(rows[0].keys()) if rows else []
        with open(file_path, "w", newline="", encoding="utf-8") as f:
            writer = csv_mod.writer(f)
            writer.writerow(columns)
            for r in rows:
                writer.writerow([r[c] for c in columns])
    else:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = create_pdf_styles()
        story = [
            Paragraph(report["name"], styles["ReportTitle"]),
            Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", styles["Normal"]),
            Spacer(1, 20),
        ]
        if rows:
            columns = list(rows[0].keys())
            table_data = [columns] + [[str(r[c]) for c in columns] for r in rows[:500]]
            story.append(create_status_table(table_data))
        else:
            story.append(Paragraph("No data available.", styles["Normal"]))
        doc.build(
            story,
            onFirstPage=lambda c, d: create_header_footer(c, d, report["name"]),
            onLaterPages=lambda c, d: create_header_footer(c, d, report["name"]),
        )
        with open(file_path, "wb") as f:
            f.write(buffer.getvalue())

    return file_path, rows


async def _execute_report_internal(
    db: asyncpg.Pool,
    report_identifier: str,
    output_format: str,
    parameters: dict[str, Any] | None = None,
    schedule_id: str | None = None,
):
    report = await db.fetchrow("SELECT * FROM report_catalog WHERE id = $1", report_identifier)
    if not report:
        report = await db.fetchrow("SELECT * FROM report_catalog WHERE slug = $1", report_identifier)
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")

    fmt = output_format if output_format in ("csv", "pdf") else "csv"
    params = parameters or {}

    exec_id = await db.fetchval(
        """INSERT INTO report_executions (report_id, schedule_id, status, parameters, output_format, started_at)
           VALUES ($1, $2, 'running', $3, $4, NOW()) RETURNING id""",
        report["id"], schedule_id, json.dumps(params), fmt,
    )

    try:
        file_path, rows = await _render_report_file(db, dict(report), exec_id, fmt)
        file_size = os.path.getsize(file_path)
        await db.execute(
            """UPDATE report_executions
               SET status='completed', file_path=$1, file_size_bytes=$2, completed_at=NOW()
               WHERE id=$3""",
            file_path, file_size, exec_id,
        )
        return {
            "id": exec_id,
            "status": "completed",
            "file_path": file_path,
            "file_size_bytes": file_size,
            "report": dict(report),
            "rows": rows,
            "output_format": fmt,
        }
    except Exception as e:
        await db.execute(
            "UPDATE report_executions SET status='failed', error_message=$1, completed_at=NOW() WHERE id=$2",
            str(e),
            exec_id,
        )
        return {"id": exec_id, "status": "failed", "error": str(e), "report": dict(report), "output_format": fmt}


def _build_email_html(report_name: str, execution_id: str, rows: list[Any]) -> str:
    preview_rows = rows[:10] if rows else []
    if preview_rows:
        headers = list(preview_rows[0].keys())
        head_html = "".join([f"<th style='padding:8px;border:1px solid #ddd'>{h}</th>" for h in headers])
        body_html = ""
        for r in preview_rows:
            body_html += "<tr>" + "".join([f"<td style='padding:8px;border:1px solid #ddd'>{r[h]}</td>" for h in headers]) + "</tr>"
        table_html = f"<table style='border-collapse:collapse;width:100%'><thead><tr>{head_html}</tr></thead><tbody>{body_html}</tbody></table>"
    else:
        table_html = "<p>No data rows returned.</p>"

    return f"""
        <h2>OctoFleet Scheduled Report</h2>
        <p><strong>Report:</strong> {report_name}</p>
        <p><strong>Execution ID:</strong> {execution_id}</p>
        <p>Attached is the generated report file. Preview:</p>
        {table_html}
    """


async def _send_email(schedule: dict, execution: dict):
    delivery_config = schedule.get("delivery_config") or {}
    recipient = delivery_config.get("email") or delivery_config.get("to")
    if not recipient:
        raise ValueError("delivery_config.email is required for email delivery")

    smtp_host = os.getenv("SMTP_HOST")
    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    smtp_user = os.getenv("SMTP_USER")
    smtp_password = os.getenv("SMTP_PASSWORD")
    smtp_from = os.getenv("SMTP_FROM") or smtp_user

    if not smtp_host or not smtp_from:
        raise ValueError("SMTP_HOST and SMTP_FROM (or SMTP_USER) must be configured")

    msg = EmailMessage()
    msg["Subject"] = f"OctoFleet Report: {execution['report']['name']}"
    msg["From"] = smtp_from
    msg["To"] = recipient
    msg.set_content("This email contains an HTML report summary and attachment.")
    msg.add_alternative(_build_email_html(execution["report"]["name"], execution["id"], execution.get("rows", [])), subtype="html")

    with open(execution["file_path"], "rb") as f:
        content = f.read()
    mime = "text/csv" if execution["output_format"] == "csv" else "application/pdf"
    maintype, subtype = mime.split("/", 1)
    msg.add_attachment(content, maintype=maintype, subtype=subtype, filename=os.path.basename(execution["file_path"]))

    with smtplib.SMTP(smtp_host, smtp_port, timeout=20) as smtp:
        if smtp_user and smtp_password:
            smtp.starttls()
            smtp.login(smtp_user, smtp_password)
        smtp.send_message(msg)


async def _record_delivery(db: asyncpg.Pool, schedule_id: str, execution_id: str, status: str, error: str | None = None):
    await db.execute(
        """
        INSERT INTO report_deliveries (schedule_id, execution_id, delivered_at, delivery_status, error_message)
        VALUES ($1, $2, NOW(), $3, $4)
        """,
        schedule_id,
        execution_id,
        status,
        error,
    )


async def run_schedule_job(schedule_id: str):
    pool = get_pool()
    if not pool:
        return
    async with pool.acquire() as conn:
        schedule = await conn.fetchrow("SELECT * FROM report_schedules WHERE id = $1 AND enabled = true", schedule_id)
        if not schedule:
            return

        execution = await _execute_report_internal(
            pool,
            report_identifier=schedule["report_id"],
            output_format=schedule["output_format"],
            parameters=schedule["parameters"] or {},
            schedule_id=schedule["id"],
        )
        await conn.execute("UPDATE report_schedules SET last_run_at = NOW() WHERE id = $1", schedule_id)

        if execution["status"] != "completed":
            await _record_delivery(pool, schedule_id, execution["id"], "failed", execution.get("error", "Execution failed"))
            return

        if schedule["delivery_method"] == "email":
            try:
                await _send_email(dict(schedule), execution)
                await _record_delivery(pool, schedule_id, execution["id"], "delivered")
            except Exception as e:
                await _record_delivery(pool, schedule_id, execution["id"], "failed", str(e))
        else:
            await _record_delivery(pool, schedule_id, execution["id"], "skipped", "Delivery method is not email")


async def _register_schedule_job(schedule: dict):
    global report_scheduler
    if not report_scheduler:
        return
    job_id = _schedule_to_job_id(schedule["id"])
    if report_scheduler.get_job(job_id):
        report_scheduler.remove_job(job_id)
    if not schedule.get("enabled"):
        return
    trigger = CronTrigger.from_crontab(schedule["cron_expression"])
    report_scheduler.add_job(run_schedule_job, trigger=trigger, id=job_id, kwargs={"schedule_id": schedule["id"]}, replace_existing=True)


@router.on_event("startup")
async def reports_startup():
    global report_scheduler
    pool = get_pool()
    if not pool:
        return

    async with pool.acquire() as conn:
        await conn.execute(
            """
            CREATE TABLE IF NOT EXISTS report_deliveries (
                id TEXT PRIMARY KEY DEFAULT gen_random_uuid()::text,
                schedule_id TEXT REFERENCES report_schedules(id) ON DELETE CASCADE,
                execution_id TEXT REFERENCES report_executions(id) ON DELETE CASCADE,
                delivered_at TIMESTAMPTZ DEFAULT NOW(),
                delivery_status TEXT NOT NULL,
                error_message TEXT
            )
            """
        )
    await _ensure_builtin_reports(pool)

    report_scheduler = AsyncIOScheduler(timezone="UTC")
    async with pool.acquire() as conn:
        schedules = await conn.fetch("SELECT * FROM report_schedules WHERE enabled = true")
    for s in schedules:
        try:
            await _register_schedule_job(dict(s))
        except Exception:
            continue
    report_scheduler.start()


@router.on_event("shutdown")
async def reports_shutdown():
    global report_scheduler
    if report_scheduler and report_scheduler.running:
        report_scheduler.shutdown(wait=False)


# --- Report Catalog ---

@router.get("/api/v1/reports/catalog", dependencies=[Depends(verify_api_key)])
async def list_report_catalog(category: Opt[str] = None, db: asyncpg.Pool = Depends(get_db)):
    if category and category.lower() != "all":
        rows = await db.fetch("SELECT * FROM report_catalog WHERE category = $1 ORDER BY name", category)
    else:
        rows = await db.fetch("SELECT * FROM report_catalog ORDER BY category, name")
    return [dict(r) for r in rows]


@router.get("/api/v1/reports/catalog/{slug}", dependencies=[Depends(verify_api_key)])
async def get_report_catalog(slug: str, db: asyncpg.Pool = Depends(get_db)):
    row = await db.fetchrow("SELECT * FROM report_catalog WHERE slug = $1", slug)
    if not row:
        raise HTTPException(status_code=404, detail="Report not found")
    return dict(row)


# --- Report Execution ---

@router.post("/api/v1/reports/execute", dependencies=[Depends(verify_api_key)])
async def execute_report(req: ExecuteReportRequest, db: asyncpg.Pool = Depends(get_db)):
    result = await _execute_report_internal(db, req.report_id, req.output_format, req.parameters)
    if result["status"] != "completed":
        return {"id": result["id"], "status": "failed", "error": result.get("error", "Execution failed")}
    return {
        "id": result["id"],
        "status": "completed",
        "file_path": result["file_path"],
        "file_size_bytes": result["file_size_bytes"],
    }


@router.get("/api/v1/reports/executions", dependencies=[Depends(verify_api_key)])
async def list_executions(
    report_id: Opt[str] = None,
    status: Opt[str] = None,
    limit: int = 50,
    offset: int = 0,
    db: asyncpg.Pool = Depends(get_db),
):
    conditions = []
    args = []
    idx = 1
    if report_id:
        conditions.append(f"re.report_id = ${idx}")
        args.append(report_id)
        idx += 1
    if status:
        conditions.append(f"re.status = ${idx}")
        args.append(status)
        idx += 1
    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
    rows = await db.fetch(
        f"""
        SELECT re.*, rc.name AS report_name, rc.slug AS report_slug
        FROM report_executions re
        LEFT JOIN report_catalog rc ON rc.id = re.report_id
        {where}
        ORDER BY re.created_at DESC
        LIMIT ${idx} OFFSET ${idx + 1}
        """,
        *args,
        limit,
        offset,
    )
    return [dict(r) for r in rows]


@router.get("/api/v1/reports/executions/{execution_id}", dependencies=[Depends(verify_api_key)])
async def get_execution(execution_id: str, db: asyncpg.Pool = Depends(get_db)):
    row = await db.fetchrow("SELECT * FROM report_executions WHERE id = $1", execution_id)
    if not row:
        raise HTTPException(status_code=404, detail="Execution not found")
    result = dict(row)
    if result.get("file_path") and result["status"] == "completed":
        result["download_url"] = f"/api/v1/reports/executions/{execution_id}/download"
    return result


@router.get("/api/v1/reports/executions/{execution_id}/download", dependencies=[Depends(verify_api_key)])
async def download_execution(execution_id: str, db: asyncpg.Pool = Depends(get_db)):
    row = await db.fetchrow("SELECT file_path, output_format FROM report_executions WHERE id = $1 AND status = 'completed'", execution_id)
    if not row or not row["file_path"] or not os.path.exists(row["file_path"]):
        raise HTTPException(status_code=404, detail="File not found")
    media = "text/csv" if row["output_format"] == "csv" else "application/pdf"
    return FileResponse(row["file_path"], media_type=media, filename=os.path.basename(row["file_path"]))


# --- Schedules ---

@router.get("/api/v1/reports/schedules", dependencies=[Depends(verify_api_key)])
async def list_schedules(db: asyncpg.Pool = Depends(get_db)):
    rows = await db.fetch(
        """SELECT s.*, rc.name as report_name, rc.slug as report_slug
           FROM report_schedules s JOIN report_catalog rc ON s.report_id = rc.id ORDER BY s.created_at DESC"""
    )
    return [dict(r) for r in rows]


@router.post("/api/v1/reports/schedules", dependencies=[Depends(verify_api_key)])
async def create_schedule(req: CreateScheduleRequest, db: asyncpg.Pool = Depends(get_db)):
    report = await db.fetchrow("SELECT id FROM report_catalog WHERE id = $1 OR slug = $1", req.report_id)
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    row = await db.fetchrow(
        """INSERT INTO report_schedules (report_id, name, cron_expression, parameters, output_format, delivery_method, delivery_config, enabled)
           VALUES ($1, $2, $3, $4, $5, $6, $7, $8) RETURNING *""",
        report["id"], req.name, req.cron_expression, json.dumps(req.parameters),
        req.output_format, req.delivery_method, json.dumps(req.delivery_config), req.enabled,
    )
    await _register_schedule_job(dict(row))
    return dict(row)


@router.put("/api/v1/reports/schedules/{schedule_id}", dependencies=[Depends(verify_api_key)])
async def update_schedule(schedule_id: str, req: UpdateScheduleRequest, db: asyncpg.Pool = Depends(get_db)):
    existing = await db.fetchrow("SELECT * FROM report_schedules WHERE id = $1", schedule_id)
    if not existing:
        raise HTTPException(status_code=404, detail="Schedule not found")

    updates = {}
    if req.name is not None:
        updates["name"] = req.name
    if req.cron_expression is not None:
        updates["cron_expression"] = req.cron_expression
    if req.parameters is not None:
        updates["parameters"] = json.dumps(req.parameters)
    if req.output_format is not None:
        updates["output_format"] = req.output_format
    if req.delivery_method is not None:
        updates["delivery_method"] = req.delivery_method
    if req.delivery_config is not None:
        updates["delivery_config"] = json.dumps(req.delivery_config)
    if req.enabled is not None:
        updates["enabled"] = req.enabled

    if updates:
        set_parts = [f"{k} = ${i + 2}" for i, k in enumerate(updates.keys())]
        await db.execute(f"UPDATE report_schedules SET {', '.join(set_parts)} WHERE id = $1", schedule_id, *updates.values())
    row = await db.fetchrow("SELECT * FROM report_schedules WHERE id = $1", schedule_id)
    await _register_schedule_job(dict(row))
    return dict(row)


@router.post("/api/v1/reports/schedules/{schedule_id}/run-now", dependencies=[Depends(verify_api_key)])
async def run_schedule_now(schedule_id: str, db: asyncpg.Pool = Depends(get_db)):
    schedule = await db.fetchrow("SELECT * FROM report_schedules WHERE id = $1", schedule_id)
    if not schedule:
        raise HTTPException(status_code=404, detail="Schedule not found")
    await run_schedule_job(schedule_id)
    return {"ok": True, "schedule_id": schedule_id}


@router.get("/api/v1/reports/schedules/{schedule_id}/deliveries", dependencies=[Depends(verify_api_key)])
async def list_schedule_deliveries(schedule_id: str, db: asyncpg.Pool = Depends(get_db)):
    rows = await db.fetch(
        """
        SELECT d.*, re.status AS execution_status, re.output_format, re.completed_at
        FROM report_deliveries d
        LEFT JOIN report_executions re ON re.id = d.execution_id
        WHERE d.schedule_id = $1
        ORDER BY d.delivered_at DESC
        LIMIT 200
        """,
        schedule_id,
    )
    return [dict(r) for r in rows]


@router.delete("/api/v1/reports/schedules/{schedule_id}", dependencies=[Depends(verify_api_key)])
async def delete_schedule(schedule_id: str, db: asyncpg.Pool = Depends(get_db)):
    global report_scheduler
    if report_scheduler and report_scheduler.get_job(_schedule_to_job_id(schedule_id)):
        report_scheduler.remove_job(_schedule_to_job_id(schedule_id))
    result = await db.execute("DELETE FROM report_schedules WHERE id = $1", schedule_id)
    return {"deleted": "DELETE 1" in result}


@router.get("/api/v1/reports/executive-dashboard", dependencies=[Depends(verify_api_key)])
async def executive_dashboard(db: asyncpg.Pool = Depends(get_db)):
    kpis = await db.fetchrow(
        """
        WITH vuln AS (
            SELECT
                SUM(CASE WHEN severity = 'critical' THEN 1 ELSE 0 END) AS critical_vulns,
                SUM(CASE WHEN severity IN ('critical','high','medium','low') THEN 1 ELSE 0 END) AS open_findings
            FROM vulnerabilities
        ),
        patch AS (
            SELECT
                COUNT(*) AS total_nodes,
                SUM(CASE WHEN COALESCE(u.update_count, 0) = 0 THEN 1 ELSE 0 END) AS up_to_date
            FROM nodes n
            LEFT JOIN (
                SELECT node_id, COUNT(*) AS update_count
                FROM node_available_updates
                GROUP BY node_id
            ) u ON u.node_id = n.id
        )
        SELECT
            (SELECT COUNT(*) FROM nodes) AS total_nodes,
            (SELECT COUNT(*) FROM nodes WHERE is_online = true) AS online_nodes,
            COALESCE((SELECT critical_vulns FROM vuln), 0) AS critical_vulns,
            COALESCE((SELECT open_findings FROM vuln), 0) AS open_findings,
            0::int AS active_alerts,
            CASE WHEN (SELECT total_nodes FROM patch) = 0
              THEN 100
              ELSE ROUND(((SELECT up_to_date FROM patch)::numeric / (SELECT total_nodes FROM patch)::numeric) * 100, 2)
            END AS patch_compliance_pct
        """
    )

    trend = await db.fetch(
        """
        WITH dates AS (
            SELECT generate_series((CURRENT_DATE - INTERVAL '6 day')::date, CURRENT_DATE::date, INTERVAL '1 day')::date AS day
        )
        SELECT d.day::text AS date,
               (SELECT COUNT(*) FROM nodes WHERE is_online = true) AS online,
               0::int AS alerts
        FROM dates d
        ORDER BY d.day
        """
    )

    top_vulnerabilities = await db.fetch(
        """
        SELECT v.cve_id AS cve,
               v.severity,
               COUNT(DISTINCT s.node_id) AS affected_nodes
        FROM vulnerabilities v
        LEFT JOIN software_current s
          ON s.name ILIKE v.software_name AND s.version = v.software_version
        GROUP BY v.cve_id, v.severity
        ORDER BY affected_nodes DESC, v.cve_id
        LIMIT 5
        """
    )

    patch_status = await db.fetchrow(
        """
        WITH updates AS (
            SELECT n.id,
                   COALESCE(COUNT(u.id), 0) AS total_updates,
                   COALESCE(SUM(CASE WHEN u.severity = 'critical' THEN 1 ELSE 0 END), 0) AS critical_updates
            FROM nodes n
            LEFT JOIN node_available_updates u ON u.node_id = n.id
            GROUP BY n.id
        )
        SELECT
            SUM(CASE WHEN total_updates = 0 THEN 1 ELSE 0 END) AS up_to_date,
            SUM(CASE WHEN total_updates > 0 THEN 1 ELSE 0 END) AS updates_available,
            SUM(CASE WHEN critical_updates > 0 THEN 1 ELSE 0 END) AS critical_updates
        FROM updates
        """
    )

    recent_jobs = await db.fetch(
        """
        SELECT j.name,
               COALESCE((SELECT status FROM job_instances ji WHERE ji.job_id = j.id ORDER BY created_at DESC LIMIT 1), 'pending') AS status,
               COALESCE((SELECT COUNT(*) FROM job_instances ji2 WHERE ji2.job_id = j.id), 0) AS node_count,
               j.created_at
        FROM jobs j
        ORDER BY j.created_at DESC
        LIMIT 10
        """
    )

    return {
        "kpis": dict(kpis) if kpis else {},
        "trend_7d": [dict(r) for r in trend],
        "top_vulnerabilities": [dict(r) for r in top_vulnerabilities],
        "patch_status": dict(patch_status) if patch_status else {"up_to_date": 0, "updates_available": 0, "critical_updates": 0},
        "recent_jobs": [dict(r) for r in recent_jobs],
    }


# --- Stats ---

@router.get("/api/v1/reports/stats", dependencies=[Depends(verify_api_key)])
async def report_stats(db: asyncpg.Pool = Depends(get_db)):
    stats = await db.fetchrow(
        """
        SELECT
            (SELECT COUNT(*) FROM report_catalog) as total_reports,
            (SELECT COUNT(*) FROM report_schedules) as total_schedules,
            (SELECT COUNT(*) FROM report_executions WHERE created_at > NOW() - INTERVAL '7 days') as executions_this_week,
            (SELECT ROUND(AVG(EXTRACT(EPOCH FROM (completed_at - started_at)))::numeric, 2)
             FROM report_executions WHERE status = 'completed') as avg_duration_sec
    """
    )
    return {
        "total_reports": stats["total_reports"],
        "total_schedules": stats["total_schedules"],
        "executions_this_week": stats["executions_this_week"],
        "avg_duration_sec": float(stats["avg_duration_sec"]) if stats["avg_duration_sec"] else 0,
    }
